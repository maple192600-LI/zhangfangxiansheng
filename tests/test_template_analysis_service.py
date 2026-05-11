"""Tests for template analysis and field dictionary services.

Verifies:
- Services do not import legacy agents.fund
- Template structure parsing works with real Excel files
- Field mapping matches placeholders to canonical fields
- Field dictionary loads defaults when no seed file exists
- Artifact service enhancements work correctly
"""
import sys
from pathlib import Path

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "backend"))
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from database import Base
from db.schemas import ArtifactKind, ParserArtifactDraftCreate, RuleArtifactDraftCreate
from services import artifact_service, template_analysis
from services.field_dictionary_service import (
    get_alias_library,
    get_field_dictionary,
    normalize_field_name,
)


@pytest.fixture()
def db_session():
    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    session = sessionmaker(bind=engine)()
    try:
        yield session
    finally:
        session.close()
        engine.dispose()


# ── Import isolation ──

def test_template_analysis_no_fund_import():
    src = Path(template_analysis.__file__).read_text(encoding="utf-8")
    assert "agents.fund" not in src
    assert "FundAgent" not in src
    assert "fund_skill_run" not in src
    assert "harness" not in src


def test_field_dictionary_no_fund_import():
    from services import field_dictionary_service
    src = Path(field_dictionary_service.__file__).read_text(encoding="utf-8")
    assert "agents.fund" not in src
    assert "FundAgent" not in src


def test_artifact_service_no_fund_import():
    src = Path(artifact_service.__file__).read_text(encoding="utf-8")
    assert "agents.fund" not in src
    assert "FundAgent" not in src
    assert "fund_skill_run" not in src


# ── Template analysis ──

def test_parse_template_structure_with_excel(tmp_path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "报表标题"
    ws["B2"] = "${月初余额}"
    ws["C3"] = "{{收入}}"
    ws["D4"] = "【支出】"
    ws["E5"] = "plain text"
    file_path = str(tmp_path / "test_template.xlsx")
    wb.save(file_path)

    result = template_analysis.parse_template_structure(file_path)
    assert "月初余额" in result["placeholders"]
    assert "收入" in result["placeholders"]
    assert "支出" in result["placeholders"]
    assert "plain text" not in result["placeholders"]


def test_parse_template_structure_missing_file():
    result = template_analysis.parse_template_structure("/nonexistent/file.xlsx")
    assert result["placeholders"] == []
    assert result["merged_cells"] == []


def test_map_placeholders_to_fields():
    stage_a = {
        "placeholders": ["月初余额", "收入", "支出", "报表标题"],
        "merged_cells": [],
        "header_rows": [],
    }
    result = template_analysis.map_placeholders_to_fields(stage_a)

    bindings = result["placeholder_bindings"]
    assert bindings["月初余额"]["primitive"] == "field"
    assert bindings["月初余额"]["value"] == "rolling_balance"
    assert bindings["收入"]["primitive"] == "field"
    assert bindings["收入"]["value"] == "amount_in"
    assert bindings["支出"]["primitive"] == "field"
    assert bindings["支出"]["value"] == "amount_out"
    assert bindings["报表标题"]["primitive"] == "const"
    assert result["matched_count"] == 3
    assert result["total_placeholders"] == 4
    assert result["confidence"] > 0.5


def test_map_placeholders_empty():
    stage_a = {"placeholders": [], "merged_cells": [], "header_rows": []}
    result = template_analysis.map_placeholders_to_fields(stage_a)
    assert result["confidence"] == 0.0
    assert result["matched_count"] == 0


def test_map_placeholders_with_custom_dictionary():
    stage_a = {"placeholders": ["自定义字段"], "merged_cells": [], "header_rows": []}
    custom_dict = {"custom_field": {"cn_name": "自定义字段", "aliases": []}}
    result = template_analysis.map_placeholders_to_fields(stage_a, field_dictionary=custom_dict)
    bindings = result["placeholder_bindings"]
    assert bindings["自定义字段"]["value"] == "custom_field"


# ── Field dictionary service ──

def test_get_field_dictionary_returns_defaults():
    fd = get_field_dictionary()
    assert "business_date" in fd
    assert "amount_in" in fd
    assert "amount_out" in fd
    assert fd["business_date"]["cn_name"] == "日期"


def test_normalize_field_name_exact():
    assert normalize_field_name("日期") == "business_date"
    assert normalize_field_name("交易日期") == "business_date"
    assert normalize_field_name("收入") == "amount_in"
    assert normalize_field_name("摘要") == "summary"


def test_normalize_field_name_no_match():
    assert normalize_field_name("不存在的字段") is None


def test_get_alias_library():
    al = get_alias_library()
    assert isinstance(al, dict)


# ── Artifact service enhancements ──

def test_get_active_parser_none_when_empty(db_session):
    result = artifact_service.get_active_parser(db_session, "A001", "bank")
    assert result is None


def test_get_active_parser_returns_approved(db_session):
    data = ParserArtifactDraftCreate(
        name="ICBC_v1", kind=ArtifactKind.bank, account_code="A001",
        code="#", primitives_imports=[],
    )
    draft = artifact_service.create_parser_draft(db_session, data)
    artifact_service.approve_parser_artifact(db_session, draft["id"], "admin")

    active = artifact_service.get_active_parser(db_session, "A001", "bank")
    assert active is not None
    assert active["status"] == "active"
    assert active["id"] == draft["id"]


def test_list_parser_versions(db_session):
    data_v1 = ParserArtifactDraftCreate(
        name="ICBC_v1", kind=ArtifactKind.bank, account_code="A001",
        code="# v1", primitives_imports=[],
    )
    data_v2 = ParserArtifactDraftCreate(
        name="ICBC_v1", kind=ArtifactKind.bank, account_code="A001",
        code="# v2", primitives_imports=[],
    )
    artifact_service.create_parser_draft(db_session, data_v1)
    artifact_service.create_parser_draft(db_session, data_v2)

    versions = artifact_service.list_parser_versions(db_session, "ICBC_v1")
    assert len(versions) == 2
    assert versions[0]["version"] == 2
    assert versions[1]["version"] == 1


def test_list_rule_versions(db_session):
    data = RuleArtifactDraftCreate(
        name="TestRule", placeholder_bindings={}, primitives_imports=[],
    )
    artifact_service.create_rule_draft(db_session, data)

    versions = artifact_service.list_rule_versions(db_session, "TestRule")
    assert len(versions) == 1
    assert versions[0]["version"] == 1


def test_list_rule_versions_empty(db_session):
    versions = artifact_service.list_rule_versions(db_session, "NonExistent")
    assert versions == []
