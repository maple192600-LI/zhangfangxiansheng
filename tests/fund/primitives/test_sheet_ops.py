"""test_sheet_ops · §P1 · 6 函数单测

read_sheet / detect_header_row / extract_headers / iter_data_rows /
is_empty_row / locate_merged_cells
"""
from __future__ import annotations

import pytest
from openpyxl import Workbook

from fund.primitives.sheet_ops import (
    detect_header_row,
    extract_headers,
    is_empty_row,
    iter_data_rows,
    locate_merged_cells,
    read_sheet,
)


# ──────────────────────────────────────────
# helpers
# ──────────────────────────────────────────
def _make_wb():
    wb = Workbook()
    ws = wb.active
    ws.title = "流水"
    # 两行 junk
    ws["A1"] = "工商银行流水明细"
    ws["A2"] = "打印日期：2026-04-23"
    # 表头行 @ row 3
    ws.cell(row=3, column=1, value="日期")
    ws.cell(row=3, column=2, value="摘要")
    ws.cell(row=3, column=3, value="对方")
    ws.cell(row=3, column=4, value="收入")
    ws.cell(row=3, column=5, value="支出")
    ws.cell(row=3, column=6, value="余额")
    # 数据行
    ws.cell(row=4, column=1, value="2026-04-22")
    ws.cell(row=4, column=2, value="转账")
    ws.cell(row=4, column=3, value="深圳XX科技")
    ws.cell(row=4, column=4, value=100.50)
    ws.cell(row=4, column=5, value=None)
    ws.cell(row=4, column=6, value=1000.50)
    # 空行
    # row 5 全空
    # 下一条
    ws.cell(row=6, column=1, value="2026-04-23")
    ws.cell(row=6, column=2, value="工资")
    ws.cell(row=6, column=4, value=5000)
    ws.cell(row=6, column=6, value=6000.50)
    return wb


# ──────────────────────────────────────────
# read_sheet
# ──────────────────────────────────────────
def test_read_sheet_by_index():
    wb = _make_wb()
    ws = read_sheet(wb, index=0)
    assert ws.title == "流水"


def test_read_sheet_by_name():
    wb = _make_wb()
    wb.create_sheet("总账")
    ws = read_sheet(wb, name="总账")
    assert ws.title == "总账"


def test_read_sheet_name_priority_over_index():
    wb = _make_wb()
    wb.create_sheet("总账")
    # index 0 = 流水, name=总账 应胜出
    ws = read_sheet(wb, index=0, name="总账")
    assert ws.title == "总账"


def test_read_sheet_missing_name_raises():
    wb = _make_wb()
    with pytest.raises(KeyError):
        read_sheet(wb, name="不存在")


def test_read_sheet_index_out_of_range_raises():
    wb = _make_wb()
    with pytest.raises(IndexError):
        read_sheet(wb, index=99)


def test_read_sheet_negative_index_raises():
    wb = _make_wb()
    with pytest.raises(IndexError):
        read_sheet(wb, index=-1)


# ──────────────────────────────────────────
# detect_header_row
# ──────────────────────────────────────────
def test_detect_header_row_finds_max_filled():
    wb = _make_wb()
    ws = read_sheet(wb)
    assert detect_header_row(ws, max_scan=10) == 3


def test_detect_header_row_empty_sheet_returns_1():
    wb = Workbook()
    ws = wb.active
    assert detect_header_row(ws, max_scan=5) == 1


def test_detect_header_row_ignores_whitespace_cells():
    wb = Workbook()
    ws = wb.active
    # row 1: 1 填充
    ws.cell(row=1, column=1, value="标题")
    # row 2: 3 个有实质内容
    ws.cell(row=2, column=1, value="日期")
    ws.cell(row=2, column=2, value="   ")
    ws.cell(row=2, column=3, value="摘要")
    ws.cell(row=2, column=4, value="金额")
    # row 2 有效填充 = 3（跳过全空白单元格），row 1 = 1
    assert detect_header_row(ws, max_scan=5) == 2


def test_detect_header_row_zero_scan_returns_1():
    wb = _make_wb()
    ws = read_sheet(wb)
    assert detect_header_row(ws, max_scan=0) == 1


# ──────────────────────────────────────────
# extract_headers
# ──────────────────────────────────────────
def test_extract_headers_basic():
    wb = _make_wb()
    ws = read_sheet(wb)
    headers = extract_headers(ws, header_row=3)
    assert headers == {
        "日期": 1, "摘要": 2, "对方": 3, "收入": 4, "支出": 5, "余额": 6,
    }


def test_extract_headers_strips_whitespace():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="  日期  ")
    ws.cell(row=1, column=2, value="\t摘要\n")
    headers = extract_headers(ws, header_row=1)
    assert headers == {"日期": 1, "摘要": 2}


def test_extract_headers_skips_empty_cells():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="日期")
    ws.cell(row=1, column=2, value=None)
    ws.cell(row=1, column=3, value="")
    ws.cell(row=1, column=4, value="金额")
    headers = extract_headers(ws, header_row=1)
    assert headers == {"日期": 1, "金额": 4}


def test_extract_headers_duplicate_keeps_first():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="金额")
    ws.cell(row=1, column=2, value="金额")
    headers = extract_headers(ws, header_row=1)
    assert headers == {"金额": 1}


# ──────────────────────────────────────────
# iter_data_rows
# ──────────────────────────────────────────
def test_iter_data_rows_yields_all_rows():
    wb = _make_wb()
    ws = read_sheet(wb)
    rows = list(iter_data_rows(ws, start_row=4))
    # row 4, 5 (空), 6 = 3 行（openpyxl 仅至 max_row）
    assert len(rows) >= 2
    first = rows[0]
    assert first[1] == "2026-04-22" or str(first[1]).startswith("2026")
    assert first[2] == "转账"


def test_iter_data_rows_keyed_by_column_index():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=3, value="C3")
    ws.cell(row=1, column=7, value="G1")
    rows = list(iter_data_rows(ws, start_row=1))
    assert rows[0][3] == "C3"
    assert rows[0][7] == "G1"


# ──────────────────────────────────────────
# is_empty_row
# ──────────────────────────────────────────
def test_is_empty_row_all_none():
    assert is_empty_row({1: None, 2: None, 3: None}) is True


def test_is_empty_row_all_whitespace():
    assert is_empty_row({1: "  ", 2: "\t", 3: "\n"}) is True


def test_is_empty_row_mixed_empty():
    assert is_empty_row({1: None, 2: "", 3: "   "}) is True


def test_is_empty_row_has_value():
    assert is_empty_row({1: "日期", 2: None}) is False


def test_is_empty_row_has_numeric():
    assert is_empty_row({1: 0, 2: None}) is False


def test_is_empty_row_empty_dict():
    assert is_empty_row({}) is True


# ──────────────────────────────────────────
# locate_merged_cells
# ──────────────────────────────────────────
def test_locate_merged_cells_empty():
    wb = Workbook()
    ws = wb.active
    assert locate_merged_cells(ws) == []


def test_locate_merged_cells_detects_ranges():
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("A1:C1")
    ws.merge_cells("B3:D5")
    merged = locate_merged_cells(ws)
    assert len(merged) == 2
    strs = sorted(str(r) for r in merged)
    assert "A1:C1" in strs
    assert "B3:D5" in strs
