"""test_template_fill · §P7 · 6 函数单测

load_template / fill / const / date_range_start / date_range_end / format_amount
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal

import pytest
from openpyxl import Workbook

from fund.primitives.template_fill import (
    const,
    date_range_end,
    date_range_start,
    fill,
    format_amount,
    load_template,
)


# ──────────────────────────────────────────
# helpers
# ──────────────────────────────────────────
def _wb_with(cells: dict) -> Workbook:
    """cells: {(row, col): value}"""
    wb = Workbook()
    ws = wb.active
    for (r, c), v in cells.items():
        ws.cell(row=r, column=c, value=v)
    return wb


# ──────────────────────────────────────────
# load_template
# ──────────────────────────────────────────
def test_load_template_roundtrip(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "${name}"
    ws["B1"] = "${amount}"
    f = tmp_path / "tmpl.xlsx"
    wb.save(f)
    wb2 = load_template(str(f))
    assert wb2.active["A1"].value == "${name}"
    assert wb2.active["B1"].value == "${amount}"


# ──────────────────────────────────────────
# fill
# ──────────────────────────────────────────
def test_fill_dollar_style_single_cell_replace():
    wb = _wb_with({(1, 1): "${name}"})
    fill(wb, "name", "张三")
    assert wb.active["A1"].value == "张三"


def test_fill_brace_style_single_cell_replace():
    wb = _wb_with({(1, 1): "{{name}}"})
    fill(wb, "name", "李四")
    assert wb.active["A1"].value == "李四"


def test_fill_inline_text_replace():
    wb = _wb_with({(1, 1): "金额: ${amount}元"})
    fill(wb, "amount", 1000)
    assert wb.active["A1"].value == "金额: 1000元"


def test_fill_decimal_coerces_to_float():
    wb = _wb_with({(1, 1): "${amount}"})
    fill(wb, "amount", Decimal("1234.56"))
    assert wb.active["A1"].value == 1234.56
    assert isinstance(wb.active["A1"].value, float)


def test_fill_preserves_date_type_on_full_token():
    wb = _wb_with({(1, 1): "${biz_date}"})
    d = date(2026, 4, 23)
    fill(wb, "biz_date", d)
    assert wb.active["A1"].value == d


def test_fill_applies_number_format():
    wb = _wb_with({(1, 1): "${amount}"})
    fill(wb, "amount", 1000, format="#,##0.00")
    assert wb.active["A1"].number_format == "#,##0.00"


def test_fill_no_hit_silent():
    wb = _wb_with({(1, 1): "nothing"})
    fill(wb, "missing", "x")  # 不抛
    assert wb.active["A1"].value == "nothing"


def test_fill_empty_placeholder_raises():
    wb = Workbook()
    with pytest.raises(ValueError, match="不可为空"):
        fill(wb, "", "x")


def test_fill_across_multiple_cells_and_sheets():
    wb = Workbook()
    ws1 = wb.active
    ws1["A1"] = "${title}"
    ws1["B2"] = "副标题 ${title} 续"
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "${title}"
    fill(wb, "title", "日报")
    assert ws1["A1"].value == "日报"
    assert ws1["B2"].value == "副标题 日报 续"
    assert ws2["A1"].value == "日报"


# ──────────────────────────────────────────
# const
# ──────────────────────────────────────────
def test_const_passthrough():
    assert const(123) == 123
    assert const("abc") == "abc"
    assert const(None) is None
    d = date(2026, 4, 1)
    assert const(d) is d


# ──────────────────────────────────────────
# date_range_start / date_range_end
# ──────────────────────────────────────────
def test_date_range_start_from_start_date():
    d = date(2026, 4, 1)
    assert date_range_start({"start_date": d}) == d


def test_date_range_start_from_period_start_fallback():
    d = date(2026, 4, 1)
    assert date_range_start({"period_start": d}) == d


def test_date_range_start_prefers_start_date():
    assert date_range_start({
        "start_date": date(2026, 4, 1),
        "period_start": date(2026, 5, 1),
    }) == date(2026, 4, 1)


def test_date_range_start_accepts_datetime():
    dt = datetime(2026, 4, 1, 12, 30)
    assert date_range_start({"start_date": dt}) == date(2026, 4, 1)


def test_date_range_start_accepts_string():
    assert date_range_start({"start_date": "2026-04-01"}) == date(2026, 4, 1)


def test_date_range_start_missing_raises():
    with pytest.raises(KeyError, match="start_date"):
        date_range_start({})


def test_date_range_end_basic():
    d = date(2026, 4, 30)
    assert date_range_end({"end_date": d}) == d


def test_date_range_end_fallback_period_end():
    d = date(2026, 4, 30)
    assert date_range_end({"period_end": d}) == d


def test_date_range_end_missing_raises():
    with pytest.raises(KeyError, match="end_date"):
        date_range_end({})


# ──────────────────────────────────────────
# format_amount
# ──────────────────────────────────────────
def test_format_amount_thousands():
    assert format_amount(Decimal("1234567.89")) == "1,234,567.89"


def test_format_amount_negative():
    assert format_amount(Decimal("-1234.5")) == "-1,234.50"


def test_format_amount_zero():
    assert format_amount(Decimal("0")) == "0.00"


def test_format_amount_digits_zero():
    assert format_amount(Decimal("1234.5"), digits=0) == "1,234"


def test_format_amount_digits_four():
    assert format_amount(Decimal("1.23456789"), digits=4) == "1.2346"


def test_format_amount_negative_digits_raises():
    with pytest.raises(ValueError, match="digits 不得为负"):
        format_amount(Decimal("1"), digits=-1)


def test_format_amount_none_returns_empty():
    assert format_amount(None) == ""


def test_format_amount_accepts_int_float():
    assert format_amount(1000) == "1,000.00"
    assert format_amount(1234.5) == "1,234.50"


def test_format_amount_non_numeric_falls_back_to_str():
    assert format_amount("abc") == "abc"


def test_format_amount_small_number():
    assert format_amount(Decimal("12.3")) == "12.30"
