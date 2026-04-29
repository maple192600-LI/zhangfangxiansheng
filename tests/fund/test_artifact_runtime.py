from __future__ import annotations

from datetime import date
from pathlib import Path
import sys

from openpyxl import Workbook
from sqlalchemy import text

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from agents.fund.skills._shared import (
    CASH_JOURNAL_PLACEHOLDERS_18,
    build_cash_rule,
    build_parser_code,
    parser_imports,
    rule_imports,
)
from db.tables import AICallLog, FundEvent, ImportBatch, ParserArtifact, RuleArtifact
from services import bank_import_service, report_service


NORMAL = "\u6b63\u5e38"
BANK_SOURCE = "\u7f51\u94f6\u5bfc\u5165"


def _session():
    import database

    return database.SessionLocal()


def _clean_runtime_tables(engine) -> None:
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM fund_events"))
        conn.execute(text("DELETE FROM operation_logs"))
        conn.execute(text("DELETE FROM parser_artifacts"))
        conn.execute(text("DELETE FROM rule_artifacts"))
        conn.execute(text("DELETE FROM ai_call_logs"))
        conn.execute(text("DELETE FROM import_batches WHERE batch_code LIKE 'P5%'"))


def _write_bank_workbook(path: Path, rows: int = 1) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "EntityId", "AccountId", "Summary", "Income", "Expense", "Balance"])
    for i in range(rows):
        ws.append([
            "2026-04-24",
            "E001",
            "A001",
            f"receipt {i + 1}",
            "1.23",
            "",
            str(100001.23 + i),
        ])
    wb.save(path)


def _write_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    for idx, placeholder in enumerate(CASH_JOURNAL_PLACEHOLDERS_18, start=1):
        ws.cell(row=idx, column=1, value=f"${{{placeholder}}}")
    wb.save(path)


def _active_parser(db) -> ParserArtifact:
    artifact = ParserArtifact(
        name="p5-runtime-bank-parser",
        kind="bank",
        account_code="A001",
        version=1,
        status="active",
        code=build_parser_code(kind="bank", default_account_code="A001"),
        primitives_imports=parser_imports(),
        sample_check_log={"ok": True},
        confidence=1,
        created_by="test",
        approved_by="test",
    )
    db.add(artifact)
    db.commit()
    db.refresh(artifact)
    return artifact


def _active_rule(db) -> RuleArtifact:
    bindings, loop, check_log = build_cash_rule(None, CASH_JOURNAL_PLACEHOLDERS_18)
    artifact = RuleArtifact(
        name="p5-runtime-cash-rule",
        version=1,
        status="active",
        placeholder_bindings=bindings,
        loop_config=loop,
        primitives_imports=rule_imports(),
        sample_check_log=check_log.model_dump(),
        confidence=1,
        created_by="test",
        approved_by="test",
    )
    db.add(artifact)
    db.commit()
    db.refresh(artifact)
    return artifact


def test_bank_commit_runs_parser_artifact_and_inserts_fund_events(primitives_db, tmp_path, monkeypatch):
    _clean_runtime_tables(primitives_db)
    upload_dir = tmp_path / "uploads"
    upload_dir.mkdir()
    monkeypatch.setattr(bank_import_service, "DATA_DIR", str(tmp_path))

    file_path = upload_dir / "P5BANK001_sample.xlsx"
    _write_bank_workbook(file_path)

    with _session() as db:
        parser = _active_parser(db)
        batch = ImportBatch(
            batch_code="P5BANK001",
            source_type="bank",
            source_name="sample.xlsx",
            status="uploaded",
        )
        db.add(batch)
        db.commit()

        result = bank_import_service.commit(db, "P5BANK001", parser.id)
        rows = db.query(FundEvent).all()

        assert result["inserted_rows"] == 1
        assert len(rows) == 1
        assert rows[0].parser_artifact_id == parser.id
        assert rows[0].batch_id == batch.id
        assert rows[0].account_code == "A001"
        assert rows[0].state == NORMAL


def test_runtime_parser_commit_1000_rows_does_not_write_ai_call_logs(primitives_db, tmp_path, monkeypatch):
    _clean_runtime_tables(primitives_db)
    upload_dir = tmp_path / "uploads"
    upload_dir.mkdir()
    monkeypatch.setattr(bank_import_service, "DATA_DIR", str(tmp_path))

    file_path = upload_dir / "P5BANK1000_sample.xlsx"
    _write_bank_workbook(file_path, rows=1000)

    with _session() as db:
        parser = _active_parser(db)
        db.add(ImportBatch(
            batch_code="P5BANK1000",
            source_type="bank",
            source_name="sample.xlsx",
            status="uploaded",
        ))
        db.commit()

        before = db.query(AICallLog).count()
        result = bank_import_service.commit(db, "P5BANK1000", parser.id)
        after = db.query(AICallLog).count()

        assert result["inserted_rows"] == 1000
        assert db.query(FundEvent).count() == 1000
        assert after == before


def test_generate_report_runs_rule_artifact_and_fills_18_placeholders(primitives_db, tmp_path):
    _clean_runtime_tables(primitives_db)
    template_path = tmp_path / "cash_journal_template.xlsx"
    _write_template(template_path)

    with _session() as db:
        db.add(FundEvent(
            business_date=date(2026, 4, 24),
            entity_code="E001",
            entity_name="Example Entity",
            account_code="A001",
            account_name="Example Account",
            summary="receipt",
            counterparty="customer",
            amount_in=300,
            amount_out=0,
            rolling_balance=100300,
            state=NORMAL,
            source=BANK_SOURCE,
        ))
        rule = _active_rule(db)
        db.commit()

        wb = report_service.generate_report(db, rule.id, {
            "template_file": str(template_path),
            "entity_code": "E001",
            "account_code": "A001",
            "start_date": date(2026, 4, 1),
            "end_date": date(2026, 4, 30),
        })

    values = [cell.value for row in wb.active.iter_rows() for cell in row]
    assert len(values) == len(CASH_JOURNAL_PLACEHOLDERS_18)
    assert all(not (isinstance(value, str) and ("${" in value or "{{" in value)) for value in values)
