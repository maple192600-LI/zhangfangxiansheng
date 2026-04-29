from __future__ import annotations

from datetime import date
from pathlib import Path
import sys

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import text

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from api.fund_agent import router as fund_router
from api.reports import router as reports_router
from database import get_db
from db.tables import FundEvent


NORMAL = "\u6b63\u5e38"
BANK_SOURCE = "\u7f51\u94f6\u5bfc\u5165"


def _session():
    import database

    return database.SessionLocal()


def _client():
    app = FastAPI()
    app.include_router(fund_router, prefix="/api")
    app.include_router(reports_router, prefix="/api")

    def override_db():
        with _session() as db:
            yield db

    app.dependency_overrides[get_db] = override_db
    return TestClient(app)


def _clean(engine):
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM fund_events"))
        conn.execute(text("DELETE FROM operation_logs"))
        conn.execute(text("DELETE FROM template_inference_job"))
        conn.execute(text("DELETE FROM rule_artifacts"))
        conn.execute(text("DELETE FROM parser_artifacts"))
        conn.execute(text("DELETE FROM ai_call_logs"))


def _write_bank_sample(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "EntityId", "AccountId", "Summary", "Income", "Expense", "Balance"])
    ws.append(["2026-04-24", "E001", "A001", "receipt", "300.00", "", "100300"])
    wb.save(path)


def _write_template(path: Path):
    placeholders = [
        "报表标题", "开始期间", "结束期间", "板块", "核算方式", "开户行", "账户信息", "银行编号",
        "月", "日", "月初余额", "摘要", "收入", "支出", "余额", "本月收入小计", "本月支出小计", "月末余额",
    ]
    wb = Workbook()
    ws = wb.active
    for idx, name in enumerate(placeholders, start=1):
        ws.cell(row=idx, column=1, value=f"${{{name}}}")
    wb.save(path)


def test_parser_skill_invoke_detail_and_approve(primitives_db, tmp_path):
    _clean(primitives_db)
    sample = tmp_path / "bank.xlsx"
    _write_bank_sample(sample)
    client = _client()

    response = client.post("/api/fund/agent/skills/parser.bank/invoke", json={
        "account_code": "A001",
        "sample_file": str(sample),
        "privacy_mode": "standard",
    })
    data = response.json()
    assert data["code"] == 0
    artifact_id = data["data"]["artifact_id"]

    detail = client.get(f"/api/fund/parsers/{artifact_id}").json()
    assert detail["code"] == 0
    assert detail["data"]["status"] == "draft"
    assert detail["data"]["sample_check_log"]["parsed_rows"] == 1

    approved = client.post(f"/api/fund/parsers/{artifact_id}/approve").json()
    assert approved["code"] == 0
    assert approved["data"]["status"] == "active"


def test_template_upload_rule_approve_and_report_download(primitives_db, tmp_path, monkeypatch):
    _clean(primitives_db)
    monkeypatch.setattr("api.fund_agent.DATA_DIR", str(tmp_path))
    monkeypatch.setattr("api.reports.DATA_DIR", str(tmp_path))
    template = tmp_path / "cash_template.xlsx"
    _write_template(template)
    client = _client()

    with template.open("rb") as f:
        uploaded = client.post(
            "/api/fund/templates/upload",
            data={"kind": "cash_journal"},
            files={"file": ("cash_template.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        ).json()
    assert uploaded["code"] == 0
    rule_id = uploaded["data"]["rule_draft_id"]
    job_id = uploaded["data"]["job_id"]

    job = client.get(f"/api/fund/templates/jobs/{job_id}").json()
    assert job["code"] == 0
    assert job["data"]["rule_artifact_id"] == rule_id

    approved = client.post(f"/api/fund/rules/{rule_id}/approve").json()
    assert approved["code"] == 0
    assert approved["data"]["status"] == "active"

    with _session() as db:
        db.add(FundEvent(
            business_date=date(2026, 4, 24),
            entity_code="E001",
            entity_name="Example Entity",
            account_code="A001",
            account_name="Example Account",
            summary="receipt",
            counterparty="customer",
            amount_in=300,
            amount_out=0,
            rolling_balance=100300,
            state=NORMAL,
            source=BANK_SOURCE,
        ))
        db.commit()

    generated = client.post("/api/reports/generate", json={
        "rule_id": rule_id,
        "account_code": "A001",
        "entity_code": "E001",
        "period_start": "2026-04-01",
        "period_end": "2026-04-30",
    }).json()
    assert generated["code"] == 0
    assert generated["data"]["placeholder_filled"] == 18

    download = client.get(generated["data"]["download_url"])
    assert download.status_code == 200
    assert download.content[:2] == b"PK"
