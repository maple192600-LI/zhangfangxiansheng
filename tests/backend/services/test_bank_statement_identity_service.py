import os
import tempfile

import pytest
from openpyxl import Workbook

from services.bank_statement_identity_service import (
    extract_identity_hints,
    extract_identity_hints_from_workbook,
)


def _make_workbook(rows):
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    return wb


def _save_temp(wb):
    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(f.name)
    f.close()
    return f.name


# ── 账号提取 ──

def test_extract_account_number_from_keyword_cell():
    wb = _make_workbook([
        ["账号：6217001234567890"],
        ["日期", "摘要", "收入", "支出", "余额"],
        ["2026-04-24", "test", 100, 0, 100],
    ])
    result = extract_identity_hints_from_workbook(wb)
    assert result["identity_hints"]["account_number"] == "6217001234567890"
    assert result["identity_hints"]["account_last_four"] == "7890"


def test_extract_account_number_adjacent_cells():
    wb = _make_workbook([
        ["账号", "6217001234567890"],
        ["日期", "摘要", "收入", "支出", "余额"],
        ["2026-04-24", "test", 100, 0, 100],
    ])
    result = extract_identity_hints_from_workbook(wb)
    assert result["identity_hints"]["account_number"] == "6217001234567890"


def test_account_number_standalone_long_digits():
    wb = _make_workbook([
        ["6217001234567890"],
        ["日期", "摘要", "收入", "支出", "余额"],
    ])
    result = extract_identity_hints_from_workbook(wb)
    assert result["identity_hints"]["account_number"] == "6217001234567890"


def test_reject_date_as_account_number():
    wb = _make_workbook([
        ["20260424"],
        ["日期", "摘要", "收入", "支出", "余额"],
    ])
    result = extract_identity_hints_from_workbook(wb)
    assert result["identity_hints"]["account_number"] == ""


def test_reject_short_number_as_account():
    wb = _make_workbook([
        ["12345"],
        ["日期", "摘要", "收入", "支出", "余额"],
    ])
    result = extract_identity_hints_from_workbook(wb)
    assert result["identity_hints"]["account_number"] == ""


# ── 银行名提取 ──

def test_extract_bank_name_from_cell():
    wb = _make_workbook([
        ["中国银行对账单"],
        ["日期", "摘要", "收入", "支出", "余额"],
    ])
    result = extract_identity_hints_from_workbook(wb)
    assert result["bank_hint"] == "中国银行"
    assert result["identity_hints"]["bank_name"] == "中国银行"


def test_extract_bank_short_name():
    wb = _make_workbook([
        ["工行流水"],
        ["日期", "摘要", "收入", "支出", "余额"],
    ])
    result = extract_identity_hints_from_workbook(wb)
    # Short names without '银行' are NOT detected — identity service
    # returns raw text only; normalization is done by match service via DB
    assert result["identity_hints"]["bank_name"] == ""


def test_extract_bank_from_filename():
    wb = _make_workbook([
        ["日期", "摘要", "收入", "支出", "余额"],
    ])
    result = extract_identity_hints_from_workbook(wb, filename="中国银行对账单.xlsx")
    assert result["identity_hints"]["bank_name"] == "中国银行"


# ── 户名/单位名提取 ──

def test_extract_account_holder_name():
    wb = _make_workbook([
        ["户名：测试建筑有限公司"],
        ["日期", "摘要", "收入", "支出", "余额"],
    ])
    result = extract_identity_hints_from_workbook(wb)
    assert result["identity_hints"]["account_name"] == "测试建筑有限公司"


def test_extract_entity_name():
    wb = _make_workbook([
        ["单位名称：测试实业发展有限公司"],
        ["日期", "摘要", "收入", "支出", "余额"],
    ])
    result = extract_identity_hints_from_workbook(wb)
    assert result["identity_hints"]["entity_name"] == "测试实业发展有限公司"


# ── 开户行提取 ──

def test_extract_branch_name():
    wb = _make_workbook([
        ["开户行：中国银行北京分行"],
        ["日期", "摘要", "收入", "支出", "余额"],
    ])
    result = extract_identity_hints_from_workbook(wb)
    assert result["identity_hints"]["branch_name"] == "中国银行北京分行"


# ── 综合提取 ──

def test_extract_all_identity_hints():
    wb = _make_workbook([
        ["中国银行对账单"],
        ["账号：6217001234567890"],
        ["户名：测试建筑有限公司"],
        ["开户行：中国银行北京分行"],
        [],
        ["日期", "摘要", "收入", "支出", "余额"],
        ["2026-04-24", "test", 100, 0, 100],
    ])
    result = extract_identity_hints_from_workbook(wb)
    h = result["identity_hints"]
    assert h["bank_name"] == "中国银行"
    assert h["account_number"] == "6217001234567890"
    assert h["account_last_four"] == "7890"
    assert h["account_name"] == "测试建筑有限公司"
    assert h["branch_name"] == "中国银行北京分行"
    assert result["confidence"] > 0.8


def test_empty_workbook_returns_empty_hints():
    wb = Workbook()
    result = extract_identity_hints_from_workbook(wb)
    assert result["confidence"] == 0.0
    assert result["identity_hints"]["account_number"] == ""


# ── 文件名线索 ──

def test_filename_last_four_recorded_in_evidence():
    wb = Workbook()
    result = extract_identity_hints_from_workbook(wb, filename="银行流水7890.xlsx")
    assert result["identity_hints"]["filename_hint"] == "7890"
    assert result["evidence"]["filename"] == "银行流水7890.xlsx"


def test_filename_weak_hint_low_confidence():
    wb = Workbook()
    result = extract_identity_hints_from_workbook(wb, filename="银行流水7890.xlsx")
    assert result["confidence"] == 0.2


def test_file_path_extraction():
    wb = _make_workbook([
        ["账号：6217001234567890"],
        ["日期", "摘要"],
    ])
    path = _save_temp(wb)
    try:
        result = extract_identity_hints(path, filename="test.xlsx")
        assert result["identity_hints"]["account_number"] == "6217001234567890"
    finally:
        os.unlink(path)


# ── evidence ──

def test_evidence_contains_cells_and_headers():
    wb = _make_workbook([
        ["中国银行"],
        ["日期", "摘要", "收入", "支出", "余额"],
        ["2026-04-24", "test", 100, 0, 100],
    ])
    result = extract_identity_hints_from_workbook(wb)
    assert len(result["evidence"]["cells"]) > 0
    assert len(result["evidence"]["headers"]) >= 3
    assert result["format_fingerprint"].startswith("fp_")


def test_standalone_last_four():
    wb = _make_workbook([
        ["尾号7890"],
        ["日期", "摘要", "收入", "支出", "余额"],
    ])
    result = extract_identity_hints_from_workbook(wb)
    assert result["identity_hints"]["account_last_four"] == "7890"


# ── _BANK_NAMES removed ──

def test_no_bank_names_hardcoded_dict():
    import services.bank_statement_identity_service as mod
    assert not hasattr(mod, "_BANK_NAMES"), "_BANK_NAMES must be removed"


def test_full_bank_name_extracted_as_raw():
    wb = _make_workbook([
        ["招商银行流水明细"],
        ["日期", "摘要"],
    ])
    result = extract_identity_hints_from_workbook(wb)
    assert result["identity_hints"]["bank_name"] == "招商银行"


# ── bank_text_candidates ──

def test_candidates_include_filename():
    wb = Workbook()
    result = extract_identity_hints_from_workbook(wb, filename="工行流水.xlsx")
    cands = result["bank_text_candidates"]
    assert "工行流水.xlsx" in cands or "工行流水" in cands


def test_candidates_include_cell_text():
    wb = _make_workbook([
        ["账号：6217001234567890"],
        ["日期", "摘要", "收入", "支出", "余额"],
    ])
    result = extract_identity_hints_from_workbook(wb)
    cands = result["bank_text_candidates"]
    assert any("6217001234567890" in c for c in cands)


def test_candidates_deduplicated():
    wb = _make_workbook([
        ["中国银行"],
        ["中国银行"],
        ["日期", "摘要"],
    ])
    result = extract_identity_hints_from_workbook(wb)
    cands = result["bank_text_candidates"]
    assert cands.count("中国银行") == 1
