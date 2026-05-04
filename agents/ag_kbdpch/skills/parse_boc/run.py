# -*- coding: utf-8 -*-
"""
中国银行流水解析 v5 — 财务思维版
- 摘要按出纳记账习惯生成
- 收入=借方, 支出=贷方
- 内部转账识别账户别名
- 输出匹配 fund_events 基础表结构
"""
import json, os, re, xlrd

# 本方实体名称（用于判断内部往来）
_OWN_ENTITY = "山西喜跃发道路建设养护集团有限公司"

# 账户别名映射（仅本方实体的银行账户）
_ACCOUNT_ALIAS = {
    '04165101040025287': '农行阳曲支行',
    '24091401040028011': '农行景洪支行',
    '139252780563': '中行并州支行',
    '04166501040003128': '农行工商街支行',
}

# 贷款子账户（中行内部，非独立银行账户）
_LOAN_ACCOUNTS = {'144270613068', '145521151705', '148024657996', '146775456631'}

# 对方名称缩写（用于摘要中精简显示）
_SHORT_NAME = {
    "山西喜跃发道路建设养护集团有限公司": "养护集团",
    "山西龙发建材有限公司": "龙发建材",
    "阳曲县华洋建筑机械设备租赁部": "华洋机械",
    "太原市鑫和谐矿业有限公司": "鑫和谐矿业",
}


def run(params):
    file_path = params.get("file_path", "")

    if not os.path.exists(file_path):
        import glob as g
        for pat in ["/**/中行银行流水.xls", "/**/*中行*.xls*", "/**/*BOC*.xls*"]:
            for c in g.glob(pat, recursive=True):
                file_path = c
                break
            else:
                continue
            break
        else:
            return {"ok": False, "error": "找不到中行流水文件"}

    ext = os.path.splitext(file_path)[1].lower()

    # ---- 读取 ----
    if ext == '.xls':
        wb = xlrd.open_workbook(file_path)
        sheets = []
        for sn in wb.sheet_names():
            ws = wb.sheet_by_name(sn)
            sheets.append((sn, [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]))
    else:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets = [(sn, [list(row) for row in wb[sn].iter_rows(values_only=True)]) for sn in wb.sheetnames]

    if not sheets:
        return {"ok": False, "error": "工作簿无工作表"}

    sheet_name, rows = sheets[0]

    # ---- 找表头 ----
    header_idx = None
    for i, row in enumerate(rows):
        if row is None or len(row) < 5:
            continue
        row_text = ' '.join([str(c) for c in row if c])
        if '交易日期' in row_text and ('交易时间' in row_text or '业务类型' in row_text):
            header_idx = i
            break

    if header_idx is None:
        return {"ok": False, "error": "未找到表头行，需含交易日期字段"}

    header = [str(c) if c else '' for c in rows[header_idx]]

    # ---- 精准列映射 ----
    mapping = {}

    for j, h in enumerate(header):
        if not h:
            continue
        hl = h.lower()

        if re.search(r'交易类型', h) and not re.search(r'业务类型', h):
            mapping['txn_type'] = j
        if re.search(r'业务类型|business type', hl):
            mapping['business_type'] = j
        if re.search(r'付款人名称', h) and not re.search(r'名义', h):
            mapping['payer_name'] = j
        if re.search(r'付款人账号', h) and not re.search(r'名义', h):
            mapping['payer_account'] = j
        if re.search(r'收款人名称', h) and not re.search(r'名义', h) and not re.search(r'(开户行|行号|行名|账号)', h):
            mapping['payee_name'] = j
        if re.search(r'收款人账号', h) and not re.search(r'名义', h):
            mapping['payee_account'] = j
        if re.search(r'交易日期|transaction.*date', hl) and not re.search(r'起息|value', hl):
            mapping['date'] = j
        if re.search(r'交易时间|transaction.*time', hl):
            mapping['time'] = j
        if re.search(r'交易金额|trade.*amount', hl):
            mapping['signed_amount'] = j
        if re.search(r'交易后余额|after.*balance', hl):
            mapping['balance'] = j
        if re.search(r'交易流水号|transaction.*reference.*number', hl):
            mapping['txn_ref'] = j
        if re.search(r'凭证号码|voucher.*number', hl):
            mapping['voucher_no'] = j
        if re.search(r'^摘要|\[摘要', h) or re.search(r'\]摘要|摘要\[', h):
            mapping['reference'] = j
        if re.search(r'^用途|\[用途|用途\[', h):
            mapping['purpose'] = j
        if re.search(r'交易附言|remark\]', hl) and not re.search(r'remarks', hl):
            mapping['remark'] = j
        if re.search(r'^备注|备注\[|\[备注|remarks\]', hl) and not re.search(r'交易附言', hl):
            mapping['remarks'] = j

    # ---- 提取交易 ----
    data_start = header_idx + 1
    while data_start < len(rows) and all(c is None or str(c).strip() == '' for c in rows[data_start]):
        data_start += 1

    transactions = []
    debug_rows = []

    for i in range(data_start, len(rows)):
        row = rows[i]
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue

        first_val = str(row[0] or '')
        if '合计' in first_val or '小计' in first_val:
            continue

        def _v(key):
            col = mapping.get(key)
            if col is None or col >= len(row):
                return None
            val = row[col]
            if val is None:
                return None
            if isinstance(val, float):
                return val
            s = str(val).strip()
            return s if s and s != 'None' else None

        def _fv(key):
            col = mapping.get(key)
            if col is None or col >= len(row):
                return None
            val = row[col]
            if val is None:
                return None
            if isinstance(val, (int, float)):
                return float(val)
            s = str(val).strip().replace(',', '')
            try:
                return float(s)
            except:
                return None

        date_str = _v('date')
        time_str = _v('time')
        txn_type = _v('txn_type')
        biz_type = _v('business_type')
        purpose = _v('purpose')
        remark = _v('remark')
        remarks = _v('remarks')
        reference = _v('reference')
        txn_ref = _v('txn_ref')
        voucher = _v('voucher_no')
        payer_name = _v('payer_name')
        payer_acct = _v('payer_account')
        payee_name = _v('payee_name')
        payee_acct = _v('payee_account')
        signed_amt = _fv('signed_amount')
        balance = _fv('balance')

        # 日期格式化
        date_display = date_str
        if date_str and re.match(r'\d{8}', date_str):
            date_display = "{}-{}-{}".format(date_str[:4], date_str[4:6], date_str[6:8])

        # ---- 方向判断 (借贷方向) ----
        # 收入 = 借方, 支出 = 贷方
        direction = None
        income = None
        expense = None

        if txn_type:
            if '来' in txn_type:
                direction = '收入'
            elif '往' in txn_type:
                direction = '支出'

        if signed_amt is not None:
            if signed_amt > 0:
                income = signed_amt
                expense = 0.0
                direction = direction or '收入'
            elif signed_amt < 0:
                income = 0.0
                expense = abs(signed_amt)
                direction = direction or '支出'
            else:
                income = 0.0
                expense = 0.0
        else:
            income = income or 0.0
            expense = expense or 0.0

        # ---- 对方 ----
        # 收入时对方是付款人，支出时对方是收款人
        if direction == '收入':
            counterparty = payer_name or payee_name or ''
        elif direction == '支出':
            counterparty = payee_name or payer_name or ''
        else:
            counterparty = payee_name or payer_name or ''

        # 对方账号
        counterparty_acct = ''
        if direction == '收入':
            counterparty_acct = str(payer_acct or '')
        elif direction == '支出':
            counterparty_acct = str(payee_acct or '')

        # ---- 判断是否内部转账 ----
        is_internal_transfer = False
        internal_alias = ''
        if counterparty == _OWN_ENTITY and counterparty_acct:
            if counterparty_acct in _ACCOUNT_ALIAS:
                is_internal_transfer = True
                internal_alias = _ACCOUNT_ALIAS[counterparty_acct]
            elif counterparty_acct in _LOAN_ACCOUNTS:
                is_internal_transfer = True
                internal_alias = '中行贷款户'

        # ---- 生成会计摘要 ----
        summary = _build_accounting_summary(
            biz_type=biz_type,
            reference=reference,
            purpose=purpose,
            remark=remark,
            remarks=remarks,
            counterparty=counterparty,
            counterparty_acct=counterparty_acct,
            direction=direction,
            is_internal_transfer=is_internal_transfer,
            internal_alias=internal_alias,
        )

        transactions.append({
            '_row': i + 1,
            'business_date': date_display,
            'direction': direction,
            'summary': summary,
            'counterparty': counterparty,
            'amount_in': income,
            'amount_out': expense,
            'balance': balance,
            '_raw': {
                'txn_type': txn_type,
                'biz_type': biz_type,
                'payer_name': payer_name,
                'payer_acct': payer_acct,
                'payee_name': payee_name,
                'payee_acct': payee_acct,
                'reference': reference,
                'purpose': purpose,
                'remark': remark,
                'remarks': remarks,
                'txn_ref': txn_ref,
                'voucher_no': voucher,
                'time': time_str,
                'is_internal': is_internal_transfer,
            }
        })

    # ---- 余额校验 ----
    balance_ok = True
    balance_issues = []
    if transactions and transactions[0]['balance'] is not None:
        for k in range(1, len(transactions)):
            prev_bal = transactions[k-1]['balance']
            cur_bal = transactions[k]['balance']
            cur_in = transactions[k]['amount_in'] or 0
            cur_out = transactions[k]['amount_out'] or 0
            if prev_bal is not None and cur_bal is not None:
                expected = prev_bal + cur_in - cur_out
                if abs(expected - cur_bal) > 0.02:
                    balance_ok = False
                    balance_issues.append({
                        'row': transactions[k]['_row'],
                        'expected': round(expected, 2),
                        'actual': cur_bal,
                    })

    # ---- 导出 CSV 预览表 ----
    import csv
    csv_path = "/zhangfangxiansheng/agents/ag_grlslj/workspace/inbox/中行银行流水_预览.csv"
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(['序号', '日期', '方向', '摘要', '对方', '收入金额', '支出金额', '余额', '对方账号', '业务类型', '用途/附言'])
        for idx, t in enumerate(transactions, 1):
            w.writerow([
                idx,
                t['business_date'],
                t['direction'],
                t['summary'],
                t['counterparty'],
                t['amount_in'] if t['amount_in'] else '',
                t['amount_out'] if t['amount_out'] else '',
                t['balance'] if t['balance'] is not None else '',
                t['_raw'].get('payer_acct', '') or t['_raw'].get('payee_acct', ''),
                t['_raw'].get('biz_type', ''),
                t['_raw'].get('purpose', '') or t['_raw'].get('remark', '') or '',
            ])

    return {
        "ok": True,
        "file_name": os.path.basename(file_path),
        "file_format": ext,
        "csv_path": csv_path,
        "meta": {
            "total_rows": len(rows),
            "header_row": header_idx,
            "data_start": data_start,
            "columns_in_file": len(header),
        },
        "transaction_count": len(transactions),
        "transactions": transactions,
        "balance_check": {
            "ok": balance_ok,
            "issues": balance_issues,
        },
        "_debug": {
            "header_idx": header_idx,
            "header": header[:30],
            "data_start": data_start,
            "mapping": dict(sorted(mapping.items())),
        }
    }


def _build_accounting_summary(biz_type=None, reference=None, purpose=None, remark=None,
                              remarks=None, counterparty=None, counterparty_acct=None,
                              direction=None, is_internal_transfer=False, internal_alias=''):
    """按出纳记账习惯生成摘要"""

    ref = str(reference).strip() if reference and str(reference).strip() != 'None' else ''
    pur = str(purpose).strip() if purpose and str(purpose).strip() != 'None' else ''
    rem = str(remark).strip() if remark and str(remark).strip() != 'None' else ''
    rems = str(remarks).strip() if remarks and str(remarks).strip() != 'None' else ''
    cp = counterparty or ''

    # ====== 1. 银行费用类 ======
    if biz_type == '短信收费' or 'SMSP' in ref:
        return '短信费'

    if biz_type == '收费' or (direction == '支出' and not cp):
        if '对公跨行转账汇款手续费' in ref:
            return '转账手续费'
        if '邮费' in ref:
            return '邮费'
        if '询证' in ref or '证明' in ref or '资信' in ref:
            return '银行证明费'
        if '账户管理' in ref or re.match(r'\d{4}/\d{2}/\d{2}', ref):
            return '账户管理费'
        # 通用费用
        return '银行手续费'

    # ====== 2. 利息类 ======
    if biz_type == '结息' or '结息' in (ref + pur):
        return '结息'

    # ====== 3. 贷款类 ======
    if biz_type == '贷款放款':
        return '贷款放款'
    if biz_type == '贷款还款':
        return '归还贷款'

    # ====== 4. 内部转账 ======
    if is_internal_transfer and internal_alias:
        if direction == '收入':
            return '{}转入'.format(internal_alias)
        else:
            return '转{}'.format(internal_alias)

    # ====== 5. 外部交易 ======
    # 取对方简称
    short_cp = _SHORT_NAME.get(cp, cp)
    # 去掉"有限公司"等后缀缩短
    if short_cp == cp:
        short_cp = cp.replace('有限公司', '').replace('有限责任公司', '').strip()

    # 用附言/用途作为款项性质
    nature = rem or pur or ''

    # 过滤掉纯编码类附言
    if nature and re.match(r'^[A-Z0-9\s\-]{8,}$', nature):
        nature = ''

    if direction == '收入':
        if '往来' in nature:
            return '收到{}往来款'.format(short_cp)
        if nature and len(nature) <= 10:
            return '收到{}{}'.format(short_cp, nature)
        return '收到{}款项'.format(short_cp)
    else:
        if nature and len(nature) <= 10:
            return '支付{}{}'.format(short_cp, nature)
        return '支付{}款项'.format(short_cp)
