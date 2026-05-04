
# -*- coding: utf-8 -*-
"""
银行流水智能解析 v3 — 多银行通用版
- 自动识别银行格式
- 联查数据中心实体判断内/外部交易
- 按各银行约定规则生成摘要
- 输出匹配 fund_events 基础表结构
"""

import json, os, re, csv

# ===== 银行识别特征 =====
BANK_SIGNATURES = {
    "中国银行": {
        "headers": ["交易类型", "Transaction Type", "付款人名称", "Payer's Name"],
        "format_hint": "中英双语38列表头",
    },
    "农业银行": {
        "headers": ["交易日期", "摘要", "对方户名", "借方金额", "贷方金额"],
        "format_hint": "中文12列左右表头",
    },
}

# ===== 标准字段定义 =====
STANDARD_FIELDS = [
    "business_date",    # 交易日期 YYYY-MM-DD
    "direction",        # 收入 / 支出
    "summary",          # 会计摘要
    "counterparty",     # 对方名称
    "amount_in",        # 收入金额
    "amount_out",       # 支出金额
    "balance",          # 余额
    "counterparty_acct",# 对方账号
    "biz_type",         # 业务类型
    "txn_ref",          # 交易流水号
]

# ===== 通用费用识别 =====
FEE_PATTERNS = [
    (r'短信|SMSP|sms', '短信费'),
    (r'对公跨行.*手续费|跨行转账.*费', '转账手续费'),
    (r'询证|验资|开户证明|资信证明|结算交易证明|结算资信', '支账户管理费'),
    (r'邮费|快件|特快|邮寄', '付款手续费'),
    (r'账户管理|账户年费|Account.*Manage', '账户管理费'),
    (r'结息|利息收入|Interest', '结息'),
    (r'贷款放款|Loan.*Disburse', '贷款放款'),
    (r'贷款还款|Loan.*Repay|还本', '归还贷款'),
    (r'缴税|税务|纳税', '缴税'),
    (r'社保|公积金|住房', '社保公积金'),
]

# ===== 核心函数 =====

def detect_bank(rows):
    """从原始数据检测银行"""
    for i, row in enumerate(rows[:20]):
        if row is None: continue
        row_text = ' '.join([str(c) for c in row if c])
        if '中国银行' in row_text or 'Bank of China' in row_text:
            return "中国银行"
        if '交易日期' in row_text and '交易时间' in row_text:
            return "中国银行"
        if '农业银行' in row_text or 'ABC' in row_text:
            return "农业银行"
    return "未知银行"


def find_header(rows, bank_name):
    """找到表头行"""
    patterns = {
        "中国银行": r'交易日期.*交易时间|交易类型.*业务类型',
        "农业银行": r'交易日期.*摘要.*对方',
    }
    pattern = patterns.get(bank_name, r'交易日期|日期')
    for i, row in enumerate(rows):
        if row is None or len(row) < 3: continue
        row_text = ' '.join([str(c) for c in row if c])
        if re.search(pattern, row_text):
            return i
    return None


def map_columns(header, bank_name):
    """列名→标准字段映射"""
    mapping = {}
    for j, h in enumerate(header):
        if not h: continue
        h_clean = str(h).strip()
        hl = h_clean.lower()
        
        # 日期
        if re.search(r'交易日期|transaction.*date', hl) and not re.search(r'起息', hl):
            mapping['date'] = j
        # 时间
        if re.search(r'交易时间|transaction.*time', hl):
            mapping['time'] = j
        # 交易类型（方向）
        if bank_name == "中国银行":
            if re.search(r'^交易类型|Transaction Type', h_clean) and '业务' not in h_clean:
                mapping['txn_type'] = j
        # 业务类型
        if re.search(r'业务类型|business type', hl):
            mapping['biz_type'] = j
        # 金额
        if re.search(r'交易金额|trade.*amount|发生额', hl):
            mapping['signed_amt'] = j
        if re.search(r'借方金额|收入金额|贷方发生额', h_clean):
            mapping['debit_amt'] = j
        if re.search(r'贷方金额|支出金额|借方发生额', h_clean):
            mapping['credit_amt'] = j
        # 余额
        if re.search(r'余额|balance', hl) and not re.search(r'可用', hl):
            mapping['balance'] = j
        # 对方
        if re.search(r'付款人名称|payer.*name', hl):
            mapping['payer_name'] = j
        if re.search(r'收款人名称|payee.*name|beneficiary.*name', hl) and not re.search(r'(行号|行名|账号)', h_clean):
            mapping['payee_name'] = j
        if re.search(r'对方.*名称|对方户名|counterparty', hl):
            mapping['counterparty'] = j
        # 账号
        if re.search(r'付款人账号|debit.*account', hl):
            mapping['payer_acct'] = j
        if re.search(r'收款人账号|payee.*account', hl):
            mapping['payee_acct'] = j
        # 摘要相关
        if re.search(r'用途|purpose', hl) and not re.search(r'摘要', h_clean):
            mapping['purpose'] = j
        if re.search(r'附言|remark\]', hl) and not re.search(r'remarks', hl):
            mapping['remark'] = j
        if re.search(r'^备注|remarks\]', hl):
            mapping['remarks'] = j
        if re.search(r'摘要|reference', hl):
            mapping['reference'] = j
        # 流水号/凭证
        if re.search(r'流水号|reference.*number', hl):
            mapping['txn_ref'] = j
        if re.search(r'凭证号|voucher', hl):
            mapping['voucher_no'] = j

    return mapping


def build_summary(row_data, mapping, internal_entities, direction, bank_name):
    """根据银行规则生成摘要"""
    def v(key):
        col = mapping.get(key)
        if col is None or col >= len(row_data): return ''
        val = row_data[col]
        return str(val).strip() if val and str(val).strip() != 'None' else ''

    biz = v('biz_type')
    ref = v('reference')
    pur = v('purpose')
    rem = v('remark')
    rems = v('remarks')

    # 费用识别
    check_text = f"{biz} {ref} {pur}"
    for pattern, label in FEE_PATTERNS:
        if re.search(pattern, check_text, re.I):
            return label

    # 内部交易（中国银行规则）
    if bank_name == "中国银行":
        cp_name = ''
        cp_acct = ''
        if direction == '收入':
            cp_name = v('payer_name')
            cp_acct = v('payer_acct')
        else:
            cp_name = v('payee_name')
            cp_acct = v('payee_acct')

        if cp_name in internal_entities:
            from entities_db import get_entity_short
            short = get_entity_short(internal_entities, cp_name)
            bank_hint = guess_bank(cp_acct)
            if direction == '收入':
                return f"{short}{bank_hint}转入"
            else:
                return f"转{short}{bank_hint}"

    # 兜底：用业务类型
    if biz: return biz
    if pur: return pur
    return '待补摘要'


def guess_bank(acct):
    """根据账号前缀猜银行"""
    if not acct: return ''
    acct = str(acct)
    if acct.startswith('041'): return '农行'
    if acct.startswith('139'): return '中行'
    if acct.startswith('240'): return '农行'
    if acct.startswith('103'): return '农行'
    if acct.startswith('201'): return '晋商'
    return ''


def run(params):
    """主入口"""
    file_path = params.get("file_path", "")

    if not os.path.exists(file_path):
        return {"ok": False, "error": f"文件不存在: {file_path}"}

    ext = os.path.splitext(file_path)[1].lower()

    # 读取文件
    if ext == '.xls':
        import xlrd
        wb = xlrd.open_workbook(file_path)
        sheets = []
        for sn in wb.sheet_names():
            ws = wb.sheet_by_name(sn)
            sheets.append((sn, [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]))
    else:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets = [(sn, [list(row) for row in wb[sn].iter_rows(values_only=True)]) for sn in wb.sheetnames]

    if not sheets:
        return {"ok": False, "error": "工作簿无工作表"}

    sheet_name, rows = sheets[0]
    bank_name = detect_bank(rows)

    # 查找表头
    header_idx = find_header(rows, bank_name)
    if header_idx is None:
        return {"ok": False, "error": "无法识别表头行", "bank_detected": bank_name}

    header = [str(c) if c else '' for c in rows[header_idx]]
    mapping = map_columns(header, bank_name)

    # 查找数据起始行
    data_start = header_idx + 1
    while data_start < len(rows) and all(c is None or str(c).strip() == '' for c in rows[data_start]):
        data_start += 1

    # 解析交易
    transactions = []
    for i in range(data_start, len(rows)):
        row = rows[i]
        if not row: continue
        if all(c is None or str(c).strip() == '' for c in row): continue

        first = str(row[0] or '').strip()
        if re.search(r'合计|小计|总计|Total', first): continue

        # 提取日期
        date_col = mapping.get('date')
        date_str = str(row[date_col]).strip() if date_col and date_col < len(row) and row[date_col] else ''
        if re.match(r'\d{8}', date_str):
            date_str = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"

        # 提取金额和方向
        direction = None
        amt_in = 0.0
        amt_out = 0.0

        txn_type_col = mapping.get('txn_type')
        signed_col = mapping.get('signed_amt')
        debit_col = mapping.get('debit_amt')
        credit_col = mapping.get('credit_amt')

        if txn_type_col and txn_type_col < len(row):
            txn_type = str(row[txn_type_col] or '')
            if '来' in txn_type:
                direction = '收入'
            elif '往' in txn_type:
                direction = '支出'

        if signed_col and signed_col < len(row):
            val = row[signed_col]
            if isinstance(val, (int, float)):
                if val > 0:
                    amt_in = float(val)
                    direction = direction or '收入'
                elif val < 0:
                    amt_out = abs(float(val))
                    direction = direction or '支出'

        if debit_col and debit_col < len(row):
            dv = row[debit_col]
            if dv and isinstance(dv, (int, float)) and dv > 0:
                amt_in = float(dv)
                direction = direction or '收入'
        if credit_col and credit_col < len(row):
            cv = row[credit_col]
            if cv and isinstance(cv, (int, float)) and cv > 0:
                amt_out = float(cv)
                direction = direction or '支出'

        # 提取余额
        bal_col = mapping.get('balance')
        balance = None
        if bal_col and bal_col < len(row):
            bv = row[bal_col]
            if isinstance(bv, (int, float)):
                balance = float(bv)

        # 提取对方
        cp_name = ''
        cp_acct = ''
        if direction == '收入':
            pn = mapping.get('payer_name')
            pa = mapping.get('payer_acct')
            cp_name = str(row[pn]).strip() if pn and pn < len(row) and row[pn] else ''
            cp_acct = str(row[pa]).strip() if pa and pa < len(row) and row[pa] else ''
        elif direction == '支出':
            pn = mapping.get('payee_name')
            pa = mapping.get('payee_acct')
            cp_name = str(row[pn]).strip() if pn and pn < len(row) and row[pn] else ''
            cp_acct = str(row[pa]).strip() if pa and pa < len(row) and row[pa] else ''

        # 生成摘要（注：internal_entities 需从数据库动态加载）
        summary = build_summary(row, mapping, {}, direction, bank_name)

        transactions.append({
            "_row": i + 1,
            "business_date": date_str,
            "direction": direction or '',
            "summary": summary,
            "counterparty": cp_name,
            "amount_in": amt_in,
            "amount_out": amt_out,
            "balance": balance,
            "_raw": {
                "biz_type": str(row[mapping.get('biz_type')]) if mapping.get('biz_type') and mapping.get('biz_type') < len(row) else '',
                "payer_name": cp_name if direction == '收入' else '',
                "payee_name": cp_name if direction == '支出' else '',
                "counterparty_acct": cp_acct,
                "reference": str(row[mapping.get('reference')]) if mapping.get('reference') and mapping.get('reference') < len(row) else '',
                "purpose": str(row[mapping.get('purpose')]) if mapping.get('purpose') and mapping.get('purpose') < len(row) else '',
                "remark": str(row[mapping.get('remark')]) if mapping.get('remark') and mapping.get('remark') < len(row) else '',
            },
        })

    # 余额校验
    balance_issues = []
    if transactions and transactions[0]['balance'] is not None:
        for k in range(1, len(transactions)):
            pb = transactions[k-1]['balance']
            cb = transactions[k]['balance']
            ci = transactions[k]['amount_in']
            co = transactions[k]['amount_out']
            if pb is not None and cb is not None:
                expected = pb + ci - co
                if abs(expected - cb) > 0.02:
                    balance_issues.append({'row': k+1, 'expected': round(expected,2), 'actual': cb})

    return {
        "ok": True,
        "file_name": os.path.basename(file_path),
        "bank_detected": bank_name,
        "file_format": ext,
        "header_row": header_idx,
        "data_start": data_start,
        "transaction_count": len(transactions),
        "transactions": transactions,
        "balance_check": {"ok": len(balance_issues) == 0, "issues": balance_issues},
        "mapping_used": {k: str(v) for k, v in mapping.items()},
    }


# ===== 参考示例：中国银行 =====
# 见 skills/parse_boc/run.py 中的完整实现
# 该文件包含养护集团体系50家内部实体的完整摘要生成规则
