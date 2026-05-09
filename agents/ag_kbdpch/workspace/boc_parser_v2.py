#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
中国银行企业网银流水解析器 - 标准摘要版
按照出纳习惯生成标准化摘要，完全兼容原始规格。
"""

import json, re, sys, os
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# 工具函数
# ============================================================================

def load_config(config_path: str) -> Dict:
    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def contains_chinese(text: str) -> bool:
    if not text:
        return False
    return bool(re.search(r'[\u4e00-\u9fff]', str(text)))


def simplify_counterparty_name(full_name: str) -> str:
    """将对方全称简化为不超过15个字符的简称"""
    if not full_name:
        return ''
    result = full_name
    # 去除前缀
    for prefix in ['山西', '太原市', '阳曲县', '太原']:
        if result.startswith(prefix):
            result = result[len(prefix):]
            break
    # 去除后缀（按长度从长到短）
    for suffix in ['道路建设养护集团有限公司', '集团有限公司', '有限责任公司',
                   '有限公司', '经营部', '合伙企业']:
        if result.endswith(suffix):
            result = result[:-len(suffix)]
            break
    if not result:
        result = full_name
    if len(result) > 15:
        result = result[:15]
    return result


def parse_amount(amount_str) -> Optional[float]:
    if not amount_str or str(amount_str).strip() == '':
        return None
    try:
        cleaned = str(amount_str).replace(',', '').strip()
        return float(cleaned)
    except ValueError:
        return None


def parse_date(date_str) -> str:
    if not date_str:
        return ''
    s = str(date_str).strip()
    # Excel序列号
    if s.isdigit() and len(s) <= 5:
        try:
            from datetime import datetime, timedelta
            d = datetime(1899, 12, 30) + timedelta(days=int(s))
            return d.strftime('%Y-%m-%d')
        except:
            pass
    # 尝试各种格式
    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y%m%d', '%d/%m/%Y']:
        try:
            from datetime import datetime
            d = datetime.strptime(s[:10], fmt)
            return d.strftime('%Y-%m-%d')
        except:
            continue
    return s[:10] if len(s) >= 10 else s


# ============================================================================
# 表头识别与字段映射
# ============================================================================

def find_header_row(df: pd.DataFrame) -> int:
    """扫描DataFrame的前N行，找到表头行"""
    for idx in range(min(20, len(df))):
        row_text = ' '.join([str(v) for v in df.iloc[idx].values if pd.notna(v)])
        if '交易日期' in row_text and '交易金额' in row_text:
            return idx
    raise ValueError("无法找到表头行，表头行必须同时包含'交易日期'和'交易金额'")


def build_field_mapping(header_row) -> Dict[str, int]:
    """根据表头行构建字段名到列索引的映射"""
    mapping = {}
    for col_idx, field_name in enumerate(header_row):
        if pd.isna(field_name):
            continue
        if '[' in str(field_name) and ']' in str(field_name):
            cn_name = str(field_name).split('[')[0].strip()
        else:
            cn_name = str(field_name).strip()
        if cn_name:
            mapping[cn_name] = col_idx
    return mapping


def get_cell_value(row_data, field_mapping: Dict[str, int], field_name: str) -> str:
    """根据字段名从一行数据中获取值"""
    if field_name not in field_mapping:
        return ''
    col_idx = field_mapping[field_name]
    if col_idx >= len(row_data):
        return ''
    value = row_data.iloc[col_idx]
    if pd.isna(value):
        return ''
    return str(value).strip()


# ============================================================================
# 摘要生成 - 完全按规格实现
# ============================================================================

SUMMARY_RULES = [
    # 第1级：精确匹配摘要规则（银行收费/利息/贷款等）
    {"biz_type": "短信收费",  "summary": "付银行短信服务费",  "counterparty_clear": True},
    {"biz_type": "结息",      "summary": "存款利息收入",      "counterparty_clear": True},
    {"biz_type": "贷款放款",   "summary": "收到银行贷款放款",  "counterparty_clear": True},
    {"biz_type": "贷款还款",   "summary": "偿还银行贷款",      "counterparty_clear": False},
    {"biz_type": "收费",      "keyword": "手续费", "summary": "付银行跨行转账手续费", "counterparty_clear": True},
    {"biz_type": "收费",      "keyword": "邮费",  "summary": "付银行邮递费",          "counterparty_clear": True},
    {"biz_type": "收费",      "keyword": "账户管理", "summary": "付银行账户管理费（年费）", "counterparty_clear": True},
    {"biz_type": "收费",      "keyword": "询证函", "summary": "付银行询证函及资信证明费",  "counterparty_clear": True},
]

INTERNAL_ACCOUNTS = {
    "139252780563": "本公司",  # 本账号
    "山西喜跃发道路建设养护集团有限公司": "本公司",  # 本公司名称
}

BANK_BENEFICIARY_NAMES = [
    "中国银行", "农业银行", "工商银行", "建设银行", "交通银行",
    "招商银行", "浦发银行", "光大银行", "民生银行"
]


def generate_summary_and_counterparty(
    row: Dict,
    our_company_name: str = "山西喜跃发道路建设养护集团有限公司"
) -> Tuple[str, str, str]:
    """
    按优先级生成标准化摘要和对方名称。
    
    优先级：
    1. 摘要规则精确匹配（业务类型/关键词）
    2. 内部划转（来账/往账且对方是本公司）
    3. 来账（外部收入）
    4. 往账（外部支出）
    5. 兜底
    
    Returns: (摘要, 对方名称, 状态)
    """
    txn_type = row.get('交易类型', '')
    biz_type = row.get('业务类型', '')
    payer_name = row.get('付款人名称', '')
    payee_name = row.get('收款人名称', '')
    purpose = row.get('用途', '') or ''
    remark_field = row.get('交易附言', '') or ''
    payee_bank = row.get('收款人开户行名', '') or ''
    payer_bank = row.get('付款人开户行名', '') or ''
    
    # 对方名称原始值（根据方向决定）
    counterparty_raw = payer_name if txn_type == '来账' else payee_name
    
    # 判断是否内部
    is_internal = (
        counterparty_raw == our_company_name or
        payer_name == our_company_name and payee_name == our_company_name
    )
    
    # === 第1级：摘要规则匹配 ===
    for rule in SUMMARY_RULES:
        if rule.get("biz_type") == biz_type:
            keyword = rule.get("keyword", "")
            if not keyword or keyword in purpose or keyword in remark_field:
                counterparty_out = '' if rule.get("counterparty_clear", False) else counterparty_raw
                return rule["summary"], counterparty_out, "正常"
    
    # === 第2级：内部划转 ===
    if is_internal:
        return "内部资金划转", "本公司", "正常"
    
    # === 第3级：来账（外部收入）===
    if txn_type == '来账':
        if payer_name and payer_name != our_company_name:
            short_name = simplify_counterparty_name(payer_name)
            # 特殊处理：从农业银行转入
            if any(b in payer_bank for b in ["农业银行", "农行"]):
                return f"收{short_name}款项", payer_name, "正常"
            return f"收{short_name}款项", payer_name, "正常"
        # 来账但无付款人：可能是结息
        return f"收到款项", counterparty_raw, "正常"
    
    # === 第4级：往账（外部支出）===
    if txn_type == '往账':
        if payee_name and payee_name != our_company_name:
            short_name = simplify_counterparty_name(payee_name)
            # 特殊处理：转农行阳曲支行
            if any(b in payee_bank for b in ["农业银行", "农行"]):
                return f"付{short_name}款项", payee_name, "正常"
            return f"付{short_name}款项", payee_name, "正常"
        return f"付出款项", counterparty_raw, "正常"
    
    # === 第5级：兜底 ===
    return "收到款项" if txn_type == '来账' else "付出款项", counterparty_raw, "正常"


# ============================================================================
# 异常检测
# ============================================================================

def detect_anomalies(df: pd.DataFrame, our_company_name: str, min_large: float = 500000) -> set:
    """
    异常检测：
    A. 同日大额收付配对（同额同日一收一付，标记为"待确认"）
    B. 大额无用途往账（大额往账但用途/附言均为空或无中文，标记"待确认"）
    """
    suspicious = set()
    
    # === 规则A：同日大额收付配对 ===
    df_sorted = df.sort_values('交易日期')
    date_groups = df_sorted.groupby('交易日期')
    
    for date_val, day_group in date_groups:
        if len(day_group) < 2:
            continue
        incomes = day_group[day_group['交易类型'] == '来账']
        expenses = day_group[day_group['交易类型'] == '往账']
        
        for _, inc in incomes.iterrows():
            if inc['收入金额'] < min_large:
                continue
            for _, exp in expenses.iterrows():
                if exp['支出金额'] < min_large:
                    continue
                # 金额误差 < 1%
                if inc['收入金额'] > 0:
                    diff = abs(inc['收入金额'] - exp['支出金额']) / inc['收入金额']
                    if diff < 0.01:
                        suspicious.add(inc.name)
                        suspicious.add(exp.name)
    
    # === 规则B：大额无用途往账 ===
    for idx, row in df.iterrows():
        if row['交易类型'] != '往账':
            continue
        if row.get('对方名称', '') == '本公司':
            continue
        if row['支出金额'] < min_large:
            continue
        purpose = str(row.get('用途', '') or '')
        remark = str(row.get('交易附言', '') or '')
        # 用途和附言都无中文
        if not contains_chinese(purpose) and not contains_chinese(remark):
            suspicious.add(idx)
    
    return suspicious


# ============================================================================
# 主解析函数
# ============================================================================

def parse_boc流水(input_file: str, config_path: str = None, our_company_name: str = "山西喜跃发道路建设养护集团有限公司") -> pd.DataFrame:
    """解析中行流水，返回标准化DataFrame"""
    
    # 读取文件（支持xls和xlsx）
    if input_file.endswith('.xls'):
        import xlrd
        wb = xlrd.open_workbook(input_file)
        ws = wb.sheet_by_index(0)
        all_rows = []
        for r in range(ws.nrows):
            all_rows.append([ws.cell_value(r, c) for c in range(ws.ncols)])
        df_raw = pd.DataFrame(all_rows)
    else:
        df_raw = pd.read_excel(input_file, header=None, engine='openpyxl')
    
    # 找表头行
    header_row_idx = find_header_row(df_raw)
    
    # 以表头行作为列名
    df = df_raw.iloc[header_row_idx:].copy()
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    
    # 构建字段映射
    field_map = build_field_mapping(df.columns)
    
    # 提取标准字段
    records = []
    for _, row_data in df.iterrows():
        txn_type = get_cell_value(row_data, field_map, '交易类型')
        if not txn_type or txn_type not in ['来账', '往账']:
            continue
        
        raw_record = {
            '交易类型': txn_type,
            '业务类型': get_cell_value(row_data, field_map, '业务类型'),
            '付款人名称': get_cell_value(row_data, field_map, '付款人名称'),
            '付款人账号': get_cell_value(row_data, field_map, '付款人账号'),
            '付款人开户行名': get_cell_value(row_data, field_map, '付款人开户行名'),
            '收款人名称': get_cell_value(row_data, field_map, '收款人名称'),
            '收款人账号': get_cell_value(row_data, field_map, '收款人账号'),
            '收款人开户行名': get_cell_value(row_data, field_map, '收款人开户行名'),
            '交易日期': get_cell_value(row_data, field_map, '交易日期'),
            '交易时间': get_cell_value(row_data, field_map, '交易时间'),
            '交易金额': get_cell_value(row_data, field_map, '交易金额'),
            '交易后余额': get_cell_value(row_data, field_map, '交易后余额'),
            '摘要': get_cell_value(row_data, field_map, '摘要'),
            '用途': get_cell_value(row_data, field_map, '用途'),
            '交易附言': get_cell_value(row_data, field_map, '交易附言'),
            '交易流水号': get_cell_value(row_data, field_map, '交易流水号'),
        }
        
        # 解析金额
        amt = parse_amount(raw_record['交易金额'])
        balance = parse_amount(raw_record['交易后余额'])
        
        if amt is None:
            continue
        
        # 计算标准化金额和方向
        if txn_type == '来账':
            income = abs(amt)
            expense = 0.0
        else:
            income = 0.0
            expense = abs(amt)
        
        std_record = {
            '交易日期_raw': raw_record['交易日期'],
            '交易类型': txn_type,
            '业务类型': raw_record['业务类型'],
            '收入金额': income,
            '支出金额': expense,
            '滚动余额': balance or 0.0,
            '付款人名称': raw_record['付款人名称'],
            '收款人名称': raw_record['收款人名称'],
            '付款人开户行名': raw_record['付款人开户行名'],
            '收款人开户行名': raw_record['收款人开户行名'],
            '原始摘要': raw_record['摘要'],
            '用途': raw_record['用途'],
            '交易附言': raw_record['交易附言'],
            '交易流水号': raw_record['交易流水号'],
        }
        records.append(std_record)
    
    df_result = pd.DataFrame(records)
    
    # 生成标准化摘要
    summaries = []
    for _, row in df_result.iterrows():
        s, c, st = generate_summary_and_counterparty(row, our_company_name)
        summaries.append({'摘要': s, '对方名称': c, '状态': st})
    
    summary_df = pd.DataFrame(summaries)
    df_result = pd.concat([df_result, summary_df], axis=1)
    
    # 解析日期
    df_result['交易日期'] = df_result['交易日期_raw'].apply(parse_date)
    
    # 异常检测
    suspicious = detect_anomalies(df_result, our_company_name)
    for idx in suspicious:
        if idx in df_result.index:
            df_result.loc[idx, '状态'] = '待确认'
    
    # 输出列排序
    output_cols = ['交易日期', '交易类型', '业务类型', '收入金额', '支出金额', '滚动余额',
                   '摘要', '对方名称', '状态', '付款人名称', '收款人名称', 
                   '付款人开户行名', '收款人开户行名', '原始摘要', '用途', '交易附言', '交易流水号']
    
    return df_result[[c for c in output_cols if c in df_result.columns]]


def write_excel(df: pd.DataFrame, output_path: str):
    """将结果写入格式化的Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "基础数据表"
    
    # 表头
    output_headers = ['序号', '状态', '交易日期', '交易方向', '业务类型', 
                      '收入金额', '支出金额', '滚动余额', '摘要', '对方名称', '流水号']
    ws.append(output_headers)
    
    # 数据行
    for i, (_, row) in enumerate(df.iterrows(), 1):
        ws.append([
            i,
            row.get('状态', '正常'),
            row['交易日期'],
            row['交易类型'],
            row.get('业务类型', ''),
            row['收入金额'],
            row['支出金额'],
            row['滚动余额'],
            row['摘要'],
            row.get('对方名称', ''),
            row.get('交易流水号', ''),
        ])
    
    # 样式
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = Font(name='微软雅黑', bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    for row in ws.iter_rows(min_row=2):
        state = row[1].value
        counterparty = row[9].value
        for cell in row:
            cell.font = Font(name='微软雅黑', size=10)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        if state == '待确认':
            for cell in row:
                cell.fill = red_fill
        elif counterparty == '本公司':
            for cell in row:
                cell.fill = yellow_fill
        for col_idx in [5, 6, 7]:
            row[col_idx-1].alignment = Alignment(horizontal='right', vertical='center')
        row[8].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    col_widths = [6, 8, 12, 10, 12, 14, 14, 14, 30, 26, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    ws.freeze_panes = 'A2'
    wb.save(output_path)
    print(f"✓ 已输出: {output_path}")
    return output_path


if __name__ == '__main__':
    input_file = 'workspace/inbox/中行银行流水.xls'
    output_file = 'workspace/outputs/zh0008_基础数据表_标准摘要.xlsx'
    
    df = parse_boc流水(input_file)
    
    # 统计
    print(f"\n解析结果：共 {len(df)} 条")
    print(f"  来账: {(df['交易类型']=='来账').sum()}, 往账: {(df['交易类型']=='往账').sum()}")
    print(f"  待确认: {(df['状态']=='待确认').sum()}, 正常: {(df['状态']=='正常').sum()}")
    print(f"  日期范围: {df['交易日期'].min()} ~ {df['交易日期'].max()}")
    print(f"  总收入: {df['收入金额'].sum():,.2f}")
    print(f"  总支出: {df['支出金额'].sum():,.2f}")
    
    write_excel(df, output_file)
    
    # 输出摘要预览
    print("\n摘要预览（前10条）:")
    for i, (_, row) in enumerate(df.head(10).iterrows(), 1):
        print(f"  [{i:2}] {row['交易日期']} {row['交易类型']:2} {row['收入金额']:>15,.2f}/{row['支出金额']:>15,.2f} | {row['摘要'][:25]:25} | {row['对方名称'][:18]:18} | {row['状态']}")