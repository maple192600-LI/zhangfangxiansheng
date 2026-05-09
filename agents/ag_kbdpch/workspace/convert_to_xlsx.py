import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 读取.xls
wb_in = xlrd.open_workbook('workspace/inbox/中行银行流水.xls')
ws_in = wb_in.sheet_by_index(0)

# 创建.xlsx
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "基础数据表"

# 写入表头
headers = ["序号", "状态", "交易日期", "交易方向", "业务类型", "收入金额", "支出金额", "滚动余额", "摘要", "对方名称", "流水号"]
ws_out.append(headers)

# 读取预览文件写入数据
import csv
with open('zh0008_preview.tsv', 'r', encoding='utf-8') as f:
    reader = csv.reader(f, delimiter='\t')
    header = next(reader)  # 跳过表头
    for row in reader:
        ws_out.append(row)

# 样式：待确认行标红
red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
for row in ws_out.iter_rows(min_row=2):
    if row[1].value == '待确认':
        for cell in row:
            cell.fill = red_fill

wb_out.save('workspace/outputs/zh0008_基础数据表.xlsx')
print("已生成: workspace/outputs/zh0008_基础数据表.xlsx")