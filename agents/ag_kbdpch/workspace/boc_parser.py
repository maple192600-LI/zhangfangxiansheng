#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
中国银行企业网银流水解析器
读取中行标准导出的XLS/XLSX文件，生成基础数据表Excel文件。

红线约束：
1. 禁止硬编码列索引，所有字段通过表头名称动态定位
2. 摘要禁止包含金额数字、金额单位
3. 摘要禁止包含敏感词
4. 可疑交易通过"状态"列标记为"待确认"，不在摘要中定性
5. 禁止在摘要中猜测业务性质
6. 单位编码和账户编码必须从外部配置获取
"""

import json
import re
import sys
import os
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================================
# 配置加载
# ============================================================================

def load_config(config_path: str) -> Dict:
    """读取JSON配置文件"""
    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)


# ============================================================================
# 表头识别与字段映射
# ============================================================================

def find_header_row(df: pd.DataFrame) -> int:
    """
    扫描DataFrame找到表头行。
    判定标准：该行同时包含'交易日期'和'交易金额'。
    """
    for idx in range(min(20, len(df))):
        row_text = ' '.join([str(v) for v in df.iloc[idx].values if pd.notna(v)])
        if '交易日期' in row_text and '交易金额' in row_text:
            return idx
    raise ValueError("无法找到表头行")


def merge_multirow_header(df: pd.DataFrame, header_row_idx: int) -> List[str]:
    """合并多行表头，处理中行某些版本拆成2行的表头"""
    max_merge_rows = 3
    end_idx = min(header_row_idx + max_merge_rows, len(df))
    
    header_rows = []
    for i in range(header_row_idx, end_idx):
        row = df.iloc[i].fillna('').astype(str).tolist()
        has_header = any('[' in str(v) and ']' in str(v) for v in row)
        if has_header:
            header_rows.append(row)
        else:
            break
    
    merged = [''] * len(df.columns)
    for row in header_rows:
        for col_idx, cell_value in enumerate(row):
            cell_str = str(cell_value).strip()
            if cell_str and cell_str != 'nan':
                merged[col_idx] = cell_str
    
    return [v.strip() for v in merged]


def build_field_mapping(header_row: List[str]) -> Dict[str, int]:
    """
    根据表头构建字段名到列索引的映射。
    字段名格式为"中文名[ 英文名 ]"，取中文名作为键。
    """
    mapping = {}
    for col_idx, field_name in enumerate(header_row):
        if '[' in field_name and ']' in field_name:
            cn_name = field_name.split('[')[0].strip()
        else:
            cn_name = field_name.strip()
        
        if cn_name:
            mapping[cn_name] = col_idx
    
    return mapping


def get_cell_value(row_data: pd.Series, field_mapping: Dict[str, int], 
                   field_name: str) -> str:
    """根据字段名从一行数据中获取值，字段不存在时返回空字符串"""
    if field_name not in field_mapping:
        return ''
    col_idx = field_mapping[field_name]
    value = row_data.iloc[col_idx] if col_idx < len(row_data) else None
    if pd.isna(value):
        return ''
    return str(value).strip()


# ============================================================================
# 对方名称简化
# ============================================================================

def simplify_counterparty_name(full_name: str) -> str:
    """
    将对方全称简化为不超过15字符的简称。
    去除前缀：山西、太原市、阳曲县
    去除后缀：集团有限公司、有限公司等
    """
    if not full_name:
        return ''
    
    result = full_name
    
    prefixes = ['山西', '太原市', '阳曲县']
    for prefix in prefixes:
        if result.startswith(prefix):
            result = result[len(prefix):]
            break
    
    suffixes = [
        '道路建设养护集团有限公司',
        '建设养护集团有限公司',
        '集团有限公司',
        '有限责任公司',
        '有限公司',
        '经营部',
        '合伙企业'
    ]
    for suffix in suffixes:
        if result.endswith(suffix):
            result = result[:-len(suffix)]
            break
    
    if not result:
        result = full_name
    
    return result[:15]


def extract_bank_short_name(bank_full_name: str) -> str:
    """从银行全称提取简称"""
    if not bank_full_name:
        return ''
    
    bank_mappings = {
        '中国银行': '中行',
        '中国农业银行': '农行',
        '中国工商银行': '工行',
        '中国建设银行': '建行',
        '中国交通银行': '交行',
        '招商银行': '招行',
    }
    
    for full_name, short_name in bank_mappings.items():
        if full_name in bank_full_name:
            return short_name
    
    return bank_full_name[:4] if len(bank_full_name) > 4 else bank_full_name


# ============================================================================
# 摘要生成
# ============================================================================

def generate_summary(vars_dict: Dict[str, Any], 
                     summary_rules: List[Dict]) -> Tuple[str, str, str]:
    """
    根据规则生成摘要和对方名称。
    优先级：配置精确匹配 > 内部划转 > 来账 > 往账 > 兜底
    """
    txn_type = vars_dict['交易类型']
    biz_type = vars_dict['业务类型']
    counterparty_full = vars_dict['对方全称']
    is_internal = vars_dict['是否内部']
    purpose = vars_dict['用途']
    remark = vars_dict['交易附言']
    original_summary = vars_dict['摘要']
    
    # 第1级：配置文件精确匹配
    for rule in summary_rules:
        match = rule.get('匹配条件', {})
        field = match.get('字段')
        value = match.get('值')
        sub_field = match.get('子字段')
        contains = match.get('包含')
        
        if field == '业务类型':
            if vars_dict.get('业务类型') != value:
                continue
        else:
            continue
        
        if sub_field and contains:
            sub_value = ''
            if sub_field == '摘要':
                sub_value = original_summary
            elif sub_field == '用途':
                sub_value = purpose
            elif sub_field == '交易附言':
                sub_value = remark
            
            if contains not in sub_value:
                continue
        
        config_summary = rule['摘要']
        
        if biz_type in ['短信收费', '结息'] or '收费' in biz_type:
            return config_summary, '', '正常'
        
        return config_summary, counterparty_full, '正常'
    
    # 第2级：内部划转
    if txn_type == '往账' and is_internal:
        payee_bank = vars_dict.get('收款人开户行', '')
        bank_short = extract_bank_short_name(payee_bank)
        summary = f'内部资金划转至{bank_short}' if bank_short else '内部资金划转'
        return summary, '本公司', '正常'
    
    # 第3级：来账
    if txn_type == '来账' and not is_internal:
        counterparty_short = simplify_counterparty_name(counterparty_full)
        return f'收{counterparty_short}款项', counterparty_full, '正常'
    
    # 第4级：往账
    if txn_type == '往账' and not is_internal:
        counterparty_short = simplify_counterparty_name(counterparty_full)
        return f'付{counterparty_short}款项', counterparty_full, '正常'
    
    # 第5级：兜底
    summary = '收到款项' if txn_type == '来账' else ('付出款项' if txn_type == '往账' else '未知交易')
    return summary, counterparty_full, '正常'


# ============================================================================
# 异常检测
# ============================================================================

def check_same_day_paired(df: pd.DataFrame, min_amount: float, 
                          error_ratio: float) -> set:
    """同日收付配对检测"""
    suspicious_indices = set()
    
    grouped = df.groupby('交易日期')
    
    for date, group in grouped:
        incomes = group[(group['交易类型'] == '来账') & (group['交易金额'] >= min_amount)]
        expenses = group[(group['交易类型'] == '往账') & (group['交易金额'].abs() >= min_amount)]
        
        for _, income_row in incomes.iterrows():
            for _, expense_row in expenses.iterrows():
                income_amount = income_row['交易金额']
                expense_amount = abs(expense_row['交易金额'])
                
                if income_amount > 0:
                    diff_ratio = abs(income_amount - expense_amount) / income_amount
                    if diff_ratio <= error_ratio:
                        suspicious_indices.add(income_row.name)
                        suspicious_indices.add(expense_row.name)
    
    return suspicious_indices


def check_large_expense_no_purpose(df: pd.DataFrame, min_amount: float) -> set:
    """大额无用途往账检测"""
    suspicious_indices = set()
    
    def contains_chinese(text: str) -> bool:
        if not text:
            return False
        return bool(re.search(r'[\u4e00-\u9fff]', text))
    
    for idx, row in df.iterrows():
        if row['交易类型'] != '往账':
            continue
        if row.get('是否内部', False):
            continue
        if abs(row['交易金额']) < min_amount:
            continue
        
        purpose = row.get('用途', '') or ''
        remark = row.get('交易附言', '') or ''
        if contains_chinese(purpose) or contains_chinese(remark):
            continue
        
        biz_type = row.get('业务类型', '') or ''
        if biz_type in ['贷款还款', '贷款放款']:
            continue
        
        suspicious_indices.add(idx)
    
    return suspicious_indices


def validate_balance(df: pd.DataFrame) -> List[str]:
    """验证滚动余额的连续性"""
    warnings = []
    
    df_sorted = df.reset_index(drop=True)
    prev_balance = None
    
    for idx in range(len(df_sorted)):
        row = df_sorted.iloc[idx]
        txn_amount = row['交易金额']
        current_balance = row['滚动余额']
        
        if prev_balance is not None:
            expected_balance = prev_balance + txn_amount
            if abs(current_balance - expected_balance) > 0.01:
                warnings.append(
                    f"第{idx+1}行余额校验失败：期望={expected_balance:.2f}，实际={current_balance:.2f}"
                )
        
        prev_balance = current_balance
    
    return warnings


# ============================================================================
# 数据处理主流程
# ============================================================================

def parse_boc_statement(input_file: str, config_path: str) -> pd.DataFrame:
    """解析中行流水文件"""
    print(f"正在读取输入文件: {input_file}")
    
    df_raw = pd.read_excel(input_file, header=None, engine='openpyxl')
    print(f"原始数据行数: {len(df_raw)}")
    
    config = load_config(config_path)
    account_mapping = config.get('账户映射', [])
    summary_rules = config.get('摘要规则', [])
    anomaly_config = config.get('异常检测', {})
    
    header_row_idx = find_header_row(df_raw)
    print(f"表头行位置: 第{header_row_idx + 1}行")
    
    # 提取账号
    account_number = ''
    for idx in range(header_row_idx):
        row_text = str(df_raw.iloc[idx, 0]) if idx < len(df_raw) else ''
        if '查询账号' in row_text:
            if len(df_raw.columns) > 1:
                account_number = str(df_raw.iloc[idx, 1]).strip()
            break
    
    print(f"流水账号: {account_number}")
    
    # 匹配账户配置
    matched_account = None
    for acc in account_mapping:
        if acc.get('流水账号') == account_number:
            matched_account = acc
            break
    
    if matched_account:
        entity_code = matched_account['单位编码']
        account_code = matched_account['账户编码']
        entity_name = matched_account['单位名称']
        account_name = matched_account['账户名称']
        print(f"匹配到账户: {entity_name}")
    else:
        entity_code = 'DW待匹配'
        account_code = 'ZH待匹配'
        entity_name = '待匹配'
        account_name = '待匹配'
        print(f"警告: 账号 {account_number} 未匹配！")
    
    # 构建字段映射
    header_row = merge_multirow_header(df_raw, header_row_idx)
    field_mapping = build_field_mapping(header_row)
    print(f"识别的字段: {len(field_mapping)}个")
    
    data_start_idx = header_row_idx + 1
    records = []
    skip_count = 0
    
    for idx in range(data_start_idx, len(df_raw)):
        row_data = df_raw.iloc[idx]
        txn_type = get_cell_value(row_data, field_mapping, '交易类型')
        
        if txn_type not in ['往账', '来账']:
            if any(str(v).strip() for v in row_data.values if pd.notna(v)):
                skip_count += 1
            continue
        
        biz_date = get_cell_value(row_data, field_mapping, '交易日期')
        biz_time = get_cell_value(row_data, field_mapping, '交易时间')
        biz_type = get_cell_value(row_data, field_mapping, '业务类型')
        payer_name = get_cell_value(row_data, field_mapping, '付款人名称')
        payee_name = get_cell_value(row_data, field_mapping, '收款人名称')
        purpose = get_cell_value(row_data, field_mapping, '用途')
        remark_field = get_cell_value(row_data, field_mapping, '交易附言')
        remarks = get_cell_value(row_data, field_mapping, '备注')
        original_summary = get_cell_value(row_data, field_mapping, '摘要')
        balance_str = get_cell_value(row_data, field_mapping, '交易后余额')
        txn_ref = get_cell_value(row_data, field_mapping, '交易流水号')
        voucher_no = get_cell_value(row_data, field_mapping, '凭证号码')
        payee_bank = get_cell_value(row_data, field_mapping, '收款人开户行名')
        
        # 金额解析
        amount_str = get_cell_value(row_data, field_mapping, '交易金额')
        try:
            amount = float(amount_str)
        except ValueError:
            amount = 0.0
        
        # 确定方向
        if txn_type == '来账':
            income_amount = abs(amount)
            expense_amount = 0.0
            counterparty_full = payer_name
        else:
            income_amount = 0.0
            expense_amount = abs(amount)
            counterparty_full = payee_name
        
        is_internal = (counterparty_full == entity_name)
        
        # 生成摘要
        vars_dict = {
            '交易类型': txn_type,
            '业务类型': biz_type,
            '对方全称': counterparty_full,
            '是否内部': is_internal,
            '用途': purpose,
            '交易附言': remark_field,
            '摘要': original_summary,
            '收款人开户行': payee_bank,
        }
        
        summary, counterparty_name, status = generate_summary(vars_dict, summary_rules)
        
        # 余额
        try:
            balance = float(balance_str) if balance_str else 0.0
        except ValueError:
            balance = 0.0
        
        records.append({
            '序号': len(records) + 1,
            '交易日期': biz_date,
            '交易时间': biz_time,
            '交易类型': txn_type,
            '业务类型': biz_type,
            '对方名称': counterparty_name,
            '收入金额': income_amount,
            '支出金额': expense_amount,
            '滚动余额': balance,
            '摘要': summary,
            '状态': status,
            '单位编码': entity_code,
            '账户编码': account_code,
            '单位名称': entity_name,
            '账户名称': account_name,
            '交易流水号': txn_ref,
            '凭证号码': voucher_no,
            '付款人名称': payer_name,
            '收款人名称': payee_name,
            '收款人开户行': payee_bank,
            '用途': purpose,
            '交易附言': remark_field,
            '备注': remarks,
            '银行原始摘要': original_summary,
            '是否内部': is_internal
        })
    
    print(f"有效数据: {len(records)}行, 跳过: {skip_count}行")
    
    df = pd.DataFrame(records)
    
    # 异常检测
    anomaly_rules = anomaly_config.get('异常检测', {})
    
    paired_config = anomaly_rules.get('同日收付配对', {})
    if paired_config.get('启用', False):
        paired_indices = check_same_day_paired(
            df, 
            paired_config.get('最小金额', 1000000),
            paired_config.get('金额匹配误差比例', 0.01)
        )
        for idx in paired_indices:
            df.loc[idx, '状态'] = '待确认'
    
    large_config = anomaly_rules.get('大额无用途往账', {})
    if large_config.get('启用', False):
        large_indices = check_large_expense_no_purpose(
            df,
            large_config.get('最小金额', 1000000)
        )
        for idx in large_indices:
            df.loc[idx, '状态'] = '待确认'
    
    # 余额校验
    for warning in validate_balance(df):
        print(f"警告: {warning}")
    
    return df


# ============================================================================
# 输出Excel
# ============================================================================

def write_excel_output(df: pd.DataFrame, output_path: str):
    """写入Excel文件"""
    sheet1_columns = [
        '序号', '交易日期', '交易时间', '交易类型', '业务类型',
        '对方名称', '收入金额', '支出金额', '滚动余额',
        '摘要', '状态', '单位编码', '账户编码'
    ]
    
    sheet2_columns = sheet1_columns + [
        '交易流水号', '凭证号码', '付款人名称', '收款人名称',
        '收款人开户行', '用途', '交易附言', '备注', '银行原始摘要'
    ]
    
    wb = openpyxl.Workbook()
    
    # Sheet1
    ws1 = wb.active
    ws1.title = '基础数据表'
    
    for col_idx, col_name in enumerate(sheet1_columns, 1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    for row_idx, row in df.iterrows():
        for col_idx, col_name in enumerate(sheet1_columns, 1):
            value = row.get(col_name, '')
            if col_name in ['收入金额', '支出金额', '滚动余额'] and pd.notna(value) and value != '':
                value = round(float(value), 2)
            cell = ws1.cell(row=row_idx + 2, column=col_idx, value=value)
            if col_name == '状态' and value == '待确认':
                cell.font = Font(color='FF0000')
    
    for col_idx in range(1, len(sheet1_columns) + 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # Sheet2
    ws2 = wb.create_sheet(title='辅助核对明细')
    
    for col_idx, col_name in enumerate(sheet2_columns, 1):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    for row_idx, row in df.iterrows():
        for col_idx, col_name in enumerate(sheet2_columns, 1):
            value = row.get(col_name, '')
            if col_name in ['收入金额', '支出金额', '滚动余额'] and pd.notna(value) and value != '':
                value = round(float(value), 2)
            cell = ws2.cell(row=row_idx + 2, column=col_idx, value=value)
            if col_name == '状态' and value == '待确认':
                cell.font = Font(color='FF0000')
    
    for col_idx in range(1, len(sheet2_columns) + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 18
    
    wb.save(output_path)
    print(f"输出文件: {output_path}")


# ============================================================================
# 主入口
# ============================================================================

def main():
    """命令行入口"""
    if len(sys.argv) < 3:
        print("用法: python boc_parser.py <输入文件> <配置文件>")
        sys.exit(1)
    
    input_file = sys.argv[1]
    config_path = sys.argv[2]
    
    if not os.path.exists(input_file):
        print(f"错误: 输入文件不存在")
        sys.exit(1)
    
    if not os.path.exists(config_path):
        print(f"错误: 配置文件不存在")
        sys.exit(1)
    
    df = parse_boc_statement(input_file, config_path)
    
    input_path = Path(input_file)
    output_file = input_path.stem + '_基础数据表.xlsx'
    output_path = input_path.parent / output_file
    
    write_excel_output(df, str(output_path))
    
    print(f"\n完成！总记录: {len(df)}, 正常: {len(df[df['状态']=='正常'])}, 待确认: {len(df[df['状态']=='待确认'])}")


if __name__ == '__main__':
    main()
