"""报表生成技能 — 从业务数据取数，生成标准报表 Excel

支持的报表类型：
- cash_journal: 现金日记账
- daily_report: 资金日报
- base_data: 基础数据表
- income_list: 收入明细表
- expense_list: 支出明细表
"""
import os
import sys
import json
from datetime import date

# 后端路径注入
_BACKEND = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "..", ".."))
if _BACKEND not in sys.path:
    sys.path.insert(0, _BACKEND)


REPORT_CONFIG = {
    "cash_journal": {
        "name": "现金日记账",
        "headers": ["日期", "上日余额", "收入", "支出", "本日余额"],
    },
    "daily_report": {
        "name": "资金日报",
        "headers": ["单位简称", "期初余额", "收入合计", "支出合计", "净变动", "期末余额"],
    },
    "base_data": {
        "name": "基础数据表",
        "headers": ["日期", "单位简称", "账户名称", "摘要", "对方", "收入", "支出", "余额", "状态"],
    },
    "income_list": {
        "name": "收入明细表",
        "headers": ["日期", "单位简称", "账户名称", "摘要", "对方", "收入金额", "余额"],
    },
    "expense_list": {
        "name": "支出明细表",
        "headers": ["日期", "单位简称", "账户名称", "摘要", "对方", "支出金额", "余额"],
    },
}


def run(params: dict) -> dict:
    """生成报表并写入 Excel

    params:
        report_type: 报表类型
        start_date: 开始日期
        end_date: 结束日期
        output_path: 输出路径（相对于工作区）
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from database import SessionLocal
    from services.report_service import cash_journal, daily_report

    report_type = params.get("report_type", "")
    if report_type not in REPORT_CONFIG:
        available = ", ".join(REPORT_CONFIG.keys())
        return {"ok": False, "error": f"不支持的报表类型: {report_type}，可用: {available}"}

    config = REPORT_CONFIG[report_type]
    start_date = params.get("start_date")
    end_date = params.get("end_date")
    output_path = params.get("output_path", f"outputs/{report_type}.xlsx")

    # 取数
    db = SessionLocal()
    try:
        rows = _fetch_data(db, report_type, start_date, end_date)
    except Exception as e:
        return {"ok": False, "error": f"取数失败: {e}"}
    finally:
        db.close()

    # 写 Excel
    try:
        # 确定输出路径：如果有 agent 工作区，写到工作区；否则写到 data/agents/system/
        agent_root = params.get("_agent_root")
        if not agent_root:
            skill_dir = os.path.dirname(os.path.abspath(__file__))
            agent_root = os.path.abspath(os.path.join(skill_dir, "..", "..", ".."))
        abs_path = os.path.join(agent_root, output_path)
        os.makedirs(os.path.dirname(abs_path), exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = config["name"]

        # 标题行
        title_font = Font(size=14, bold=True)
        ws.append([config["name"]])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(config["headers"]))
        ws.cell(row=1, column=1).font = title_font

        # 表头
        header_fill = PatternFill(start_color="E8F0E8", end_color="E8F0E8", fill_type="solid")
        header_font = Font(bold=True, size=11)
        ws.append(config["headers"])
        for col_idx in range(1, len(config["headers"]) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font

        # 数据行
        for row in rows:
            ws.append(row)

        # 自动列宽
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, len(config["headers"]) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(2, ws.max_row + 1)
                if not isinstance(ws.cell(row=r, column=col_idx).value, type(None))
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

        wb.save(abs_path)
        wb.close()

        return {
            "ok": True,
            "file_path": output_path,
            "report_type": report_type,
            "report_name": config["name"],
            "row_count": len(rows),
        }
    except Exception as e:
        return {"ok": False, "error": f"写入 Excel 失败: {e}"}


def _fetch_data(db, report_type: str, start_date: str = None, end_date: str = None) -> list:
    """从数据库取数"""
    from services.report_service import cash_journal, daily_report, income_list, expense_list
    from datetime import datetime as _dt

    # 字符串日期转 date 对象
    sd = _dt.strptime(start_date, "%Y-%m-%d").date() if start_date else _dt.now().date()
    ed = _dt.strptime(end_date, "%Y-%m-%d").date() if end_date else _dt.now().date()

    if report_type == "cash_journal":
        params = {"start_date": sd, "end_date": ed}
        blocks = cash_journal(db, **params)
        rows = []
        for block in blocks:
            for r in block.get("rows", []):
                rows.append([
                    r.get("business_date"), r.get("prev_balance"),
                    r.get("income"), r.get("expense"), r.get("day_balance"),
                ])
        return rows

    elif report_type == "daily_report":
        params = {"start_date": sd, "end_date": ed}
        result = daily_report(db, **params)
        return [
            [r.get("entity_name"), r.get("opening_balance"), r.get("total_income"),
             r.get("total_expense"), r.get("net_change"), r.get("ending_balance")]
            for r in result
        ]

    elif report_type == "base_data":
        from db.tables import FundEvent
        from datetime import datetime as _dt
        query = db.query(FundEvent).order_by(FundEvent.business_date.desc())
        if start_date:
            query = query.filter(FundEvent.business_date >= _dt.strptime(start_date, "%Y-%m-%d").date())
        if end_date:
            query = query.filter(FundEvent.business_date <= _dt.strptime(end_date, "%Y-%m-%d").date())
        events = query.limit(5000).all()
        return [
            [e.business_date.isoformat() if e.business_date else "",
             e.entity_name or "", e.account_name or "",
             e.summary or "", e.counterparty or "",
             float(e.amount_in or 0), float(e.amount_out or 0),
             float(e.rolling_balance or 0), e.state or ""]
            for e in events
        ]

    elif report_type == "income_list":
        params = {"page": 1, "page_size": 10000}
        if sd: params["start_date"] = sd
        if ed: params["end_date"] = ed
        result = income_list(db, **params)
        return [
            [r.get("business_date"), r.get("entity_name"), r.get("account_name"),
             r.get("summary_text"), r.get("counterparty_name"),
             r.get("amount"), r.get("rolling_balance")]
            for r in result.get("items", [])
        ]

    elif report_type == "expense_list":
        params = {"page": 1, "page_size": 10000}
        if sd: params["start_date"] = sd
        if ed: params["end_date"] = ed
        result = expense_list(db, **params)
        return [
            [r.get("business_date"), r.get("entity_name"), r.get("account_name"),
             r.get("summary_text"), r.get("counterparty_name"),
             r.get("amount"), r.get("rolling_balance")]
            for r in result.get("items", [])
        ]

    return []
