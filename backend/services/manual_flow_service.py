"""手工流水服务 — Track A 快速录入 + Track B Excel上传，收敛到统一预览/提交流程"""
import json
import os
import re
import uuid
from datetime import datetime, date as date_type
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy.orm import Session
from sqlalchemy import text

from config import DATA_DIR
from core import artifact_runtime
from core.ai_call import chat
from core.parser_engine import (
    detect_format,
    read_file_from_bytes,
    detect_header_row,
    normalize_date,
    normalize_amount,
)
from core.pii_masker import mask_row
from core.ai_parse_utils import (
    strip_thinking_tags, extract_and_parse_json,
    find_active_agent, call_agent_for_mapping,
    shorten_headers, format_sample_text,
    restore_mapping_keys, validate_mapping_confidence,
)
from db.tables import FundEvent, ImportBatch, Account, Entity, AccountAlias, ManualFieldPool
from services.manual_scheme_service import get_scheme_by_code
from services import log_service


# ── 实体/账户匹配 ──────────────────────────────

def _match_entity(db: Session, match_key: str) -> Optional[Tuple[int, str]]:
    """匹配法人：entity_code > short_name > alias > name 部分匹配"""
    if not match_key or not match_key.strip():
        return None
    key = match_key.strip()

    # 1. entity_code 精确
    row = db.query(Entity).filter(Entity.entity_code == key, Entity.status == "enabled").first()
    if row:
        return (row.id, row.short_name)

    # 2. short_name 精确
    row = db.query(Entity).filter(Entity.short_name == key, Entity.status == "enabled").first()
    if row:
        return (row.id, row.short_name)

    # 3. 别名匹配（通过 account_aliases）
    alias = db.query(AccountAlias).filter(AccountAlias.alias_text == key).first()
    if alias:
        acct = db.query(Account).filter(Account.id == alias.account_id).first()
        if acct:
            ent = db.query(Entity).filter(Entity.id == acct.entity_id, Entity.status == "enabled").first()
            if ent:
                return (ent.id, ent.short_name)

    # 4. name 包含匹配
    rows = db.query(Entity).filter(Entity.name.contains(key), Entity.status == "enabled").all()
    if len(rows) == 1:
        return (rows[0].id, rows[0].short_name)

    return None


def _match_account(db: Session, match_key: str, entity_id: int = None) -> Optional[Tuple[int, str]]:
    """匹配账户：account_code > alias > account_alias"""
    if not match_key or not match_key.strip():
        return None
    key = match_key.strip()

    q = db.query(Account).filter(Account.status == "enabled")
    if entity_id:
        q = q.filter(Account.entity_id == entity_id)

    # 1. account_code 精确
    row = q.filter(Account.account_code == key).first()
    if row:
        return (row.id, row.account_alias)

    # 2. alias
    alias_q = db.query(AccountAlias).filter(AccountAlias.alias_text == key)
    alias = alias_q.first()
    if alias:
        acct = db.query(Account).filter(Account.id == alias.account_id, Account.status == "enabled")
        if entity_id:
            acct = acct.filter(Account.entity_id == entity_id)
        acct = acct.first()
        if acct:
            return (acct.id, acct.account_alias)

    # 3. account_alias 精确
    row = q.filter(Account.account_alias == key).first()
    if row:
        return (row.id, row.account_alias)

    return None


# ── 行验证 ──────────────────────────────

def _validate_row(parsed: Dict) -> List[str]:
    """返回异常代码列表"""
    errors = []
    if not parsed.get("_entity_id"):
        errors.append("ENTITY_MATCH_FAILED")
    if not parsed.get("_account_id"):
        errors.append("ACCOUNT_MATCH_FAILED")
    biz_date = parsed.get("business_date")
    if not biz_date:
        errors.append("DATE_INVALID")
    elif isinstance(biz_date, str) and not _is_valid_date(biz_date):
        errors.append("DATE_INVALID")
    if not parsed.get("summary_text"):
        errors.append("SUMMARY_MISSING")
    inc = parsed.get("income_amount") or 0
    exp = parsed.get("expense_amount") or 0
    has_balance = parsed.get("previous_balance_input") or parsed.get("ending_balance_input")
    if inc == 0 and exp == 0 and not has_balance:
        errors.append("AMOUNT_INVALID")
    elif inc > 0 and exp > 0:
        errors.append("AMOUNT_INVALID")
    return errors


def _is_valid_date(s: str) -> bool:
    try:
        datetime.strptime(s[:10], "%Y-%m-%d")
        return True
    except (ValueError, IndexError):
        try:
            datetime.strptime(s[:10], "%Y/%m/%d")
            return True
        except (ValueError, IndexError):
            return False


# ── 列映射构建 ──────────────────────────────

def _build_column_mapping(db: Session, scheme_code: str) -> Dict[str, str]:
    """field_name_cn → field_code"""
    from db.tables import ManualFieldPool
    scheme = get_scheme_by_code(db, scheme_code)
    if not scheme:
        return {}
    pool = {f.field_code: f.field_name_cn for f in db.query(ManualFieldPool).all()}
    mapping = {}
    for fc in scheme.selected_fields:
        cn = pool.get(fc)
        if cn:
            mapping[cn] = fc
    return mapping


# ── 批次创建 ──────────────────────────────

def _create_batch(db: Session, source_type: str, source_name: str = None, status: str = "uploaded") -> ImportBatch:
    batch = ImportBatch(
        batch_code=f"M-{uuid.uuid4().hex[:8].upper()}",
        source_type=source_type,
        source_name=source_name,
        status=status,
    )
    db.add(batch)
    db.commit()
    db.refresh(batch)
    return batch


# ── Track A: 快速录入保存 ──────────────────────────────

def quick_entry_save(db: Session, rows: List[Dict], scheme_code: str = "manual_multi_subject_basic") -> Dict:
    batch = _create_batch(db, "manual_quick", "quick_entry", status="uploaded")
    inserted = 0
    for raw in rows:
        parsed = _process_row(db, raw)
        event = _canonical_event_from_manual(parsed)
        event["state"] = "待确认"
        db.add(FundEvent(**event, batch_id=batch.id))
        inserted += 1
    db.commit()
    log_service.write_log(
        db,
        action="manual_quick_entry",
        module="manual_flow",
        detail={"batch_code": batch.batch_code, "inserted_rows": inserted, "scheme_code": scheme_code},
        batch_id=batch.id,
    )
    return {"batch_code": batch.batch_code, "inserted_rows": inserted}


# ── Track B: Excel上传 ──────────────────────────────

def upload_workbook(db: Session, file_data: bytes, filename: str, scheme_code: str) -> Dict:
    fmt = detect_format(filename)
    if fmt not in ("xlsx", "xls", "csv"):
        raise ValueError(f"不支持的文件格式: {filename}")

    rows = read_file_from_bytes(file_data, filename, fmt)
    if not rows:
        raise ValueError("文件内容为空")

    header_idx = detect_header_row(rows)
    headers = [h.strip() for h in rows[header_idx] if h and h.strip()]
    data_rows = rows[header_idx + 1:]
    # 过滤空行
    data_rows = [r for r in data_rows if any(c and c.strip() for c in r)]

    batch = _create_batch(db, "manual_excel", filename)

    # Save file with batch_code prefix for stable lookup
    upload_dir = os.path.join(DATA_DIR, "uploads")
    os.makedirs(upload_dir, exist_ok=True)
    safe_name = re.sub(r'[\/:*?"<>|]', '_', filename or 'upload.xlsx')
    file_path = os.path.join(upload_dir, f"{batch.batch_code}_{safe_name}")
    with open(file_path, "wb") as f:
        f.write(file_data)

    return {
        "batch_code": batch.batch_code,
        "batch_id": batch.id,
        "file_name": filename,
        "row_count": len(data_rows),
        "headers": headers,
        "scheme_code": scheme_code,
    }


def preview_manual(db: Session, batch_code: str, scheme_code: str = None) -> Dict:
    batch = db.query(ImportBatch).filter(ImportBatch.batch_code == batch_code).first()
    if not batch:
        raise ValueError("批次不存在")

    # 快速录入场景：已有 FundEvent，直接构建预览
    if batch.source_type == "manual_quick":
        return _preview_from_events(db, batch)

    # Excel 上传场景：如果有 FundEvent 直接构建预览，否则从文件解析
    if batch.source_type in ("manual_excel", "manual_file"):
        existing_events = db.query(FundEvent).filter(FundEvent.batch_id == batch.id).count()
        if existing_events > 0:
            return _preview_from_events(db, batch)
        # Parse file and create FundEvents
        return _preview_from_excel_file(db, batch, scheme_code)

    raise ValueError(f"不支持的批次类型: {batch.source_type}")


def _preview_from_events(db: Session, batch: ImportBatch) -> Dict:
    """快速录入：从已有的 FundEvent 记录构建预览"""
    events = db.query(FundEvent).filter(FundEvent.batch_id == batch.id).order_by(FundEvent.id).all()

    parsed_rows = []
    abnormal_rows = []

    for i, ev in enumerate(events):
        row = _event_to_dict(ev)
        row["_row_no"] = i + 1
        # Validate row to determine if it's valid or abnormal
        errors = _validate_fund_event(ev)
        row["_errors"] = errors
        row["abnormal_code"] = ",".join(errors) if errors else None
        if errors:
            abnormal_rows.append(row)
        else:
            parsed_rows.append(row)

    batch.status = "previewed"
    db.commit()

    return {
        "batch_code": batch.batch_code,
        "total_count": len(events),
        "valid_count": len(parsed_rows),
        "abnormal_count": len(abnormal_rows),
        "parsed_rows": parsed_rows,
        "abnormal_rows": abnormal_rows,
    }


def _preview_from_file(db: Session, batch: ImportBatch, scheme_code: str) -> Dict:
    """Excel上传：重新读取上传文件构建预览"""
    mapping = _build_column_mapping(db, scheme_code)

    upload_dir = os.path.join(DATA_DIR, "uploads")
    source = batch.source_name or ""
    fmt = detect_format(source)
    file_path = None
    for f in os.listdir(upload_dir):
        if batch.batch_code.replace("M-", "manual_") in f or source in f:
            file_path = os.path.join(upload_dir, f)
            break

    if not file_path or not os.path.isfile(file_path):
        raise ValueError("上传文件已丢失，请重新上传")

    with open(file_path, "rb") as fh:
        file_data = fh.read()
    rows = read_file_from_bytes(file_data, source, fmt)
    header_idx = detect_header_row(rows)
    headers = [h.strip() for h in rows[header_idx]]
    data_rows = rows[header_idx + 1:]
    data_rows = [r for r in data_rows if any(c and c.strip() for c in r)]

    parsed_rows = []
    abnormal_rows = []

    # 删除旧的 fund_events（文件类预览允许重复）
    db.query(FundEvent).filter(FundEvent.batch_id == batch.id).delete()

    for i, raw in enumerate(data_rows):
        row_dict = _map_raw_row(headers, raw, mapping)
        parsed = _process_row(db, row_dict)
        errors = _validate_row(parsed)
        parsed["parse_status"] = "valid" if not errors else "abnormal"
        parsed["abnormal_code"] = ",".join(errors) if errors else None
        parsed["_row_no"] = i + 1
        parsed["_raw"] = [c.strip() if c else "" for c in raw]
        _create_fund_event(db, batch.id, "manual", parsed)
        if errors:
            abnormal_rows.append(parsed)
        else:
            parsed_rows.append(parsed)

    batch.status = "previewed"
    db.commit()

    return {
        "batch_code": batch.batch_code,
        "total_count": len(data_rows),
        "valid_count": len(parsed_rows),
        "abnormal_count": len(abnormal_rows),
        "parsed_rows": parsed_rows,
        "abnormal_rows": abnormal_rows,
    }


def commit_manual(
    db: Session,
    batch_code: str,
    confirm_rows: List[int] = None,
    fixes: List[Dict] = None,
    parser_artifact_id: Optional[int] = None,
) -> Dict:
    if parser_artifact_id is None:
        if isinstance(confirm_rows, int):
            parser_artifact_id = confirm_rows
        else:
            raise ValueError("提交手工 Excel 需要 parser_artifact_id")
    batch = db.query(ImportBatch).filter(ImportBatch.batch_code == batch_code).first()
    if not batch:
        raise ValueError("批次不存在")
    file_path = _find_manual_upload(batch)
    if not file_path:
        raise ValueError("原始上传文件未找到")

    db.query(FundEvent).filter(FundEvent.batch_id == batch.id).delete()
    rows = list(
        artifact_runtime.run_parser(
            db,
            parser_artifact_id,
            file_path,
            {"batch_code": batch_code},
        )
    )
    for row in rows:
        db.add(FundEvent(**row, batch_id=batch.id, parser_artifact_id=parser_artifact_id))
    batch.status = "committed"
    batch.updated_at = datetime.now()
    db.commit()
    log_service.write_log(
        db,
        action="manual_commit",
        module="manual_flow",
        detail={
            "batch_code": batch_code,
            "parser_artifact_id": parser_artifact_id,
            "inserted_rows": len(rows),
        },
        batch_id=batch.id,
    )
    return {
        "batch_code": batch_code,
        "parser_artifact_id": parser_artifact_id,
        "inserted_rows": len(rows),
    }


# ── 导出模板 ──────────────────────────────

def export_template(db: Session, scheme_code: str, include_example: bool = False) -> bytes:
    import openpyxl
    from io import BytesIO

    scheme = get_scheme_by_code(db, scheme_code)
    if not scheme:
        raise ValueError(f"方案不存在: {scheme_code}")

    pool_map = {f.field_code: f for f in db.query(ManualFieldPool).all()}
    headers = []
    hints = []
    for fc in scheme.selected_fields:
        field = pool_map.get(fc)
        if field:
            headers.append(field.field_name_cn)
            hints.append(f"{'必填' if field.is_core else '选填'} ({field.data_type})")
        else:
            headers.append(fc)
            hints.append("")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "手工流水录入"
    ws.append(headers)
    ws.append(hints)

    if include_example:
        example = []
        for fc in scheme.selected_fields:
            if fc == "entity_match_key":
                example.append("E001")
            elif fc == "account_match_key":
                example.append("A001")
            elif fc == "business_date":
                example.append("2026-04-17")
            elif fc == "summary_text":
                example.append("测试摘要")
            elif fc == "counterparty_name":
                example.append("对方公司")
            elif fc == "income_amount":
                example.append("10000.00")
            elif fc == "expense_amount":
                example.append("")
            elif fc == "note_text":
                example.append("示例行")
            else:
                example.append("")
        ws.append(example)

    # 说明页
    ws2 = wb.create_sheet("字段说明")
    ws2.append(["字段代码", "中文名称", "数据类型", "是否必填", "说明"])
    for fc in scheme.selected_fields:
        field = pool_map.get(fc)
        if field:
            ws2.append([field.field_code, field.field_name_cn, field.data_type,
                        "是" if field.is_core else "否", field.field_name_cn])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 内部辅助 ──────────────────────────────

def _process_row(db: Session, row: Dict) -> Dict:
    """匹配实体/账户 + 规范化日期/金额"""
    parsed = dict(row)

    entity_key = str(row.get("entity_match_key", "")).strip()
    account_key = str(row.get("account_match_key", "")).strip()

    ent = _match_entity(db, entity_key)
    if ent:
        parsed["_entity_id"] = ent[0]
        parsed["_entity_name"] = ent[1]
    else:
        parsed["_entity_id"] = None
        parsed["_entity_name"] = entity_key

    acct = _match_account(db, account_key, parsed.get("_entity_id"))
    if acct:
        parsed["_account_id"] = acct[0]
        parsed["_account_name"] = acct[1]
    else:
        parsed["_account_id"] = None
        parsed["_account_name"] = account_key

    # 日期规范化
    raw_date = str(row.get("business_date", "")).strip()
    parsed["business_date"] = normalize_date(raw_date) if raw_date else raw_date

    # 金额规范化
    for amt_field in ("income_amount", "expense_amount", "previous_balance_input", "ending_balance_input"):
        raw_val = row.get(amt_field)
        if raw_val is not None and str(raw_val).strip():
            parsed[amt_field] = normalize_amount(str(raw_val))
        elif raw_val is not None:
            parsed[amt_field] = None

    return parsed


def _map_raw_row(headers: List[str], raw: List[str], mapping: Dict[str, str]) -> Dict:
    """原始行 → field_code dict"""
    result = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        fc = mapping.get(h)
        val = raw[i].strip() if i < len(raw) and raw[i] else ""
        if fc:
            result[fc] = val
        else:
            # 尝试直接作为 field_code
            result[h] = val
    return result


def _validate_fund_event(ev: FundEvent) -> List[str]:
    """Validate a FundEvent ORM row, return error codes."""
    errors = []
    if not ev.business_date:
        errors.append("DATE_INVALID")
    if not ev.entity_code or not ev.entity_code.strip():
        errors.append("ENTITY_MATCH_FAILED")
    if not ev.account_code or not ev.account_code.strip():
        errors.append("ACCOUNT_MATCH_FAILED")
    inc = float(ev.amount_in) if ev.amount_in else 0
    exp = float(ev.amount_out) if ev.amount_out else 0
    if inc == 0 and exp == 0:
        errors.append("AMOUNT_INVALID")
    elif inc > 0 and exp > 0:
        errors.append("AMOUNT_INVALID")
    return errors


def _preview_from_excel_file(db: Session, batch: ImportBatch, scheme_code: str = None) -> Dict:
    """Parse uploaded Excel file for manual_excel/manual_file batches and create FundEvents."""
    mapping = _build_column_mapping(db, scheme_code or "manual_multi_subject_basic")

    upload_dir = os.path.join(DATA_DIR, "uploads")
    file_path = None
    # Prefer batch_code-based lookup
    for f in os.listdir(upload_dir):
        if f.startswith(batch.batch_code + "_") and os.path.isfile(os.path.join(upload_dir, f)):
            file_path = os.path.join(upload_dir, f)
            break
    if not file_path:
        file_path = _find_manual_upload(batch)
    if not file_path:
        raise ValueError("上传文件已丢失，请重新上传")

    source = batch.source_name or ""
    fmt = detect_format(source)
    with open(file_path, "rb") as fh:
        file_data = fh.read()
    rows = read_file_from_bytes(file_data, source, fmt)
    header_idx = detect_header_row(rows)
    headers = [h.strip() for h in rows[header_idx] if h and h.strip()]
    data_rows = rows[header_idx + 1:]
    data_rows = [r for r in data_rows if any(c and str(c).strip() for c in r)]

    # Delete old fund_events if re-previewing
    db.query(FundEvent).filter(FundEvent.batch_id == batch.id).delete()

    parsed_rows = []
    abnormal_rows = []
    for i, raw in enumerate(data_rows):
        row_dict = _map_raw_row(headers, raw, mapping)
        parsed = _process_row(db, row_dict)
        errors = _validate_row(parsed)
        parsed["parse_status"] = "valid" if not errors else "abnormal"
        parsed["abnormal_code"] = ",".join(errors) if errors else None
        parsed["_row_no"] = i + 1
        _create_fund_event(db, batch.id, "手工录入", parsed)
        if errors:
            abnormal_rows.append(parsed)
        else:
            parsed_rows.append(parsed)

    batch.status = "previewed"
    db.commit()

    return {
        "batch_code": batch.batch_code,
        "source_type": batch.source_type,
        "source_name": batch.source_name,
        "total_count": len(data_rows),
        "valid_count": len(parsed_rows),
        "abnormal_count": len(abnormal_rows),
        "parsed_rows": parsed_rows,
        "abnormal_rows": abnormal_rows,
    }


def _event_to_dict(ev: FundEvent) -> Dict:
    """FundEvent ORM → 预览用的 dict"""
    return {
        "_entity_id": ev.entity.id if ev.entity else None,
        "_entity_name": ev.entity_name,
        "_account_id": ev.account.id if ev.account else None,
        "_account_name": ev.account_name,
        "business_date": str(ev.business_date) if ev.business_date else "",
        "summary_text": ev.summary,
        "counterparty_name": ev.counterparty,
        "income_amount": float(ev.amount_in) if ev.amount_in else None,
        "expense_amount": float(ev.amount_out) if ev.amount_out else None,
        "parse_status": ev.state,
        "abnormal_code": None if ev.state == "正常" else ev.state,
    }


def _create_fund_event(db: Session, batch_id: int, source_type: str, parsed: Dict):
    income = parsed.get("income_amount")
    expense = parsed.get("expense_amount")

    raw_entity = str(parsed.get("entity_match_key") or "").strip()
    entity_code = raw_entity if parsed.get("_entity_id") and raw_entity else None
    entity_name = parsed.get("_entity_name") or raw_entity or ""

    raw_account = str(parsed.get("account_match_key") or "").strip()
    account_code = raw_account if parsed.get("_account_id") and raw_account else None
    account_name = parsed.get("_account_name") or raw_account or ""

    biz_date = parsed.get("business_date")
    if isinstance(biz_date, str) and biz_date:
        try:
            biz_date = datetime.strptime(biz_date[:10].replace("/", "-"), "%Y-%m-%d").date()
        except (ValueError, IndexError):
            biz_date = None

    ev = FundEvent(
        batch_id=batch_id,
        business_date=biz_date,
        entity_code=entity_code,
        entity_name=entity_name,
        account_code=account_code,
        account_name=account_name,
        summary=parsed.get("summary_text", ""),
        counterparty=parsed.get("counterparty_name", ""),
        amount_in=float(income) if income else 0,
        amount_out=float(expense) if expense else 0,
        rolling_balance=parsed.get("ending_balance_input"),
        state="待确认",
        source="手工录入",
    )
    db.add(ev)
    db.flush()


def _find_manual_upload(batch: ImportBatch) -> Optional[str]:
    upload_dir = os.path.join(DATA_DIR, "uploads")
    if not os.path.isdir(upload_dir):
        return None
    # Prefer batch_code-based lookup
    for filename in os.listdir(upload_dir):
        if filename.startswith(batch.batch_code + "_") and os.path.isfile(os.path.join(upload_dir, filename)):
            return os.path.join(upload_dir, filename)
    # Fallback to source_name suffix
    source = batch.source_name or ""
    for filename in os.listdir(upload_dir):
        if source and filename.endswith(source) and os.path.isfile(os.path.join(upload_dir, filename)):
            return os.path.join(upload_dir, filename)
    return None


def _canonical_event_from_manual(parsed: Dict) -> Dict:
    biz_date = parsed.get("business_date")
    if isinstance(biz_date, str) and biz_date:
        try:
            biz_date = datetime.strptime(biz_date[:10].replace("/", "-"), "%Y-%m-%d").date()
        except (ValueError, IndexError):
            biz_date = None
    else:
        biz_date = None
    income = parsed.get("income_amount") or 0
    expense = parsed.get("expense_amount") or 0

    # If _entity_id exists, matching succeeded → use the match key as code
    # If _entity_id is None, matching failed → code must be None, keep name for display
    raw_entity = str(parsed.get("entity_match_key") or "").strip()
    entity_code = raw_entity if parsed.get("_entity_id") and raw_entity else None
    entity_name = parsed.get("_entity_name") or raw_entity or ""

    raw_account = str(parsed.get("account_match_key") or "").strip()
    account_code = raw_account if parsed.get("_account_id") and raw_account else None
    account_name = parsed.get("_account_name") or raw_account or ""

    return {
        "business_date": biz_date,
        "entity_code": entity_code,
        "entity_name": entity_name,
        "account_code": account_code,
        "account_name": account_name,
        "summary": parsed.get("summary_text") or "",
        "counterparty": parsed.get("counterparty_name") or "",
        "amount_in": income,
        "amount_out": expense,
        "rolling_balance": parsed.get("ending_balance_input"),
        "state": "待确认",
        "source": "手工录入",
    }


# ──────────────────────────────────────────
# AI 智能解析（手工流水列映射）
# ──────────────────────────────────────────

def ai_parse_headers(
    db: Session,
    headers: list,
    sample_rows: list = None,
    agent_id: int = None,
    scheme_code: str = None,
) -> Dict[str, Any]:
    """通过指定智能体分析手工流水表头，自动生成列映射。"""
    agent = find_active_agent(db, agent_id)
    if not agent or not agent.ai_config:
        return {"ok": False, "error": "未找到可用的 AI 智能体，请先在「AI智能体」中创建并配置"}

    # 获取方案的可用字段
    pool = db.query(ManualFieldPool).filter(ManualFieldPool.status == "active").all()
    field_map = {f.field_code: f.field_name_cn for f in pool}
    available_fields = dict(field_map)
    if scheme_code:
        scheme = get_scheme_by_code(db, scheme_code)
        if scheme:
            available_fields = {fc: field_map.get(fc, fc) for fc in scheme.selected_fields if fc in field_map}

    field_desc = "\n".join(f"  - {k}: {v}" for k, v in available_fields.items())
    short = shorten_headers(headers)
    header_list = ", ".join(f'"{h}"' for h in short)
    sample_text = format_sample_text(
        sample_rows, max_rows=2,
        mask_fn=lambda row: mask_row(row, short),
    )

    user_message = f"""请分析以下手工流水表的表头列名，将每列映射到对应的标准字段。

## 表头列名
{header_list}
{sample_text}

## 可用标准字段
{field_desc}

## 要求
1. 返回 JSON 格式：{{"Excel列名": "字段code"}}
2. 只映射能匹配的列，不确定的不映射
3. business_date、summary_text 是核心必填字段
4. 如果列名含义不明确，参考样本数据判断
5. entity_match_key 用于匹配法人简称/编码，account_match_key 用于匹配账户名称/编号

请严格返回如下 JSON 格式（不要包含其他文字）:
{{"mapping": {{"Excel列名1": "field_code1", "Excel列名2": "field_code2"}}}}"""

    ai_result = call_agent_for_mapping(agent, user_message)
    if not ai_result.get("ok"):
        return ai_result

    parsed = ai_result["parsed"]
    mapping = parsed.get("mapping", {})
    mapping = restore_mapping_keys(mapping, headers)

    valid_mapping = {k: v for k, v in mapping.items() if v in field_map}
    if not valid_mapping:
        return {"ok": False, "error": "AI 未能识别出有效的列映射"}

    return {
        "ok": True,
        "mapping": valid_mapping,
        "confidence": validate_mapping_confidence(valid_mapping, len(headers)),
        "matched_count": len(valid_mapping),
        "total_columns": len(headers),
    }
