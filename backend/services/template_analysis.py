"""Template structure analysis and field mapping service.

Pure logic — no LLM, no database writes, no legacy agent dependency.
"""
from __future__ import annotations

import re
from typing import Any


def parse_template_structure(template_file: str) -> dict[str, Any]:
    """Stage A: parse Excel template to extract placeholders and merged cells.

    Returns:
        {
            "placeholders": [str, ...],
            "merged_cells": [{"min_row", "max_row", "min_col", "max_col"}, ...],
            "header_rows": [int, ...],
        }
    """
    placeholders: list[str] = []
    merged_cells: list[dict[str, int]] = []
    header_rows: list[int] = []

    try:
        from openpyxl import load_workbook

        wb = load_workbook(template_file, read_only=True, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        found = re.findall(
                            r'\$\{([^}]+)\}|\{\{([^}]+)\}\}|【([^】]+)】',
                            cell.value,
                        )
                        for match in found:
                            ph = match[0] or match[1] or match[2]
                            if ph and ph not in placeholders:
                                placeholders.append(ph)

            for mc in ws.merged_cells.ranges:
                merged_cells.append({
                    "min_row": mc.min_row,
                    "max_row": mc.max_row,
                    "min_col": mc.min_col,
                    "max_col": mc.max_col,
                })
        wb.close()
    except Exception:
        pass

    return {
        "placeholders": placeholders,
        "merged_cells": merged_cells,
        "header_rows": header_rows,
    }


def map_placeholders_to_fields(
    stage_a: dict[str, Any],
    field_dictionary: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Stage B: map detected placeholders to canonical fields.

    Args:
        stage_a: output of parse_template_structure()
        field_dictionary: optional dict of {field_key: {cn_name, aliases}}.
            Falls back to field_dictionary_service defaults.

    Returns:
        {
            "placeholder_bindings": {placeholder_name: {"primitive", "value"}, ...},
            "confidence": float,
            "matched_count": int,
            "total_placeholders": int,
        }
    """
    if field_dictionary is None:
        from services.field_dictionary_service import get_field_dictionary
        field_dictionary = get_field_dictionary()

    placeholders = stage_a.get("placeholders", [])
    bindings: dict[str, dict[str, Any]] = {}
    matched_count = 0

    for ph in placeholders:
        best_match = _find_best_field_match(ph, field_dictionary)
        if best_match:
            bindings[ph] = {"primitive": "field", "value": best_match}
            matched_count += 1
        else:
            bindings[ph] = {"primitive": "const", "value": ph}

    confidence = matched_count / max(len(placeholders), 1) if placeholders else 0.0

    return {
        "placeholder_bindings": bindings,
        "confidence": confidence,
        "matched_count": matched_count,
        "total_placeholders": len(placeholders),
    }


def _find_best_field_match(
    placeholder: str,
    field_dictionary: dict[str, dict[str, Any]],
) -> str | None:
    """Find the best matching canonical field for a placeholder string."""
    best_match: str | None = None
    best_score = 0

    for field_key, field_info in field_dictionary.items():
        aliases = field_info.get("aliases", [])
        cn_name = field_info.get("cn_name", "")
        all_names = [cn_name] + aliases

        for name in all_names:
            if name in placeholder or placeholder in name:
                score = len(name) / max(len(placeholder), 1)
                if score > best_score:
                    best_score = score
                    best_match = field_key

    return best_match
