"""导出服务 — 使用 openpyxl 生成 xlsx"""
import os
from datetime import datetime
from typing import Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import EXPORT_DIR


# 报表类型 → 中文名 + 表头 + 数据查询函数
EXPORT_CONFIG = {
    "base_data": {
        "name": "基础数据表",
        "headers": ["日期", "单位简称", "账户名称", "摘要", "对方", "收入", "支出", "余额", "状态"],
    },
    "daily_report": {
        "name": "资金日报",
        "headers": ["单位简称", "期初余额", "收入合计", "支出合计", "净变动", "期末余额"],
    },
    "cash_journal": {
        "name": "现金日记账",
        "headers": ["日期", "上日余额", "收入", "支出", "本日余额"],
    },
    "account_balance": {
        "name": "账户余额表",
        "headers": ["单位简称", "账户名称", "期初余额", "本期收入", "本期支出", "期末余额"],
    },
    "income_list": {
        "name": "收入明细表",
        "headers": ["日期", "单位简称", "账户名称", "摘要", "对方", "收入金额", "余额"],
    },
    "expense_list": {
        "name": "支出明细表",
        "headers": ["日期", "单位简称", "账户名称", "摘要", "对方", "支出金额", "余额"],
    },
}


def generate_export(db, export_type: str, start_date: Optional[str], end_date: Optional[str], entity_id: Optional[int] = None) -> str:
    """生成 Excel 导出文件，返回文件路径"""
    from services.base_data_service import query_base_data
    from services.report_service import (
        daily_report, cash_journal, account_balance,
        income_list, expense_list,
    )

    config = EXPORT_CONFIG.get(export_type)
    if not config:
        raise ValueError(f"不支持的导出类型: {export_type}")

    wb = Workbook()
    ws = wb.active
    ws.title = config["name"]

    # 写表头
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="F7F4EE", end_color="F7F4EE", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="EEE7DA"),
        right=Side(style="thin", color="F0EADF"),
    )

    for col_idx, header in enumerate(config["headers"], 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 冻结首行
    ws.freeze_panes = "A2"

    # 获取数据
    rows = _fetch_data(db, export_type, start_date, end_date, entity_id)

    # 写数据行
    money_format = '#,##0.00'
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            # 金额列右对齐 + 千分位
            if isinstance(value, (int, float)) and col_idx in _get_money_columns(export_type):
                cell.number_format = money_format
                cell.alignment = Alignment(horizontal="right")

    # 自动列宽
    for col_idx in range(1, len(config["headers"]) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(rows) + 2)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(30, max(max_len + 2, 10))

    # 保存
    os.makedirs(EXPORT_DIR, exist_ok=True)
    now = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"{config['name']}_{start_date or 'all'}_{end_date or 'all'}_{now}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)

    return filepath


def _get_money_columns(export_type: str):
    """返回金额列的索引 (1-based)"""
    mapping = {
        "base_data": [6, 7, 8],
        "daily_report": [2, 3, 4, 5, 6],
        "cash_journal": [2, 3, 4, 5],
        "account_balance": [3, 4, 5, 6],
        "income_list": [6, 7],
        "expense_list": [6, 7],
    }
    return mapping.get(export_type, [])


def _fetch_data(db, export_type, start_date, end_date, entity_id) -> list:
    """根据类型获取数据行"""
    params = {"start_date": start_date, "end_date": end_date}
    if entity_id:
        params["entity_id"] = entity_id

    try:
        if export_type == "base_data":
            result = query_base_data(db, **params, page=1, page_size=10000)
            return [
                [r.get("business_date"), r.get("entity_name"), r.get("account_name"),
                 r.get("summary_text"), r.get("counterparty_name"),
                 r.get("income_amount"), r.get("expense_amount"),
                 r.get("rolling_balance"), r.get("abnormal_code") or "正常"]
                for r in result.get("items", [])
            ]
        elif export_type == "daily_report":
            rows = daily_report(db, **params)
            return [
                [r.get("entity_name"), r.get("opening_balance"), r.get("total_income"),
                 r.get("total_expense"), r.get("net_change"), r.get("ending_balance")]
                for r in rows
            ]
        elif export_type == "account_balance":
            rows = account_balance(db, **params)
            return [
                [r.get("entity_name"), r.get("account_name"), r.get("opening_balance"),
                 r.get("period_income"), r.get("period_expense"), r.get("ending_balance")]
                for r in rows if not r.get("is_subtotal")
            ]
        elif export_type == "income_list":
            result = income_list(db, **params, page=1, page_size=10000)
            return [
                [r.get("business_date"), r.get("entity_name"), r.get("account_name"),
                 r.get("summary_text"), r.get("counterparty_name"),
                 r.get("amount"), r.get("rolling_balance")]
                for r in result.get("items", [])
            ]
        elif export_type == "expense_list":
            result = expense_list(db, **params, page=1, page_size=10000)
            return [
                [r.get("business_date"), r.get("entity_name"), r.get("account_name"),
                 r.get("summary_text"), r.get("counterparty_name"),
                 r.get("amount"), r.get("rolling_balance")]
                for r in result.get("items", [])
            ]
        elif export_type == "cash_journal":
            blocks = cash_journal(db, **params)
            rows = []
            for block in blocks:
                for r in block.get("rows", []):
                    rows.append([
                        r.get("business_date"), r.get("prev_balance"),
                        r.get("income"), r.get("expense"), r.get("day_balance"),
                    ])
            return rows
    except Exception:
        return []
    return []
