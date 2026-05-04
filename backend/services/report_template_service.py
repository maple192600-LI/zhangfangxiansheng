"""报表模板服务 — CRUD + Excel解析 + 默认模板管理"""
import io
import json
import logging
from datetime import datetime
from typing import Dict, List, Optional

from sqlalchemy.orm import Session
from sqlalchemy import text

from db.tables import ReportTemplate

logger = logging.getLogger(__name__)
logger.info("[report_template_service] xls compat loaded v2")


def _load_workbook_any(file_path: str, data_only: bool = True, read_only: bool = False):
    """兼容加载 .xls / .xlsx，统一返回 openpyxl Workbook。

    .xls 用 xlrd 读出后转写成内存 .xlsx 再 openpyxl 打开，下游代码无需改。
    """
    from openpyxl import Workbook, load_workbook
    if not file_path.lower().endswith(".xls"):
        return load_workbook(file_path, read_only=read_only, data_only=data_only)

    import xlrd
    src = xlrd.open_workbook(file_path)
    dst = Workbook()
    dst.remove(dst.active)
    for si in range(src.nsheets):
        sh = src.sheet_by_index(si)
        ws = dst.create_sheet(title=(sh.name[:31] or f"Sheet{si+1}"))
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                value = cell.value
                if cell.ctype == 3:
                    try:
                        value = xlrd.xldate_as_datetime(cell.value, src.datemode)
                    except Exception:
                        value = str(cell.value)
                elif cell.ctype == 0:
                    value = None
                if value is not None and value != "":
                    ws.cell(row=r + 1, column=c + 1, value=value)
    return dst

CODE_PREFIX = "MB"
CODE_DIGITS = 4

# ── 中文表头 → 后端字段名映射 ──────────────────────────────
# key: 可能出现的中文表头文字（strip后小写比较）
# value: 后端数据字段名
HEADER_FIELD_MAP: Dict[str, str] = {
    # 通用
    "日期": "business_date",
    "业务日期": "business_date",
    "交易日期": "business_date",
    "发生日期": "business_date",
    "月": "month",
    "日": "day",
    "年": "year",
    # 单位/实体
    "单位简称": "entity_name",
    "单位名称": "entity_name",
    "单位编码": "entity_code",
    "往来单位": "entity_name",
    "法人名称": "entity_name",
    "会计组织": "entity_name",
    "实体名称": "entity_name",
    # 账户
    "账户名称": "account_name",
    "账户简称": "account_name",
    "账户别名": "account_name",
    "账户信息": "account_name",
    "银行账户": "account_name",
    "账号": "account_code",
    "账户编码": "account_code",
    "银行编号": "account_code",
    "开户行": "account_bank",
    # 金额类
    "收入": "income",
    "收入金额": "income_amount",
    "借方金额": "income_amount",
    "本期收入": "total_income",
    "收入合计": "total_income",
    "支出": "expense",
    "支出金额": "expense_amount",
    "贷方金额": "expense_amount",
    "本期支出": "total_expense",
    "支出合计": "total_expense",
    "金额": "amount",
    # 余额类
    "余额": "day_balance",
    "滚动余额": "rolling_balance",
    "当前余额": "rolling_balance",
    "结余": "day_balance",
    "本日余额": "day_balance",
    "日终余额": "day_balance",
    "期末余额": "ending_balance",
    "余额合计": "ending_balance",
    "上日余额": "prev_balance",
    "期初余额": "opening_balance",
    "期初": "opening_balance",
    "上期余额": "opening_balance",
    # 净变动
    "净变动": "net_change",
    "净额": "net_change",
    "本期净额": "net_change",
    # 文字类
    "摘要": "summary_text",
    "摘要说明": "summary_text",
    "对方": "counterparty_name",
    "对方名称": "counterparty_name",
    "对方单位": "counterparty_name",
    "对方账号": "counterparty_account",
    "币种": "currency",
    "币种标识": "currency",
    "方向": "direction",
    "收支方向": "direction",
    "状态": "parse_status",
    "来源": "source_type",
    "数据来源": "source_type",
    "异常代码": "abnormal_code",
    "备注": "remark",
}


def _match_field_key(header_name: str) -> str:
    """将中文表头名匹配到后端字段名，匹配不到则返回 col_N 形式"""
    name = header_name.strip()
    mapped = HEADER_FIELD_MAP.get(name)
    if mapped:
        return mapped
    return None


# 根据报表类型修正字段名（同一中文表头在不同报表中对应不同字段）
REPORT_TYPE_FIELD_REMAP = {
    "cash_journal": {
        "opening_balance": "prev_balance",   # 现金日记账的"期初余额"是上日余额
        "ending_balance": "day_balance",     # 现金日记账的"期末余额"是本日余额
        "total_income": "income",            # 现金日记账的"本期收入"是当日收入
        "total_expense": "expense",          # 现金日记账的"本期支出"是当日支出
        "period_income": "income",
        "period_expense": "expense",
    },
    "income_list": {
        "income_amount": "amount",
        "income": "amount",
    },
    "expense_list": {
        "expense_amount": "amount",
        "expense": "amount",
    },
}


def _remap_field_keys(columns: list, report_type: str) -> list:
    """根据报表类型修正 field_key"""
    remap = REPORT_TYPE_FIELD_REMAP.get(report_type, {})
    if not remap:
        return columns
    for col in columns:
        old_key = col.get("field_key", "")
        if old_key in remap:
            col["field_key"] = remap[old_key]
    return columns


def excel_to_html(file_path: str, max_rows: int = 800) -> str:
    """把 .xls / .xlsx 完整转成 HTML <table> 用于网页渲染。

    保留：合并单元格（colspan/rowspan）、对齐、粗体、背景色、所有行（含空行边框）。
    重排（不直接套用 Excel 原值）：
    - 字号统一映射：标题 ≥ 18px、副标题 14px、表头 14px、正文 14px、小注 12px
    - 列宽用比例分配，整张表撑满容器宽度（用户屏幕多大显示多大）
    - 行高最低 32px（避免空行被压成一条线）
    - 所有 td 都有边框（包括空 td，让"表格"看起来像表格）
    """
    from openpyxl.utils import get_column_letter
    wb = _load_workbook_any(file_path, data_only=True, read_only=False)
    ws = wb.worksheets[0]

    # 1. 合并单元格
    merge_anchor = {}
    skip_cells = set()
    for rng in ws.merged_cells.ranges:
        anchor = (rng.min_row, rng.min_col)
        merge_anchor[anchor] = {
            "colspan": rng.max_col - rng.min_col + 1,
            "rowspan": rng.max_row - rng.min_row + 1,
        }
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if (r, c) != anchor:
                    skip_cells.add((r, c))

    # 2. 列宽：取 Excel 原始字符宽度作为相对比例，归一化为百分比，撑满容器
    n_cols = ws.max_column or 1
    raw_widths = []
    for ci in range(1, n_cols + 1):
        cl = get_column_letter(ci)
        w = ws.column_dimensions[cl].width if cl in ws.column_dimensions else None
        raw_widths.append(float(w) if w and w > 0 else 10.0)
    total_w = sum(raw_widths) or 1
    col_pct = [round(w / total_w * 100, 3) for w in raw_widths]

    # 3. 行高 — 最低 32px，让空行也可见
    def _row_h(r):
        rd = ws.row_dimensions.get(r)
        if rd and rd.height:
            return max(int(rd.height * 1.33), 32)
        return 32

    # 4. 字号映射：根据 Excel 原字号决定 web 视觉权重
    def _web_font_size(orig_size):
        if not orig_size:
            return 14
        s = float(orig_size)
        if s >= 18:
            return 22  # 大标题
        if s >= 14:
            return 16  # 副标题
        if s >= 12:
            return 14  # 表头/正文
        return 13      # 小注

    rows_html = []
    max_r = min(ws.max_row or 0, max_rows)

    for r in range(1, max_r + 1):
        cells_html = []
        for c in range(1, n_cols + 1):
            if (r, c) in skip_cells:
                continue
            cell = ws.cell(row=r, column=c)
            value = cell.value
            text = "" if value is None else str(value)
            text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br>")

            attrs = []
            if (r, c) in merge_anchor:
                m = merge_anchor[(r, c)]
                if m["colspan"] > 1:
                    attrs.append(f'colspan="{m["colspan"]}"')
                if m["rowspan"] > 1:
                    attrs.append(f'rowspan="{m["rowspan"]}"')

            # 默认边框 + 内边距 — 所有 td 都有，包括空 td（让"表格"看起来像表格）
            styles = ["border:1px solid #c8c2b5", "padding:6px 8px", "word-break:break-word"]

            # 对齐
            ha = (cell.alignment.horizontal or "").lower() if cell.alignment else ""
            va = (cell.alignment.vertical or "").lower() if cell.alignment else ""
            if ha in ("left", "center", "right"):
                styles.append(f"text-align:{ha}")
            elif text and any(ch.isdigit() for ch in text) and not any(ch.isalpha() for ch in text):
                styles.append("text-align:right")
            if va in ("top", "middle", "bottom"):
                styles.append(f"vertical-align:{va}")
            else:
                styles.append("vertical-align:middle")

            # 字体（用 web 友好字号）
            if cell.font:
                if cell.font.bold:
                    styles.append("font-weight:600")
                styles.append(f"font-size:{_web_font_size(cell.font.size)}px")
                if cell.font.color and cell.font.color.rgb and isinstance(cell.font.color.rgb, str):
                    rgb = cell.font.color.rgb
                    if len(rgb) == 8:
                        styles.append(f"color:#{rgb[2:]}")

            # 背景
            try:
                if cell.fill and cell.fill.patternType == "solid":
                    fg = cell.fill.fgColor
                    if fg and fg.rgb and isinstance(fg.rgb, str) and len(fg.rgb) == 8:
                        styles.append(f"background-color:#{fg.rgb[2:]}")
            except Exception:
                pass

            attr_str = " ".join(attrs)
            style_str = ";".join(styles)
            cells_html.append(f'<td {attr_str} style="{style_str}">{text}</td>')

        rows_html.append(f'<tr style="height:{_row_h(r)}px">{"".join(cells_html)}</tr>')

    # 列宽 colgroup（百分比 → 撑满容器）
    colgroup = "".join(f'<col style="width:{p}%">' for p in col_pct)

    table_style = (
        "border-collapse:collapse;"
        "border:2px solid #8a857a;"
        "font-family:'Microsoft YaHei','PingFang SC','Helvetica Neue',sans-serif;"
        "font-size:14px;"
        "line-height:1.5;"
        "background:#fff;"
        "width:100%;"
        "table-layout:fixed;"
    )
    html = (
        '<div class="excel-template-wrap">'
        f'<table class="excel-template-table" style="{table_style}">'
        f'<colgroup>{colgroup}</colgroup>'
        f'<tbody>{"".join(rows_html)}</tbody>'
        '</table>'
        '</div>'
    )
    wb.close()
    return html


def _next_code(db: Session) -> str:
    """生成下一个模板编码 MBxxxx"""
    last = (
        db.query(ReportTemplate.template_code)
        .filter(ReportTemplate.template_code.like(f"{CODE_PREFIX}%"))
        .order_by(ReportTemplate.id.desc())
        .first()
    )
    if last:
        try:
            num = int(last[0][len(CODE_PREFIX):]) + 1
        except (ValueError, IndexError):
            num = 1
    else:
        num = 1
    return f"{CODE_PREFIX}{str(num).zfill(CODE_DIGITS)}"


def _columns_to_json(columns: list) -> str:
    """将列配置列表序列化为JSON"""
    return json.dumps([c.dict() if hasattr(c, "dict") else c for c in columns], ensure_ascii=False)


def _json_to_columns(json_str: str) -> list:
    """将JSON反序列化为列配置列表"""
    if not json_str:
        return []
    try:
        return json.loads(json_str)
    except json.JSONDecodeError:
        return []


def list_templates(
    db: Session,
    report_type: Optional[str] = None,
    status: Optional[str] = None,
) -> List[Dict]:
    """查询模板列表"""
    q = db.query(ReportTemplate)
    if report_type:
        q = q.filter(ReportTemplate.report_type == report_type)
    if status:
        q = q.filter(ReportTemplate.status == status)
    else:
        q = q.filter(ReportTemplate.status != "deleted")
    q = q.order_by(ReportTemplate.report_type, ReportTemplate.is_default.desc(), ReportTemplate.id)
    return [_to_dict(t) for t in q.all()]


def get_template(db: Session, template_id: int) -> Optional[Dict]:
    """获取单个模板"""
    t = db.query(ReportTemplate).filter(ReportTemplate.id == template_id).first()
    return _to_dict(t) if t else None


def get_default_template(db: Session, report_type: str) -> Optional[Dict]:
    """获取某报表类型的默认模板"""
    t = (
        db.query(ReportTemplate)
        .filter(
            ReportTemplate.report_type == report_type,
            ReportTemplate.is_default == True,
            ReportTemplate.status == "active",
        )
        .first()
    )
    return _to_dict(t) if t else None


def create_template(db: Session, data: dict) -> Dict:
    """创建模板"""
    code = data.get("template_code") or _next_code(db)
    columns = data.get("columns", [])
    # 根据报表类型修正字段名
    report_type = data.get("report_type", "")
    columns = _remap_field_keys(columns, report_type)
    columns_json = _columns_to_json(columns)

    # 如果设为默认，先清除同类型的其他默认
    if data.get("is_default"):
        db.query(ReportTemplate).filter(
            ReportTemplate.report_type == data["report_type"],
            ReportTemplate.is_default == True,
        ).update({"is_default": False})

    t = ReportTemplate(
        template_code=code,
        template_name=data["template_name"],
        report_type=data["report_type"],
        columns_json=columns_json,
        layout_json=json.dumps(data.get("layout"), ensure_ascii=False) if data.get("layout") else None,
        source_file_path=data.get("source_file_path"),
        group_by=data.get("group_by"),
        is_default=data.get("is_default", False),
        status="active",
        created_by=data.get("created_by", "admin"),
        remark=data.get("remark"),
    )
    db.add(t)
    db.commit()
    db.refresh(t)
    return _to_dict(t)


def update_template(db: Session, template_id: int, data: dict) -> Optional[Dict]:
    """更新模板"""
    t = db.query(ReportTemplate).filter(ReportTemplate.id == template_id).first()
    if not t:
        return None

    if "template_name" in data and data["template_name"] is not None:
        t.template_name = data["template_name"]
    if "columns" in data and data["columns"] is not None:
        cols = _remap_field_keys(data["columns"], t.report_type)
        t.columns_json = _columns_to_json(cols)
    if "layout" in data:
        t.layout_json = json.dumps(data["layout"], ensure_ascii=False) if data["layout"] else None
    if "group_by" in data:
        t.group_by = data["group_by"]
    if "remark" in data:
        t.remark = data["remark"]
    if "status" in data and data["status"] is not None:
        t.status = data["status"]

    if data.get("is_default"):
        db.query(ReportTemplate).filter(
            ReportTemplate.report_type == t.report_type,
            ReportTemplate.is_default == True,
            ReportTemplate.id != template_id,
        ).update({"is_default": False})
        t.is_default = True

    t.updated_at = datetime.now()
    db.commit()
    db.refresh(t)
    return _to_dict(t)


def delete_template(db: Session, template_id: int) -> bool:
    """软删除模板"""
    t = db.query(ReportTemplate).filter(ReportTemplate.id == template_id).first()
    if not t:
        return False
    t.status = "deleted"
    t.updated_at = datetime.now()
    db.commit()
    return True


def set_default(db: Session, template_id: int) -> Optional[Dict]:
    """设为默认模板"""
    t = db.query(ReportTemplate).filter(ReportTemplate.id == template_id).first()
    if not t or t.status != "active":
        return None

    db.query(ReportTemplate).filter(
        ReportTemplate.report_type == t.report_type,
        ReportTemplate.is_default == True,
    ).update({"is_default": False})

    t.is_default = True
    t.updated_at = datetime.now()
    db.commit()
    db.refresh(t)
    return _to_dict(t)


def parse_excel_headers(file_path: str, report_type: str = None) -> List[Dict]:
    """解析Excel文件的表头行，自动识别表头位置并映射到后端字段名。

    策略：
    1. 始终使用第一个 worksheet（不依赖 wb.active）
    2. 扫描前 20 行，找到非空单元格最多的行作为表头行
    3. 跳过合并标题行（只有 1-2 个非空单元格的行）
    4. 中文表头名自动映射到后端数据字段名（field_key）
    """
    from openpyxl.utils import get_column_letter

    wb = _load_workbook_any(file_path, read_only=False, data_only=True)
    if not wb.worksheets:
        wb.close()
        return []

    ws = wb.worksheets[0]

    # 提取列宽映射
    col_widths = {}
    try:
        for col_letter, dim in ws.column_dimensions.items():
            if dim.width:
                col_widths[col_letter] = int(dim.width * 8)  # Excel字符宽→像素约×8
    except Exception:
        pass

    # 扫描前 20 行，找最佳表头行
    best_row_idx = None
    best_count = 0
    sample_rows = []
    max_scan = min(20, ws.max_row or 20)
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan), 1):
        non_empty = [c for c in row if c.value is not None and str(c.value).strip()]
        count = len(non_empty)
        sample_rows.append((row_idx, row, count))
        if count >= 3 and count > best_count:
            best_count = count
            best_row_idx = row_idx

    if best_row_idx is None:
        for row_idx, row, count in sample_rows:
            if count >= 2:
                best_row_idx = row_idx
                break

    if best_row_idx is None:
        wb.close()
        return []

    # 取最佳行的数据，映射字段名
    target_row = sample_rows[best_row_idx - 1][1]
    headers = []
    used_field_keys = set()

    for col_idx, cell in enumerate(target_row, 1):
        if cell.value is not None and str(cell.value).strip():
            header_name = str(cell.value).strip()

            # 自动映射中文表头到后端字段名
            mapped_key = _match_field_key(header_name)
            if mapped_key:
                field_key = mapped_key
                # 处理重复字段名（如两个"单位简称"列）
                if field_key in used_field_keys:
                    field_key = f"{field_key}_{col_idx}"
            else:
                field_key = f"col_{col_idx}"

            used_field_keys.add(field_key)

            # 列宽：优先从Excel获取，否则默认
            col_letter = get_column_letter(col_idx)
            width = col_widths.get(col_letter, 120)

            headers.append({
                "field_key": field_key,
                "header_name": header_name,
                "width": width,
                "align": "right" if field_key in (
                    "opening_balance", "total_income", "total_expense", "net_change",
                    "ending_balance", "income", "expense", "day_balance", "prev_balance",
                    "income_amount", "expense_amount", "rolling_balance", "amount",
                    "period_income", "period_expense",
                ) else "center" if col_idx <= 2 else "left",
                "visible": True,
                "format": "money" if field_key in (
                    "opening_balance", "total_income", "total_expense", "net_change",
                    "ending_balance", "income", "expense", "day_balance", "prev_balance",
                    "income_amount", "expense_amount", "rolling_balance", "amount",
                    "period_income", "period_expense",
                ) else None,
                "sort_order": col_idx,
            })

    wb.close()
    return headers


# ── 占位符 → 字段名映射 ──────────────────────────────
PLACEHOLDER_MAP: Dict[str, str] = {
    "{{板块}}": "entity_name",
    "{{核算组织}}": "entity_name",
    "{{往来单位}}": "entity_name",
    "{{开户行}}": "account_bank",
    "{{账户信息}}": "account_name",
    "{{银行编号}}": "account_code",
    "{{核算方式}}": None,
    "{{月}}": "month",
    "{{日}}": "day",
    "{{月初余额}}": "prev_balance",
    "{{期初余额}}": "prev_balance",
    "{{摘要}}": "summary_text",
    "{{收入}}": "income",
    "{{支出}}": "expense",
    "{{余额}}": "day_balance",
    "{{本月收入小计}}": "income",
    "{{本月支出小计}}": "expense",
    "{{月末余额}}": "day_balance",
    "{{报表标题}}": "_report_title",
    "{{开始期间}}": "_start_date",
    "{{结束期间}}": "_end_date",
}


def parse_excel_layout(file_path: str) -> Dict:
    """解析Excel完整布局：标题区、信息区、表头、数据模板、汇总行。
    返回完整的布局信息，用于前端按Excel样式渲染报表。
    """
    from openpyxl.utils import get_column_letter
    import re

    wb = _load_workbook_any(file_path, read_only=False, data_only=True)
    if not wb.worksheets:
        wb.close()
        return {}
    ws = wb.worksheets[0]

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    # 列宽
    col_widths = []
    for ci in range(1, max_col + 1):
        cl = get_column_letter(ci)
        dim = ws.column_dimensions.get(cl)
        col_widths.append(int((dim.width or 8) * 8) if dim and dim.width else 64)

    # 合并单元格映射 → (row, col) → (merge_id, colspan, rowspan)
    merge_map: Dict[tuple, Dict] = {}
    merge_ranges = []
    for mr in ws.merged_cells.ranges:
        rmin, rmax = mr.min_row, mr.max_row
        cmin, cmax = mr.min_col, mr.max_col
        colspan = cmax - cmin + 1
        rowspan = rmax - rmin + 1
        merge_info = {
            "col_start": cmin - 1, "col_end": cmax - 1,
            "row_start": rmin - 1, "row_end": rmax - 1,
            "colspan": colspan, "rowspan": rowspan,
        }
        merge_ranges.append(merge_info)
        for r in range(rmin, rmax + 1):
            for c in range(cmin, cmax + 1):
                merge_map[(r, c)] = merge_info

    # 找表头行（非空单元格最多的行，且非合并标题行）
    header_row_idx = None
    best_count = 0
    for row_idx in range(1, min(max_row + 1, 21)):
        non_empty = 0
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is not None and str(val).strip() and not str(val).strip().startswith("{{"):
                non_empty += 1
        if non_empty >= 3 and non_empty > best_count:
            # 排除合并标题行（整行合并的情况）
            merge = merge_map.get((row_idx, 1))
            if not merge or merge.get("colspan", 1) < max_col:
                best_count = non_empty
                header_row_idx = row_idx

    # 解析所有行
    rows = []
    for row_idx in range(1, max_row + 1):
        cells = []
        skip_cols = set()
        for col_idx in range(1, max_col + 1):
            if col_idx in skip_cols:
                continue

            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            text = str(val).strip() if val is not None else ""

            # 检查是否是被合并的单元格（非左上角）
            merge = merge_map.get((row_idx, col_idx))
            if merge and (row_idx != merge["row_start"] + 1 or col_idx != merge["col_start"] + 1):
                # 被合并的从属单元格，跳过
                skip_cols.add(col_idx)
                continue

            colspan = 1
            rowspan = 1
            if merge and row_idx == merge["row_start"] + 1 and col_idx == merge["col_start"] + 1:
                colspan = merge["colspan"]
                rowspan = merge["rowspan"]
                for dr in range(merge["row_start"] + 1, merge["row_end"] + 2):
                    for dc in range(merge["col_start"] + 1, merge["col_end"] + 2):
                        if dr != row_idx or dc != col_idx:
                            skip_cols.add(dc)

            # 检测占位符
            is_placeholder = bool(re.match(r'^\{\{.*\}\}$', text))
            field_key = None
            if is_placeholder:
                field_key = PLACEHOLDER_MAP.get(text)

            cells.append({
                "col": col_idx - 1,  # 0-indexed
                "colspan": colspan,
                "rowspan": rowspan,
                "text": text,
                "is_placeholder": is_placeholder,
                "field_key": field_key,
            })

        # 判断行类型
        has_placeholder = any(c["is_placeholder"] for c in cells)
        has_non_empty = any(c["text"] for c in cells)

        if row_idx == header_row_idx:
            row_type = "header"
        elif has_placeholder:
            row_type = "data"
        elif has_non_empty and row_idx < (header_row_idx or 999):
            row_type = "title"
        elif has_non_empty and "小计" in " ".join(c["text"] for c in cells):
            row_type = "subtotal"
        elif has_non_empty and row_idx > (header_row_idx or 0):
            row_type = "info"
        elif has_non_empty:
            row_type = "title"
        else:
            row_type = "empty"

        rows.append({"row": row_idx, "type": row_type, "cells": cells})

    # 重新分类行（更精确）
    for r in rows:
        texts = " ".join(c["text"] for c in r["cells"] if c["text"])
        has_ph = any(c["is_placeholder"] for c in r["cells"])

        if r["row"] == header_row_idx:
            r["type"] = "header"
        elif "小计" in texts or "合计" in texts or "总计" in texts:
            r["type"] = "subtotal"
        elif has_ph and r["row"] < (header_row_idx or 999):
            r["type"] = "info"
        elif has_ph:
            r["type"] = "data"
        elif texts and r["row"] < (header_row_idx or 999):
            r["type"] = "title"
        elif texts and "说明" in texts:
            r["type"] = "instruction"
        elif texts:
            r["type"] = "info"
        else:
            r["type"] = "empty"

    # 找小计行位置
    subtotal_rows = [r for r in rows if r["type"] == "subtotal"]
    first_subtotal_row = subtotal_rows[0]["row"] if subtotal_rows else max_row

    # 剔除空行和说明行，只保留有意义的行
    meaningful_rows = []
    for r in rows:
        if r["row"] > first_subtotal_row:
            break  # 小计之后的不保留（第二个块、说明行等）
        if r["type"] in ("title", "info", "header", "data", "subtotal"):
            # 对于数据行中的空行（没有占位符也没有文本的），跳过
            if r["type"] == "data" and not any(c["is_placeholder"] or c["text"] for c in r["cells"]):
                continue
            meaningful_rows.append(r)

    # 找左侧合并列（cash journal中 A-E 列在数据块中是合并的）
    merge_left_cols = []
    if header_row_idx:
        for r in rows:
            if r["row"] == header_row_idx + 1:  # 第一行数据行
                for c in r["cells"]:
                    if c.get("rowspan", 1) > 3 and c["col"] < max_col // 2:
                        merge_left_cols.append(c["col"])
                break

    wb.close()

    return {
        "col_count": max_col,
        "col_widths": col_widths,
        "header_row_idx": (header_row_idx or 1) - 1,
        "merge_left_cols": merge_left_cols,
        "rows": meaningful_rows,
    }


def _to_dict(t: ReportTemplate) -> Dict:
    layout = None
    if t.layout_json:
        try:
            layout = json.loads(t.layout_json)
        except json.JSONDecodeError:
            layout = None
    return {
        "id": t.id,
        "template_code": t.template_code,
        "template_name": t.template_name,
        "report_type": t.report_type,
        "columns": _json_to_columns(t.columns_json),
        "layout": layout,
        "group_by": t.group_by,
        "is_default": t.is_default,
        "status": t.status,
        "created_by": t.created_by,
        "remark": t.remark,
        "created_at": t.created_at.isoformat() if t.created_at else None,
        "updated_at": t.updated_at.isoformat() if t.updated_at else None,
    }
