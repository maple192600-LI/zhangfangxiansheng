"""银行对账单身份线索提取服务

从银行对账单文件（xlsx）和文件名中提取银行名称、账号、户名、单位名、开户行等身份线索。
输出供 bank_account_match_service 使用。

契约锚点：docs/14_BANK_IMPORT_GENERALIZATION.md
"""
import re
from typing import Optional

from openpyxl import load_workbook


_BANK_TEXT_RE = re.compile(r"([一-鿿]+银行)")

_ACCOUNT_NUMBER_KW = ["账号", "账户号码", "账户号", "帐号", "Account"]
_ACCOUNT_HOLDER_KW = ["户名", "账户名称", "账户名", "户名全称"]
_ENTITY_NAME_KW = ["单位名称", "公司名称"]
_BRANCH_KW = ["开户行", "开户银行", "开户支行"]

_MAX_SCAN_ROWS = 30
_MAX_SCAN_COLS = 20
_MAX_CANDIDATES = 80


def _add_candidate(candidates: list, text: str):
    text = str(text).strip()
    if not text:
        return
    if len(text) > 120:
        text = text[:120]
    if text not in candidates:
        candidates.append(text)


def extract_identity_hints(file_path: str, filename: Optional[str] = None) -> dict:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        return extract_identity_hints_from_workbook(wb, filename)
    finally:
        wb.close()


def extract_identity_hints_from_workbook(wb, filename: Optional[str] = None) -> dict:
    hints = {
        "account_number": "",
        "account_last_four": "",
        "account_name": "",
        "entity_name": "",
        "bank_name": "",
        "branch_name": "",
        "filename_hint": "",
    }
    evidence = {"cells": [], "headers": [], "filename": filename or ""}
    candidates = []

    if filename:
        _add_candidate(candidates, filename)
        fn_base = filename.rsplit(".", 1)[0] if "." in filename else filename
        if fn_base != filename:
            _add_candidate(candidates, fn_base)
        fn_bank = _extract_bank_text(filename)
        if fn_bank:
            hints["bank_name"] = fn_bank
        fn_last4 = _extract_last_four_from_filename(filename)
        if fn_last4:
            hints["filename_hint"] = fn_last4

    for sheet in wb.worksheets:
        _scan_sheet(sheet, hints, evidence, candidates)
        break

    _add_candidate(candidates, hints["bank_name"])
    _add_candidate(candidates, hints["branch_name"])

    if hints["account_number"] and not hints["account_last_four"]:
        digits = re.sub(r"\D", "", hints["account_number"])
        if len(digits) >= 4:
            hints["account_last_four"] = digits[-4:]

    return {
        "bank_hint": hints["bank_name"],
        "format_fingerprint": _fingerprint(evidence),
        "identity_hints": hints,
        "evidence": evidence,
        "confidence": _confidence(hints),
        "bank_text_candidates": candidates[:_MAX_CANDIDATES],
    }


def _scan_sheet(sheet, hints: dict, evidence: dict, candidates: list):
    max_row = min(_MAX_SCAN_ROWS, sheet.max_row or _MAX_SCAN_ROWS)
    max_col = min(_MAX_SCAN_COLS, sheet.max_column or _MAX_SCAN_COLS)

    cells_text = []
    for row_idx in range(1, max_row + 1):
        row_vals = []
        prev_text = ""
        for col_idx in range(1, max_col + 1):
            val = sheet.cell(row=row_idx, column=col_idx).value
            if val is None:
                prev_text = ""
                continue
            text = str(val).strip()
            if not text:
                prev_text = ""
                continue
            row_vals.append(text)
            _add_candidate(candidates, text)
            _extract_from_cell(text, hints)
            if prev_text:
                _extract_adjacent(prev_text, text, hints)
            prev_text = text
        if row_vals:
            cells_text.append(row_vals)

    evidence["cells"] = cells_text[:10]

    for row_idx in range(1, min(6, max_row + 1)):
        headers = []
        for col_idx in range(1, max_col + 1):
            v = sheet.cell(row=row_idx, column=col_idx).value
            if v and str(v).strip():
                headers.append(str(v).strip())
        if len(headers) >= 3:
            evidence["headers"] = headers
            for h in headers:
                _add_candidate(candidates, h)
            break


def _extract_from_cell(text: str, hints: dict):
    if not hints["bank_name"]:
        bank = _extract_bank_text(text)
        if bank:
            hints["bank_name"] = bank

    if not hints["account_number"]:
        acc = _extract_account_number(text)
        if acc:
            hints["account_number"] = acc

    if not hints["account_name"]:
        v = _after_keyword(text, _ACCOUNT_HOLDER_KW)
        if v:
            hints["account_name"] = v

    if not hints["entity_name"]:
        v = _after_keyword(text, _ENTITY_NAME_KW)
        if v:
            hints["entity_name"] = v

    if not hints["branch_name"]:
        v = _after_keyword(text, _BRANCH_KW)
        if v:
            hints["branch_name"] = v

    if not hints["account_last_four"] and not hints["account_number"]:
        last4 = _standalone_last_four(text)
        if last4:
            hints["account_last_four"] = last4


def _extract_adjacent(prev_text: str, current_text: str, hints: dict):
    pairs = [
        (_ACCOUNT_NUMBER_KW, "account_number", True),
        (_ACCOUNT_HOLDER_KW, "account_name", False),
        (_ENTITY_NAME_KW, "entity_name", False),
        (_BRANCH_KW, "branch_name", False),
    ]
    for kw_list, key, is_number in pairs:
        if hints[key]:
            continue
        stripped = prev_text.strip().rstrip("：: ")
        for kw in kw_list:
            if stripped == kw:
                value = current_text.strip()
                if value and len(value) <= 200:
                    if is_number:
                        digits = re.sub(r"[\s\-]", "", value)
                        if _is_plausible_account(digits):
                            hints[key] = digits
                    else:
                        hints[key] = value
                break


def _extract_bank_text(text: str) -> str:
    """Extract raw bank name text containing '银行'. No normalization —
    the match service resolves against banks master data."""
    m = _BANK_TEXT_RE.search(text)
    if m:
        return m.group(1)
    return ""


def _extract_account_number(text: str) -> str:
    for kw in _ACCOUNT_NUMBER_KW:
        idx = text.find(kw)
        if idx >= 0:
            remainder = text[idx + len(kw):]
            match = re.search(r"[\d\s\-]{10,25}", remainder)
            if match:
                digits = re.sub(r"[\s\-]", "", match.group())
                if _is_plausible_account(digits):
                    return digits
    digits = re.sub(r"[\s\-]", "", text)
    if _is_plausible_account(digits):
        return digits
    return ""


def _is_plausible_account(digits: str) -> bool:
    if len(digits) < 10 or len(digits) > 25:
        return False
    if re.match(r"^(19|20)\d{6}$", digits):
        return False
    return True


def _after_keyword(text: str, keywords: list) -> str:
    for kw in keywords:
        for sep in ["：", ":"]:
            idx = text.find(kw + sep)
            if idx >= 0:
                value = text[idx + len(kw) + len(sep):].strip()
                value = value.split("：")[0].split(":")[0].strip()
                if value and len(value) <= 200:
                    return value
        if text.startswith(kw):
            value = text[len(kw):].strip()
            value = value.split("：")[0].split(":")[0].strip()
            if value and len(value) <= 200:
                return value
    return ""


def _standalone_last_four(text: str) -> str:
    t = text.strip()
    m = re.match(r"^尾号\s*(\d{4})$", t)
    if m:
        return m.group(1)
    if re.match(r"^\d{4}$", t):
        return t
    return ""


def _extract_last_four_from_filename(filename: str) -> str:
    m = re.search(r"(\d{4})\.\w+$", filename)
    if m:
        return m.group(1)
    m = re.search(r"尾号\s*(\d{4})", filename)
    if m:
        return m.group(1)
    return ""


def _confidence(hints: dict) -> float:
    filled = sum(1 for v in hints.values() if v)
    if filled == 0:
        return 0.0
    if hints.get("account_number"):
        return min(0.8 + 0.05 * (filled - 1), 0.95)
    if hints.get("account_last_four") and hints.get("bank_name"):
        return 0.6
    if hints.get("entity_name") or hints.get("account_name"):
        return 0.4
    return 0.2


def _fingerprint(evidence: dict) -> str:
    headers = evidence.get("headers", [])
    if not headers:
        return "unknown"
    key = "|".join(h[:10] for h in headers[:5])
    return f"fp_{abs(hash(key)) % 10000:04d}"
