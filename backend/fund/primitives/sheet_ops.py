"""sheet_ops · §P1 · 6 个 Excel 表格操作基元

契约锚点：docs/30_contracts/25_primitives_whitelist.md §P1

| 函数              | 输入                              | 输出                       |
|-------------------|-----------------------------------|----------------------------|
| read_sheet        | wb, index/name                    | Worksheet                  |
| detect_header_row | sheet, max_scan                   | int (1-based)              |
| extract_headers   | sheet, header_row                 | {表头文本: col_index_1based} |
| iter_data_rows    | sheet, start_row                  | Iterator[{col_index: value}] |
| is_empty_row      | row (dict)                        | bool                       |
| locate_merged_cells | sheet                           | list[MergedCellRange]      |
"""
from __future__ import annotations

from typing import Iterator, Optional

from openpyxl import Workbook
from openpyxl.worksheet.merge import MergedCellRange
from openpyxl.worksheet.worksheet import Worksheet


def read_sheet(wb: Workbook, index: int = 0, name: Optional[str] = None) -> Worksheet:
    """按索引或名称取工作表。

    Runtime 预加载 wb；Agent 不得自己调 openpyxl.load_workbook。
    优先级：name > index。
    """
    if name is not None:
        if name not in wb.sheetnames:
            raise KeyError(f"工作表 {name!r} 不存在；可选：{wb.sheetnames!r}")
        return wb[name]
    sheets = wb.sheetnames
    if not sheets:
        raise IndexError("工作簿无任何工作表")
    if index < 0 or index >= len(sheets):
        raise IndexError(f"工作表索引越界 {index}（共 {len(sheets)} 张）")
    return wb[sheets[index]]


def detect_header_row(sheet: Worksheet, max_scan: int = 20) -> int:
    """扫描前 max_scan 行，返回"字段最多"的一行（1-based openpyxl 约定）。

    平分时取靠前的那行；整表无数据返回 1。
    """
    if max_scan <= 0:
        return 1
    best_row = 1
    best_count = -1
    for i, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=max_scan, values_only=True),
        start=1,
    ):
        filled = sum(
            1
            for v in row
            if v is not None and (not isinstance(v, str) or v.strip())
        )
        if filled > best_count:
            best_count = filled
            best_row = i
    return best_row


def extract_headers(sheet: Worksheet, header_row: int) -> dict[str, int]:
    """取表头行为 {去空白表头文本: col_index_1based}。

    - 空单元格跳过
    - 重复表头保留第一个出现的列
    """
    headers: dict[str, int] = {}
    for col_idx, cell in enumerate(sheet[header_row], start=1):
        val = cell.value
        if val is None:
            continue
        text = str(val).strip()
        if not text:
            continue
        if text in headers:
            continue
        headers[text] = col_idx
    return headers


def iter_data_rows(sheet: Worksheet, start_row: int) -> Iterator[dict]:
    """从 start_row（1-based）开始迭代数据行。

    yield：{col_index_1based: value}，caller 自己按 header 字典映射。
    """
    for row in sheet.iter_rows(min_row=start_row, values_only=False):
        row_dict: dict[int, object] = {}
        for cell in row:
            row_dict[cell.column] = cell.value
        yield row_dict


def is_empty_row(row: dict) -> bool:
    """全空 / 分隔行判定：所有 value ∈ {None, 空白字符串} → True。"""
    for v in row.values():
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return False
    return True


def locate_merged_cells(sheet: Worksheet) -> list[MergedCellRange]:
    """返回合并单元格区域列表（用于处理合计行/标题行）。"""
    return list(sheet.merged_cells.ranges)
