"""template_fill · §P7 · 6 个报表模板填写基元

契约锚点：docs/30_contracts/25_primitives_whitelist.md §P7

| 函数              | 职责                              |
|-------------------|-----------------------------------|
| load_template     | 只读加载 Excel 空模板             |
| fill              | 把 value 填到 ${ph} / {{ph}} 位置 |
| const             | 常量占位                          |
| date_range_start  | 从 ctx 读期间起                   |
| date_range_end    | 从 ctx 读期间止                   |
| format_amount     | 金额格式化（千分位 + 固定位数）   |
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Optional

from openpyxl import Workbook, load_workbook


# §P7 · 占位符语法（两种）
_PLACEHOLDER_STYLES: tuple[tuple[str, str], ...] = (
    ("${", "}"),
    ("{{", "}}"),
)


def load_template(path: str) -> Workbook:
    """加载 Excel 模板（空白 wb 原样读入；不 data_only，保留公式）。"""
    return load_workbook(filename=path, keep_vba=False, data_only=False)


def _iter_placeholder_hits(wb: Workbook, placeholder: str):
    """遍历 wb 所有 sheet 所有 cell，返回命中 ${ph} / {{ph}} 的单元格。

    Yields: (cell, original_text, token)
    """
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                for prefix, suffix in _PLACEHOLDER_STYLES:
                    token = f"{prefix}{placeholder}{suffix}"
                    if token in v:
                        yield cell, v, token
                        break


def fill(
    wb: Workbook,
    placeholder: str,
    value: Any,
    format: Optional[str] = None,  # noqa: A002 (契约签名)
) -> None:
    """把 value 填到所有命中 ${placeholder} / {{placeholder}} 的单元格。

    - 若 format 给定 → 设置 number_format
    - value 为 Decimal → 自动转 float
    - 占位符是单元格唯一内容 → 原值替换（保留类型，比如 date）
    - 占位符混在文本中 → 字符串替换
    - 无任何命中 → 静默（Agent 层 check_placeholder_binding 验证完整性）
    """
    if not placeholder:
        raise ValueError("placeholder 不可为空")
    hits = list(_iter_placeholder_hits(wb, placeholder))
    if not hits:
        return

    if isinstance(value, Decimal):
        final_value: Any = float(value)
    else:
        final_value = value

    for cell, original, token in hits:
        if original == token:
            cell.value = final_value
        else:
            cell.value = original.replace(token, str(final_value))
        if format:
            cell.number_format = format


def const(value: Any) -> Any:
    """常量占位：原样返回 value。"""
    return value


def _coerce_date(v: Any) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        # 延迟导入避免 primitives 内循环
        from .value_parsers import parse_date

        return parse_date(v)
    return None


def date_range_start(ctx: dict) -> date:
    """从 ctx 读期间开始；keys: start_date | period_start。"""
    for key in ("start_date", "period_start"):
        d = _coerce_date(ctx.get(key))
        if d is not None:
            return d
    raise KeyError(
        f"ctx 缺少 start_date / period_start；keys={list(ctx.keys())}"
    )


def date_range_end(ctx: dict) -> date:
    """从 ctx 读期间结束；keys: end_date | period_end。"""
    for key in ("end_date", "period_end"):
        d = _coerce_date(ctx.get(key))
        if d is not None:
            return d
    raise KeyError(
        f"ctx 缺少 end_date / period_end；keys={list(ctx.keys())}"
    )


def format_amount(value, digits: int = 2) -> str:
    """格式化金额：千分位 + 固定小数位。

    - None → ''
    - 非数值（Decimal 失败） → str(value)
    - digits < 0 → ValueError
    """
    if value is None:
        return ""
    if digits < 0:
        raise ValueError(f"digits 不得为负：{digits}")
    if not isinstance(value, Decimal):
        try:
            value = Decimal(str(value))
        except (InvalidOperation, ValueError):
            return str(value)
    q = Decimal(10) ** -digits if digits > 0 else Decimal(1)
    quantized = value.quantize(q)
    sign = "-" if quantized < 0 else ""
    absval = -quantized if quantized < 0 else quantized
    int_str, _, dec_str = f"{absval:f}".partition(".")
    rev = int_str[::-1]
    grouped = ",".join(rev[i : i + 3] for i in range(0, len(rev), 3))
    int_str_grouped = grouped[::-1]
    if dec_str:
        return f"{sign}{int_str_grouped}.{dec_str}"
    return f"{sign}{int_str_grouped}"
