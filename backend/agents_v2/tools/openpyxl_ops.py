"""openpyxl 读写工具 — 读取 Excel 文件内容"""
import os

from agents_v2.tool_registry import register_tool, ToolContext
from agents_v2.workspace import safe_path


@register_tool(read_only=True)
def openpyxl_read(path: str, sheet_name: str = None, max_rows: int = 100, ctx: ToolContext = None) -> dict:
    """读取工作区内的 Excel 文件（xlsx/xls），返回二维表格数据。path 为相对路径。"""
    abs_path = safe_path(ctx.agent_code, path)
    if not abs_path or not os.path.isfile(abs_path):
        return {"ok": False, "error": f"文件不存在: {path}"}

    ext = os.path.splitext(abs_path)[1].lower()
    if ext == ".csv":
        return _read_csv(abs_path, max_rows)
    if ext in (".xlsx", ".xls"):
        return _read_excel(abs_path, sheet_name, max_rows)
    return {"ok": False, "error": f"不支持的文件格式: {ext}"}


def _read_excel(filepath: str, sheet_name: str, max_rows: int) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"ok": False, "error": f"Sheet 不存在: {sheet_name}", "sheets": wb.sheetnames}
        ws = wb[sheet_name]
    else:
        ws = wb.active
        sheet_name = ws.title

    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        rows.append([str(c) if c is not None else "" for c in row])

    wb.close()
    headers = rows[0] if rows else []
    return {
        "ok": True,
        "sheet_name": sheet_name,
        "sheets": wb.sheetnames if hasattr(wb, "sheetnames") else [sheet_name],
        "headers": headers,
        "rows": rows[1:] if len(rows) > 1 else [],
        "total_data_rows": len(rows) - 1 if rows else 0,
        "truncated": len(rows) >= max_rows,
    }


def _read_csv(filepath: str, max_rows: int) -> dict:
    import csv
    rows = []
    with open(filepath, "r", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows.append(row)

    headers = rows[0] if rows else []
    return {
        "ok": True,
        "sheet_name": None,
        "headers": headers,
        "rows": rows[1:] if len(rows) > 1 else [],
        "total_data_rows": len(rows) - 1 if rows else 0,
        "truncated": False,
    }
