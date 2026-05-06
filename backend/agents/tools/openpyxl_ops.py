"""openpyxl 读写工具 — 读取/写入 Excel 文件"""
import os
import tempfile

from agents.tool_registry import register_tool, ToolContext
from agents.workspace import safe_path, get_agent_root


@register_tool(read_only=True)
def openpyxl_read(path: str, sheet_name: str = None, max_rows: int = 100, ctx: ToolContext = None) -> dict:
    """读取工作区内的 Excel 文件，返回结构化的二维表格数据。

    使用场景：
    - 读取银行流水 Excel 文件（file_parse 适合纯文本提取，本工具适合结构化读取）
    - 读取报表模板，获取精确的单元格数据
    - 需要知道 sheet 名称、表头、行数据时

    参数：
    - path: 必需，相对路径
    - sheet_name: 可选，指定读取哪个 sheet，不填读第一个
    - max_rows: 可选，最大读取行数，默认 100

    返回：{"ok": true, "sheet_name": "Sheet名", "sheets": ["Sheet1", ...], "headers": [...], "rows": [[...], ...]}
    """
    abs_path = safe_path(ctx.agent_code, path)
    if not abs_path or not os.path.isfile(abs_path):
        return {"ok": False, "error": f"文件不存在: {path}"}

    ext = os.path.splitext(abs_path)[1].lower()
    if ext == ".csv":
        return _read_csv(abs_path, max_rows)
    if ext in (".xlsx", ".xls"):
        return _read_excel(abs_path, sheet_name, max_rows)
    return {"ok": False, "error": f"不支持的文件格式: {ext}"}


def _read_excel(filepath: str, sheet_name: str, max_rows: int) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"ok": False, "error": f"Sheet 不存在: {sheet_name}", "sheets": wb.sheetnames}
        ws = wb[sheet_name]
    else:
        ws = wb.active
        sheet_name = ws.title

    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        rows.append([str(c) if c is not None else "" for c in row])

    wb.close()
    headers = rows[0] if rows else []
    return {
        "ok": True,
        "sheet_name": sheet_name,
        "sheets": wb.sheetnames if hasattr(wb, "sheetnames") else [sheet_name],
        "headers": headers,
        "rows": rows[1:] if len(rows) > 1 else [],
        "total_data_rows": len(rows) - 1 if rows else 0,
        "truncated": len(rows) >= max_rows,
    }


def _read_csv(filepath: str, max_rows: int) -> dict:
    import csv
    rows = []
    with open(filepath, "r", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows.append(row)

    headers = rows[0] if rows else []
    return {
        "ok": True,
        "sheet_name": None,
        "headers": headers,
        "rows": rows[1:] if len(rows) > 1 else [],
        "total_data_rows": len(rows) - 1 if rows else 0,
        "truncated": False,
    }


@register_tool(read_only=False)
def openpyxl_write(path: str, headers: list, rows: list, sheet_name: str = "Sheet1", ctx: ToolContext = None) -> dict:
    """将二维数据写入 Excel 文件。如文件已存在则覆盖创建新文件。

    使用场景：生成导出报表、创建数据汇总 Excel。

    参数：
    - path: 必需，输出文件相对路径
    - headers: 必需，表头列表，如 ["日期", "摘要", "收入", "支出"]
    - rows: 必需，数据行列表，如 [["2025-01-01", "测试", 100, 0]]
    - sheet_name: 可选，Sheet 名称，默认 "Sheet1"
    """
    import openpyxl
    abs_path = safe_path(ctx.agent_code, path)
    if not abs_path:
        return {"ok": False, "error": f"非法路径: {path}"}

    # 确保目录存在
    os.makedirs(os.path.dirname(abs_path), exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 写表头
    if headers:
        ws.append(headers)

    # 写数据行
    for row in rows:
        ws.append(row)

    wb.save(abs_path)
    wb.close()
    return {"ok": True, "path": path, "rows_written": len(rows), "columns": len(headers) if headers else 0}


@register_tool(read_only=False)
def openpyxl_edit(path: str, cells: dict, ctx: ToolContext = None) -> dict:
    """编辑现有 Excel 文件中的指定单元格。用于填充报表模板中的占位符。

    使用场景：
    - 填充报表模板（将占位符替换为实际数据）
    - 修改现有报表中的特定单元格

    参数：
    - path: 必需，Excel 文件相对路径
    - cells: 必需，单元格修改字典，键为单元格坐标（如 "B3"），值为要写入的内容（字符串或数字）
      示例：{"B3": "2025年1月", "C5": 10000.00, "D5": "=C5*0.13"}

    返回：{"ok": true, "path": "路径", "cells_updated": 修改的单元格数量}
    """
    import openpyxl
    abs_path = safe_path(ctx.agent_code, path)
    if not abs_path or not os.path.isfile(abs_path):
        return {"ok": False, "error": f"文件不存在: {path}"}

    try:
        wb = openpyxl.load_workbook(abs_path)
        ws = wb.active

        updated = 0
        for cell_ref, value in cells.items():
            try:
                ws[cell_ref] = value
                updated += 1
            except Exception:
                pass

        wb.save(abs_path)
        wb.close()
        return {"ok": True, "path": path, "cells_updated": updated}
    except Exception as e:
        return {"ok": False, "error": f"编辑 Excel 失败: {e}"}
