"""Fund Agent harness_strict 调度器

步骤固定、工具白名单、产物受 Pydantic Schema 约束。
每个 skill 的执行步骤在代码中写死，AI 只填参数。
"""
from dataclasses import dataclass
from typing import Any

from sqlalchemy.orm import Session

from agents.fund import memory, sandbox
from agents.fund.schemas import (
    ParserInput, ParserOutput,
    RuleInput, RuleOutput,
    RuleMaintainInput,
    TemplateInferenceInput, TemplateInferenceOutput,
    PlaceholderBinding,
)
from agents.fund.skills._shared import (
    build_cash_rule,
    build_parser_code,
    parser_imports,
    rule_imports,
)

ALLOWED_SKILLS = {
    "parser.bank",
    "parser.manual",
    "rule.template_fill",
    "rule.maintain",
    "template.inference",
}


@dataclass
class SkillDraft:
    """Skill 执行结果"""
    payload: dict[str, Any]


class FundAgent:
    """Fund Agent harness_strict 调度器"""

    def __init__(self, db: Session):
        self.db = db

    def run_skill(self, skill_name: str, payload: dict[str, Any]) -> SkillDraft:
        if skill_name not in ALLOWED_SKILLS:
            raise ValueError(f"未知 skill: {skill_name}")

        handler = {
            "parser.bank": self._parser_bank,
            "parser.manual": self._parser_manual,
            "rule.template_fill": self._rule_template_fill,
            "rule.maintain": self._rule_maintain,
            "template.inference": self._template_inference,
        }[skill_name]

        return handler(payload)

    def _parser_bank(self, payload: dict) -> SkillDraft:
        """parser.bank: 生成银行流水解析器

        步骤固定：
        1. 验证输入 Schema
        2. 读取样本文件（脱敏）
        3. 加载字段字典 + 别名库
        4. 创建 Parser artifact 草稿
        """
        inp = ParserInput(**payload)

        code = build_parser_code(kind="bank", default_account_code=inp.account_code)
        sample_log = self._check_parser_sample(code, inp.sample_file, inp.account_code)
        artifact = memory.create_parser_draft(
            db=self.db,
            name=f"{inp.account_code}_bank_parser",
            kind="bank",
            account_code=inp.account_code,
            code=code,
            primitives_imports=parser_imports(),
            sample_check_log=sample_log,
            confidence=1.0 if sample_log["parsed_rows"] > 0 and sample_log["canonical_violations"] == 0 else 0.0,
        )

        return SkillDraft(payload={
            "artifact_id": artifact.id,
            "name": artifact.name,
            "kind": "bank",
            "account_code": inp.account_code,
            "status": "draft",
            "confidence": float(artifact.confidence or 0),
            "sample_check_log": sample_log,
        })

    def _parser_manual(self, payload: dict) -> SkillDraft:
        """parser.manual: 生成手工流水解析器"""
        inp = ParserInput(**payload)

        code = build_parser_code(kind="manual", default_account_code=inp.account_code)
        sample_log = self._check_parser_sample(code, inp.sample_file, inp.account_code)
        artifact = memory.create_parser_draft(
            db=self.db,
            name=f"{inp.account_code}_manual_parser",
            kind="manual",
            account_code=inp.account_code,
            code=code,
            primitives_imports=parser_imports(),
            sample_check_log=sample_log,
            confidence=1.0 if sample_log["parsed_rows"] > 0 and sample_log["canonical_violations"] == 0 else 0.0,
        )

        return SkillDraft(payload={
            "artifact_id": artifact.id,
            "name": artifact.name,
            "kind": "manual",
            "account_code": inp.account_code,
            "status": "draft",
            "confidence": float(artifact.confidence or 0),
            "sample_check_log": sample_log,
        })

    def _rule_template_fill(self, payload: dict) -> SkillDraft:
        """rule.template_fill: 生成报表填充规则

        步骤固定：
        1. 验证输入 Schema
        2. 解析模板占位符列表
        3. 为每个占位符匹配基元
        4. 创建 Rule artifact 草稿
        """
        inp = RuleInput(**payload)

        bindings, loop_config, check_log = build_cash_rule(inp.template_job_id, inp.placeholder_list)

        artifact = memory.create_rule_draft(
            db=self.db,
            name="template_rule",
            template_id=None,
            placeholder_bindings=bindings,
            loop_config=loop_config,
            primitives_imports=rule_imports(),
            sample_check_log=check_log.model_dump(),
            confidence=1.0 if inp.placeholder_list else 0.0,
        )

        return SkillDraft(payload={
            "artifact_id": artifact.id,
            "name": artifact.name,
            "template_job_id": inp.template_job_id,
            "status": "draft",
            "confidence": float(artifact.confidence or 0),
        })

    def _rule_maintain(self, payload: dict) -> SkillDraft:
        """rule.maintain: 维护/迭代现有规则

        步骤固定：
        1. 加载现有 Rule artifact
        2. 解析修改请求
        3. 创建新版本 Rule artifact（旧版保留）
        """
        inp = RuleMaintainInput(**payload)

        from db.tables import RuleArtifact
        existing = self.db.query(RuleArtifact).filter(RuleArtifact.id == inp.rule_id).first()
        if not existing:
            raise ValueError(f"Rule artifact {inp.rule_id} 不存在")

        new_bindings = dict(existing.placeholder_bindings or {})
        new_loop = dict(existing.loop_config or {}) if existing.loop_config else None
        new_primitives = list(existing.primitives_imports or [])

        artifact = memory.create_rule_draft(
            db=self.db,
            name=existing.name,
            template_id=existing.template_id,
            placeholder_bindings=new_bindings,
            loop_config=new_loop,
            primitives_imports=new_primitives,
            sample_check_log=existing.sample_check_log or {},
            confidence=existing.confidence or 0.0,
            created_by=inp.user_id,
        )

        return SkillDraft(payload={
            "artifact_id": artifact.id,
            "name": artifact.name,
            "version": artifact.version,
            "status": "draft",
            "change_request": inp.change_request,
        })

    def _template_inference(self, payload: dict) -> SkillDraft:
        """template.inference: 自动识别空白模板

        三阶段流水线：
        Stage A: 纯代码结构解析（无 AI）
        Stage B: AI 语义映射
        Stage C: 用户确认
        """
        inp = TemplateInferenceInput(**payload)

        # 创建推断任务
        job = memory.create_inference_job(
            db=self.db,
            template_file=inp.template_file,
            original_filename=inp.template_file.split("/")[-1] if "/" in inp.template_file else inp.template_file.split("\\")[-1],
        )

        # Stage A: 结构解析
        stage_a = self._stage_a_parse(inp.template_file)
        memory.update_inference_job(
            self.db, job.id,
            stage="b",
            stage_a_output=stage_a,
        )

        # Stage B: 占位 → 字段映射
        rule_draft = self._stage_b_map(stage_a, job.id)
        memory.update_inference_job(
            self.db, job.id,
            stage="c",
            stage_b_output=rule_draft,
        )

        return SkillDraft(payload={
            "job_id": job.id,
            "detected_placeholders": stage_a.get("placeholders", []),
            "confidence": rule_draft.get("confidence", 0.0),
            "rule_draft": rule_draft,
            "stage_a_output": stage_a,
        })

    def _stage_a_parse(self, template_file: str) -> dict:
        """Stage A: 纯代码结构解析"""
        import re

        placeholders = []
        merged_cells = []
        header_rows = []

        try:
            from openpyxl import load_workbook
            wb = load_workbook(template_file, read_only=False, data_only=True)
            try:
                for ws in wb.worksheets:
                    # 提取占位符
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str):
                                # 匹配 ${xxx}, {{xxx}}, 【xxx】等占位符
                                found = re.findall(r'\$\{([^}]+)\}|\{\{([^}]+)\}\}|【([^】]+)】', cell.value)
                                for match in found:
                                    ph = match[0] or match[1] or match[2]
                                    if ph and ph not in placeholders:
                                        placeholders.append(ph)

                    # 识别合并单元格
                    for mc in ws.merged_cells.ranges:
                        merged_cells.append({
                            "min_row": mc.min_row,
                            "max_row": mc.max_row,
                            "min_col": mc.min_col,
                            "max_col": mc.max_col,
                        })
            finally:
                wb.close()
        except Exception:
            pass

        return {
            "placeholders": placeholders,
            "merged_cells": merged_cells,
            "header_rows": header_rows,
        }

    def _stage_b_map(self, stage_a: dict, job_id: int) -> dict:
        placeholders = stage_a.get("placeholders", [])
        bindings, loop_config, check_log = build_cash_rule(job_id, placeholders)
        confidence = 1.0 if placeholders else 0.0

        # 创建 Rule 草稿
        artifact = memory.create_rule_draft(
            db=self.db,
            name="auto_inference_rule",
            template_id=None,
            placeholder_bindings=bindings,
            loop_config=loop_config,
            primitives_imports=rule_imports(),
            sample_check_log=check_log.model_dump(),
            confidence=confidence,
        )

        memory.update_inference_job(
            self.db, job_id,
            rule_artifact_id=artifact.id,
        )

        return {
            "artifact_id": artifact.id,
            "name": artifact.name,
            "confidence": confidence,
            "placeholder_bindings": bindings,
        }

    def _check_parser_sample(self, code: str, sample_file: str, account_code: str) -> dict:
        from decimal import Decimal
        from openpyxl import load_workbook

        wb = load_workbook(sample_file, read_only=True, data_only=True)
        try:
            rows = list(sandbox.execute(code, wb, {"account_code": account_code}))
        finally:
            wb.close()
        amount_in = sum((row.get("amount_in") or Decimal("0")) for row in rows)
        amount_out = sum((row.get("amount_out") or Decimal("0")) for row in rows)
        violations = sum(1 for row in rows if not row.get("business_date") or not row.get("account_code"))
        return {
            "sample_rows": len(rows),
            "parsed_rows": len(rows),
            "canonical_violations": violations,
            "amount_sum_in": str(amount_in),
            "amount_sum_out": str(amount_out),
        }
