"""主数据 API 路由

板块 / 法人 / 账户 / 别名 — 共 14 个端点 + 批量导入。
严格按 docs/30_contracts/23_api_contracts.md 实现。
"""
from typing import Optional

from fastapi import APIRouter, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from core.response import error, success
from database import get_db
from db.schemas import (
    AccountCreate, AccountUpdate, AliasCreate,
    DivisionCreate, DivisionUpdate, EntityCreate, EntityUpdate,
    InitialBalanceSet,
)
from services import master_data_service as svc

router = APIRouter()


# ──────────────────────────────────────────
# 板块
# ──────────────────────────────────────────

@router.get("/divisions")
def get_divisions(
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    items = svc.list_divisions(db, status=status)
    return success([{
        "id": d.id, "name": d.name, "sort_order": d.sort_order,
        "status": d.status, "created_at": d.created_at, "updated_at": d.updated_at,
    } for d in items])


@router.post("/divisions")
def create_division(body: DivisionCreate, db: Session = Depends(get_db)):
    try:
        obj = svc.create_division(db, body)
    except ValueError as e:
        return error(2001, str(e))
    return success({
        "id": obj.id, "name": obj.name, "sort_order": obj.sort_order,
        "status": obj.status,
    })


@router.put("/divisions/{division_id}")
def update_division(
    division_id: int, body: DivisionUpdate, db: Session = Depends(get_db),
):
    try:
        obj = svc.update_division(db, division_id, body)
    except ValueError as e:
        return error(2001, str(e))
    return success({
        "id": obj.id, "name": obj.name, "sort_order": obj.sort_order,
        "status": obj.status,
    })


# ──────────────────────────────────────────
# 法人
# ──────────────────────────────────────────

@router.get("/entities")
def get_entities(
    division_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    result = svc.list_entities(
        db, division_id=division_id, status=status, keyword=keyword,
        page=page, page_size=page_size,
    )
    return success(result.model_dump())


@router.post("/entities")
def create_entity(body: EntityCreate, db: Session = Depends(get_db)):
    try:
        obj = svc.create_entity(db, body)
    except ValueError as e:
        return error(2001, str(e))
    return success({
        "id": obj.id, "entity_code": obj.entity_code, "name": obj.name,
        "short_name": obj.short_name, "status": obj.status,
    })


@router.put("/entities/{entity_id}")
def update_entity(
    entity_id: int, body: EntityUpdate, db: Session = Depends(get_db),
):
    try:
        obj = svc.update_entity(db, entity_id, body)
    except ValueError as e:
        return error(2001, str(e))
    return success({
        "id": obj.id, "entity_code": obj.entity_code, "name": obj.name,
        "short_name": obj.short_name, "status": obj.status,
    })


# ──────────────────────────────────────────
# 账户
# ──────────────────────────────────────────

@router.get("/accounts")
def get_accounts(
    entity_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    account_type: Optional[str] = Query(None),
    instrument_type: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    result = svc.list_accounts(
        db, entity_id=entity_id, status=status, keyword=keyword,
        account_type=account_type, instrument_type=instrument_type,
        page=page, page_size=page_size,
    )
    return success(result.model_dump())


@router.get("/accounts/tree")
def get_accounts_tree(db: Session = Depends(get_db)):
    tree = svc.get_accounts_tree(db)
    return success([g.model_dump() for g in tree])


@router.post("/accounts")
def create_account(body: AccountCreate, db: Session = Depends(get_db)):
    try:
        obj = svc.create_account(db, body)
    except ValueError as e:
        return error(2001, str(e))
    return success({
        "id": obj.id, "account_code": obj.account_code,
        "account_alias": obj.account_alias, "status": obj.status,
    })


@router.put("/accounts/{account_id}")
def update_account(
    account_id: int, body: AccountUpdate, db: Session = Depends(get_db),
):
    try:
        obj = svc.update_account(db, account_id, body)
    except ValueError as e:
        return error(2001, str(e))
    return success({
        "id": obj.id, "account_code": obj.account_code,
        "account_alias": obj.account_alias, "status": obj.status,
    })


@router.post("/accounts/{account_id}/initial-balance")
def set_initial_balance(
    account_id: int, body: InitialBalanceSet, db: Session = Depends(get_db),
):
    try:
        obj = svc.set_initial_balance(db, account_id, body)
    except ValueError as e:
        return error(3001, str(e))
    return success({
        "id": obj.id, "initial_balance": float(obj.initial_balance),
        "balance_date": obj.balance_date.isoformat(),
    })


# ──────────────────────────────────────────
# 别名
# ──────────────────────────────────────────

@router.get("/accounts/{account_id}/aliases")
def get_aliases(account_id: int, db: Session = Depends(get_db)):
    items = svc.list_aliases(db, account_id)
    return success([{
        "id": a.id, "account_id": a.account_id,
        "alias_text": a.alias_text, "alias_type": a.alias_type,
        "created_at": a.created_at,
    } for a in items])


@router.post("/accounts/{account_id}/aliases")
def create_alias(
    account_id: int, body: AliasCreate, db: Session = Depends(get_db),
):
    try:
        obj = svc.create_alias(db, account_id, body)
    except ValueError as e:
        return error(2001, str(e))
    return success({
        "id": obj.id, "alias_text": obj.alias_text, "alias_type": obj.alias_type,
    })


@router.delete("/accounts/{account_id}/aliases/{alias_id}")
def delete_alias(
    account_id: int, alias_id: int, db: Session = Depends(get_db),
):
    try:
        svc.delete_alias(db, account_id, alias_id)
    except ValueError as e:
        return error(2001, str(e))
    return success(None, message="别名已删除")


# ──────────────────────────────────────────
# 批量导入
# ──────────────────────────────────────────

@router.get("/accounts/template")
def download_template():
    """下载账户批量导入模板"""
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "账户导入模板"

    # 表头
    headers = [
        ("板块名称", "必填，如：养护板块"),
        ("法人编码", "必填，如：E001"),
        ("法人名称", "必填，如：山西喜跃发实业发展有限公司"),
        ("法人简称", "必填，如：实业公司"),
        ("账户编码", "必填，如：A001"),
        ("账户别名", "必填，如：中行手工户"),
        ("开户银行", "选填"),
        ("开户行", "选填"),
        ("账号", "选填"),
        ("账户类型", "必填，如：银行账户/现金/票据/其他"),
        ("工具类型", "必填，如：银行存款/现金/票据/受限资金"),
        ("录入方式", "manual 或 bank_import"),
        ("币种", "如 CNY"),
        ("期初余额", "数字"),
        ("余额日期", "如 2026-03-01"),
        ("备注", "选填"),
    ]

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="E8F0E6", end_color="E8F0E6", fill_type="solid")
    hint_font = Font(size=9, color="888888")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, (name, hint) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

        hint_cell = ws.cell(row=2, column=ci, value=hint)
        hint_cell.font = hint_font
        hint_cell.border = thin_border

    # 示例行
    example = [
        "养护板块", "E001", "山西喜跃发实业发展有限公司", "实业公司",
        "A001", "中行手工户", "中国银行", "太原分行", "6217xxxx0001",
        "银行账户", "银行存款", "manual", "CNY", 200000, "2026-03-01", "无网银手工户",
    ]
    for ci, val in enumerate(example, 1):
        cell = ws.cell(row=3, column=ci, value=val)
        cell.border = thin_border

    # 列宽
    col_widths = [14, 10, 30, 12, 10, 14, 12, 12, 18, 12, 12, 14, 8, 12, 14, 20]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=account_import_template.xlsx"},
    )


@router.post("/accounts/import")
async def batch_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """批量导入账户（含自动创建板块和法人）"""
    if not file.filename:
        return error(1001, "缺少文件名")
    data = await file.read()
    if not data:
        return error(1001, "文件为空")
    try:
        result = svc.batch_import_accounts(db, data, file.filename)
    except ValueError as e:
        return error(3001, str(e))
    return success(result)
