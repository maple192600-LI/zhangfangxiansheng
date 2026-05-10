"""银行流水解析引擎

文件格式检测、表头读取、日期/金额规范化。
支持 xls (xlrd) / xlsx (openpyxl) / csv。
"""
import csv
import io
from datetime import date, datetime
from typing import Any, Dict, List, Optional

import openpyxl


def detect_format(filename: str) -> str:
    name = filename.lower()
    if name.endswith(".xlsx"):
        return "xlsx"
    if name.endswith(".xls"):
        return "xls"
    if name.endswith(".csv"):
        return "csv"
    return "unknown"


def read_file(filepath: str, fmt: str) -> List[List[str]]:
    """读取文件，返回原始二维字符串列表"""
    if fmt == "xlsx":
        return _read_xlsx(filepath)
    if fmt == "xls":
        return _read_xls(filepath)
    if fmt == "csv":
        return _read_csv(filepath)
    raise ValueError(f"不支持的文件格式: {fmt}")


def _read_xlsx(filepath: str) -> List[List[str]]:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c) if c is not None else "" for c in row])
    wb.close()
    return rows


def _read_xls(filepath: str) -> List[List[str]]:
    import xlrd
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)
    rows = []
    for r in range(ws.nrows):
        row = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            if cell.ctype == 3:  # XL_CELL_DATE
                val = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                row.append(val.strftime("%Y-%m-%d") if val else "")
            else:
                row.append(str(cell.value) if cell.value != "" else "")
        rows.append(row)
    return rows


def _read_csv(filepath: str) -> List[List[str]]:
    rows = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(row)
    return rows


def read_file_from_bytes(data: bytes, filename: str, fmt: str) -> List[List[str]]:
    """从内存字节读取文件"""
    if fmt == "csv":
        text = data.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        return [row for row in reader]

    # xlsx / xls 需要临时文件
    import tempfile, os
    suffix = ".xlsx" if fmt == "xlsx" else ".xls"
    fd, tmp = tempfile.mkstemp(suffix=suffix)
    try:
        os.write(fd, data)
        os.close(fd)
        return read_file(tmp, fmt)
    finally:
        os.unlink(tmp)


def detect_header_row(rows: List[List[str]], max_scan: int = 30) -> int:
    """自动检测表头所在行

    策略：
    1. 找到所有包含 3+ 个非空单元格的候选行（扫描范围扩大到 30 行）
    2. 对每个候选行评分：关键词命中数 × 非空列数
    3. 优先选择关键词命中多且列数多的行（真正的表头通常 6+ 列且命中多个关键词）
    4. 排除元数据行（含冒号分隔的 key:value 或 key  value 模式）
    """
    import re
    # 银行流水常见交易表头关键词（按权重分类）
    strong_keywords = {"交易日", "交易日期", "借方金额", "贷方金额", "摘要", "余额",
                       "对方户名", "对方名称", "收入金额", "支出金额", "用途",
                       "起息日", "流水号"}
    general_keywords = {"交易", "日期", "金额", "摘要", "对方", "余额", "收入", "支出",
                        "用途", "账号", "开户", "凭证", "币种", "备注", "类型",
                        "交易时间", "交易类型", "贷方", "借方", "发生额"}

    candidates = []
    for i, row in enumerate(rows[:max_scan]):
        non_empty = sum(1 for c in row if str(c).strip())
        if non_empty >= 3:
            candidates.append((i, row))

    if not candidates:
        return 0

    best_idx = candidates[0][0]
    best_score = 0

    for idx, row in candidates:
        cells = [str(c).strip() for c in row if str(c).strip()]
        non_empty = len(cells)
        all_text = " ".join(cells)

        # 计算关键词命中
        strong_hits = sum(1 for kw in strong_keywords if kw in all_text)
        general_hits = sum(1 for kw in general_keywords if kw in all_text)
        total_hits = strong_hits * 3 + general_hits

        # 排除元数据行：含冒号分隔的 key:value 对，或类似"接口版本 2.0 银行码 5456"的元数据模式
        has_colon_meta = bool(re.search(r'[：:]', all_text))
        # 检测 key-value 对模式：中文标签后面紧跟数字/日期/代码（非中文的值）
        kv_pairs = re.findall(r'[\u4e00-\u9fff]{2,6}\s+[\d.:/\-]+', all_text)
        # 如果含数字值的 KV 对 >= 2，且关键词命中少，这是元数据行
        if kv_pairs and len(kv_pairs) >= 2 and total_hits < 5:
            continue

        # 评分：关键词权重 × 非空列数（真正表头通常 6+ 列）
        # 元数据行通常 2-4 列有价值信息，表头通常 6+ 列
        score = total_hits * non_empty

        if has_colon_meta:
            score = score // 3  # 含冒号的大概率是元数据，降权

        if score > best_score:
            best_score = score
            best_idx = idx

    return best_idx


def normalize_date(val: str) -> Optional[str]:
    """将各种日期格式统一为 YYYY-MM-DD，支持带时间的格式"""
    if not val:
        return None
    val = val.strip()
    # 先尝试带时间的格式
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M",
                "%Y/%m/%d %H:%M", "%Y.%m.%d %H:%M:%S"):
        try:
            dt = datetime.strptime(val, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    # 纯日期格式
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            dt = datetime.strptime(val, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    # 尝试从混合格式中提取
    import re
    m = re.match(r"(\d{4})[./\-](\d{1,2})[./\-](\d{1,2})", val)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None


def normalize_amount(val: str) -> Optional[float]:
    """将金额字符串转为 float"""
    if not val:
        return None
    val = val.strip().replace(",", "").replace("，", "").replace(" ", "")
    if not val or val == "-":
        return None
    try:
        return round(float(val), 2)
    except ValueError:
        return None
