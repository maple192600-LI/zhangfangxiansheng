"""Excel 兼容加载工具。

让 .xls 文件也能用 openpyxl 接口处理：
- 读到 .xls → 用 xlrd 读所有 sheet/cell → 重写到内存 .xlsx → 返回 openpyxl Workbook
- 读到 .xlsx → 直接走 openpyxl

下游代码无需改 — 拿到的都是 openpyxl Workbook 对象。
"""
from __future__ import annotations

import io
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook as _load_xlsx


def _is_xls_path(path_or_bytes: Any) -> bool:
    if isinstance(path_or_bytes, (str, Path)):
        return str(path_or_bytes).lower().endswith(".xls")
    return False


def _xls_bytes_to_xlsx_workbook(data: bytes) -> Workbook:
    import xlrd  # 延迟 import — 仅 .xls 路径需要

    src = xlrd.open_workbook(file_contents=data)
    dst = Workbook()
    # 删掉默认 sheet
    dst.remove(dst.active)

    for sheet_idx in range(src.nsheets):
        src_sheet = src.sheet_by_index(sheet_idx)
        dst_sheet = dst.create_sheet(title=src_sheet.name[:31] or f"Sheet{sheet_idx + 1}")
        for r in range(src_sheet.nrows):
            for c in range(src_sheet.ncols):
                cell = src_sheet.cell(r, c)
                value: Any = cell.value
                # xlrd 日期 → datetime
                if cell.ctype == 3:  # XL_CELL_DATE
                    try:
                        value = xlrd.xldate_as_datetime(cell.value, src.datemode)
                    except Exception:
                        value = str(cell.value)
                elif cell.ctype == 2 and float(value).is_integer():
                    value = int(value)
                elif cell.ctype == 0:  # empty
                    value = None
                if value is not None:
                    dst_sheet.cell(row=r + 1, column=c + 1, value=value)
    return dst


def load_workbook_compat(path_or_bytes: Any, data_only: bool = True, read_only: bool = False) -> Workbook:
    """统一加载 .xls / .xlsx，返回 openpyxl Workbook。

    - path_or_bytes: 文件路径（str/Path）或 bytes
    - data_only / read_only: 透传给 openpyxl（仅 xlsx 路径有效）

    .xls 经 xlrd 读出后转成内存 .xlsx，故 data_only/read_only 对其无意义。
    """
    if isinstance(path_or_bytes, bytes):
        # 字节流 — 嗅探签名（.xls 是 OLE2: D0 CF 11 E0；.xlsx 是 ZIP: 50 4B 03 04）
        head = path_or_bytes[:4]
        if head[:2] == b"PK":
            return _load_xlsx(io.BytesIO(path_or_bytes), data_only=data_only, read_only=read_only)
        return _xls_bytes_to_xlsx_workbook(path_or_bytes)

    path_str = str(path_or_bytes)
    if path_str.lower().endswith(".xls"):
        with open(path_str, "rb") as f:
            return _xls_bytes_to_xlsx_workbook(f.read())
    return _load_xlsx(path_str, data_only=data_only, read_only=read_only)
