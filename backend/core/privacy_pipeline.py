"""Privacy masking helpers for Fund Agent skill calls."""
from __future__ import annotations

from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any, Mapping

PRIVACY_MODES = {"standard", "strict", "offline"}


class OfflineModeError(ValueError):
    """Raised when a skill call is blocked by offline privacy mode."""


def validate_privacy_mode(mode: str | None) -> str:
    value = (mode or "standard").strip().lower()
    if value not in PRIVACY_MODES:
        raise ValueError(f"未知隐私模式: {mode}")
    return value


def mask_for_llm(rows: Any, mode: str = "standard") -> Any:
    """Return the only data shape allowed to be shown to LLM-backed skills."""
    mode = validate_privacy_mode(mode)
    normalized = _normalize_rows(rows)
    if mode == "offline":
        raise OfflineModeError("已设为离线模式，不允许调用 Fund Agent skill")
    if mode == "strict":
        return {"headers": normalized["headers"], "rows": []}
    return {
        "headers": normalized["headers"],
        "rows": [
            {key: _mask_value(value) for key, value in row.items()}
            for row in normalized["rows"]
        ],
    }


def sample_preview_from_workbook(sample_file: str, *, limit: int = 10) -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(sample_file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    iterator = ws.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(iterator, ())]
    rows = []
    for raw in iterator:
        if len(rows) >= limit:
            break
        row = {
            headers[idx] or f"col_{idx + 1}": value
            for idx, value in enumerate(raw)
            if idx < len(headers)
        }
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    wb.close()
    return {"headers": headers, "rows": rows}


def _normalize_rows(rows: Any) -> dict[str, Any]:
    if rows is None:
        return {"headers": [], "rows": []}
    if isinstance(rows, Mapping):
        headers = list(rows.get("headers") or [])
        raw_rows = rows.get("rows") or []
        return {"headers": headers, "rows": [_row_to_mapping(row, headers) for row in raw_rows]}
    if isinstance(rows, list):
        if not rows:
            return {"headers": [], "rows": []}
        if all(isinstance(row, Mapping) for row in rows):
            headers = list(dict.fromkeys(key for row in rows for key in row.keys()))
            return {"headers": headers, "rows": [dict(row) for row in rows]}
        headers = [str(value) for value in rows[0]]
        return {"headers": headers, "rows": [_row_to_mapping(row, headers) for row in rows[1:]]}
    return {"headers": [], "rows": []}


def _row_to_mapping(row: Any, headers: list[str]) -> dict[str, Any]:
    if isinstance(row, Mapping):
        return dict(row)
    if isinstance(row, (list, tuple)):
        return {headers[idx] if idx < len(headers) else f"col_{idx + 1}": value for idx, value in enumerate(row)}
    return {"value": row}


def _mask_value(value: Any) -> Any:
    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, (int, float, Decimal)):
        return _mask_amount(value)
    text = str(value)
    if _digit_count(text) >= 10:
        return _mask_digits(text)
    if _looks_like_number(text):
        return _mask_amount(text)
    if not text:
        return text
    if any(ch.isdigit() for ch in text):
        return _mask_digits(text)
    return _mask_name(text)


def _mask_amount(value: Any) -> str:
    try:
        amount = Decimal(str(value).replace(",", ""))
    except (InvalidOperation, ValueError):
        return "***"
    rounded = (amount / Decimal("1000")).quantize(Decimal("1"), rounding=ROUND_HALF_UP) * Decimal("1000")
    return str(rounded)


def _looks_like_number(text: str) -> bool:
    stripped = text.replace(",", "").replace(" ", "")
    if stripped.startswith("-"):
        stripped = stripped[1:]
    return bool(stripped) and stripped.replace(".", "", 1).isdigit()


def _mask_digits(text: str) -> str:
    return "".join(ch if not ch.isdigit() else "*" for ch in text)


def _digit_count(text: str) -> int:
    return sum(1 for ch in text if ch.isdigit())


def _mask_name(text: str) -> str:
    first = text.strip()[:1]
    if not first:
        return ""
    return f"{first}***"
